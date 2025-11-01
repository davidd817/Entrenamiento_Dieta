# dieta_app.py
import streamlit as st
import pandas as pd
import random
from typing import List, Dict, Tuple
from scipy.optimize import minimize
import numpy as np
from datetime import datetime
import plotly.graph_objects as go
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from io import BytesIO





# ============================================================================
# CLASE PARA MARCA DE AGUA
# ============================================================================
class MarcaAguaCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
        
    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        for page_dict in self.pages:
            self.__dict__.update(page_dict)
            self.draw_watermark()
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
        
    def draw_watermark(self):
        self.saveState()
        self.setFont('Helvetica', 8)
        self.setFillColor(colors.Color(0.7, 0.7, 0.7, alpha=0.3))
        self.drawRightString(A4[0] - 1.5*cm, 1*cm, "David López - Plan Nutricional Personalizado")
        self.restoreState()

# ============================================================================
# FUNCIÓN PARA CALCULAR RACIONES EXACTAS
# ============================================================================

def calcular_racion_exacta(gramos_objetivo_macro, alimento, macro_tipo='proteina'):
    contenido_macro_100g = alimento[macro_tipo]
    
    if contenido_macro_100g == 0 or contenido_macro_100g < 0.1:
        return None
    
    gramos_alimento = (gramos_objetivo_macro / contenido_macro_100g) * 100
    
    if gramos_alimento > 750:
        return None
    
    unidades_exactas = gramos_alimento / alimento['gramos']
    
    # REDONDEAR SOLO A ENTEROS O MEDIOS
    if unidades_exactas < 0.25:
        return None
    elif unidades_exactas < 0.75:
        unidades = 0.5
    elif unidades_exactas < 1.25:
        unidades = 1
    elif unidades_exactas < 1.75:
        unidades = 1.5
    elif unidades_exactas < 2.25:
        unidades = 2
    elif unidades_exactas < 2.75:
        unidades = 2.5
    elif unidades_exactas < 3.25:
        unidades = 3
    elif unidades_exactas < 3.75:
        unidades = 3.5
    elif unidades_exactas < 4.25:
        unidades = 4
    elif unidades_exactas < 4.75:
        unidades = 4.5
    elif unidades_exactas < 5.25:
        unidades = 5
    else:
        unidades = round(unidades_exactas * 2) / 2
        if unidades > 10:
            return None
    
    gramos_alimento_redondeado = unidades * alimento['gramos']
    gramos_macro_real = (gramos_alimento_redondeado / 100) * contenido_macro_100g
    
    # CREAR DESCRIPCIÓN CON FORMATO CORRECTO
    unidad_base = alimento['unidad']
    
    # Pluralizar si es necesario (para cantidades > 1)
    if unidades > 1:
        # Reglas de pluralización en español
        if unidad_base.endswith('z'):
            unidad_plural = unidad_base[:-1] + 'ces'
        elif unidad_base.endswith('ón'):
            unidad_plural = unidad_base[:-2] + 'ones'
        elif unidad_base.endswith(('a', 'e', 'i', 'o', 'u')):
            unidad_plural = unidad_base + 's'
        elif unidad_base.endswith('s'):
            unidad_plural = unidad_base
        else:
            unidad_plural = unidad_base + 'es'
        
        # Formatear cantidad
        if unidades % 1 == 0:  # Es entero
            desc_unidad = f"{int(unidades)} {unidad_plural}"
        else:  # Es medio (x.5)
            desc_unidad = f"{unidades} {unidad_plural}".replace('.', ',')
    else:
        # Para cantidades < 2 (0.5, 1, 1.5)
        if unidades == 0.5:
            desc_unidad = f"0,5 {unidad_base}"
        elif unidades == 1:
            desc_unidad = f"1 {unidad_base}"  # AQUÍ: ahora sí incluye el "1"
        elif unidades == 1.5:
            desc_unidad = f"1,5 {unidad_base}"  # Y aquí también plural si lo prefieres
        else:
            desc_unidad = f"{unidades} {unidad_base}".replace('.', ',')
    
    return {
        'gramos_alimento': round(gramos_alimento_redondeado, 0),
        'unidades': unidades,
        'descripcion': desc_unidad,
        'gramos_macro_real': round(gramos_macro_real, 1)
    }

# ============================================================================
# FUNCIÓN PARA SELECCIONAR MEJORES OPCIONES PARA UNA COMIDA
# ============================================================================
def seleccionar_mejores_opciones_comida(cantidad_macro, diccionario_alimentos, macro_tipo, num_opciones_por_categoria=2):
    """
    Selecciona 2 alimentos de CADA subgrupo disponible (o los que sean viables)
    """
    # Primero, organizar por categoría
    opciones_por_categoria = {}
    
    for categoria, alimentos in diccionario_alimentos.items():
        opciones_por_categoria[categoria] = []
        
        for alimento in alimentos:
            racion = calcular_racion_exacta(cantidad_macro, alimento, macro_tipo)
            
            if racion is None:
                continue
            
            practicidad = 1 / (abs(racion['unidades'] - 1.5) + 0.5)
            
            opciones_por_categoria[categoria].append({
                'alimento': alimento,
                'racion': racion,
                'practicidad': practicidad,
                'categoria': categoria
            })
        
        # Ordenar por practicidad dentro de cada categoría
        opciones_por_categoria[categoria].sort(key=lambda x: x['practicidad'], reverse=True)
    
    # Ahora seleccionar 2 de cada categoría
    seleccionadas = []
    
    for categoria, opciones in opciones_por_categoria.items():
        # Tomar las 2 mejores de esta categoría (o las que haya disponibles)
        for opcion in opciones[:num_opciones_por_categoria]:
            seleccionadas.append(opcion)
    
    return seleccionadas




# ============================================================================
# FUNCIÓN PARA SELECCIONAR VARIEDAD DE ALIMENTOS 
# ============================================================================

def seleccionar_alimentos_variados(diccionario_alimentos, limite_por_categoria=3):
    """
    Selecciona 3 alimentos de CADA subgrupo para la guía general
    """
    alimentos_seleccionados = []
    
    for categoria, alimentos in diccionario_alimentos.items():
        contador = 0
        
        for alimento in alimentos:
            if contador >= limite_por_categoria:
                break
            
            alimentos_seleccionados.append(alimento)
            contador += 1
    
    return alimentos_seleccionados

# ============================================================================
# GENERADOR SIMPLIFICADO DE EJEMPLOS DE COMIDAS
# ============================================================================

class GeneradorComidasOptimas:
    """
    Genera ejemplos de comidas completas ajustadas a macros objetivo
    """
    
    def __init__(self):
        self.margen_error = 0.08  # 8% de margen permitido
    
    def generar_ejemplos_comida(self, comida_objetivo: Dict, num_generales=7, num_vegetarianas=3, num_veganas=2) -> List[Dict]:
        """
        Genera ejemplos de comidas ajustadas a los macros objetivo
        """
        ejemplos = []
        
        # Plantillas de combinaciones
        plantillas_generales = [
            {'prot': ['Carnes magras'], 'carb': ['Cereales integrales'], 'grasa': ['Aceites']},
            {'prot': ['Pescados blancos'], 'carb': ['Tubérculos'], 'grasa': ['Aceites']},
            {'prot': ['Pescados azules'], 'carb': ['Cereales integrales'], 'grasa': []},
            {'prot': ['Huevos'], 'carb': ['Pan'], 'grasa': ['Aguacates y aceitunas']},
            {'prot': ['Carnes magras'], 'carb': ['Frutas'], 'grasa': ['Frutos secos']},
            {'prot': ['Mariscos'], 'carb': ['Cereales refinados'], 'grasa': ['Aceites']},
            {'prot': ['Lácteos'], 'carb': ['Frutas'], 'grasa': ['Frutos secos']},
        ]
        
        plantillas_vegetarianas = [
            {'prot': ['Legumbres'], 'carb': ['Cereales integrales'], 'grasa': ['Semillas']},
            {'prot': ['Huevos'], 'carb': ['Tubérculos'], 'grasa': ['Aceites']},
            {'prot': ['Lácteos'], 'carb': ['Cereales integrales'], 'grasa': ['Frutos secos']},
        ]
        
        plantillas_veganas = [
            {'prot': ['Legumbres'], 'carb': ['Cereales integrales'], 'grasa': ['Semillas']},
            {'prot': ['Derivados de soja'], 'carb': ['Tubérculos'], 'grasa': ['Aguacates y aceitunas']},
        ]
        
        # Generar ejemplos generales
        for i, plantilla in enumerate(plantillas_generales[:num_generales]):
            ejemplo = self._construir_comida_ajustada(comida_objetivo, plantilla, f'General {i+1}', i)
            if ejemplo:
                ejemplos.append(ejemplo)
        
        # Generar ejemplos vegetarianos
        for i, plantilla in enumerate(plantillas_vegetarianas[:num_vegetarianas]):
            ejemplo = self._construir_comida_ajustada(comida_objetivo, plantilla, f'Vegetariana {i+1}', i)
            if ejemplo:
                ejemplos.append(ejemplo)
        
        # Generar ejemplos veganos
        for i, plantilla in enumerate(plantillas_veganas[:num_veganas]):
            ejemplo = self._construir_comida_ajustada(comida_objetivo, plantilla, f'Vegana {i+1}', i)
            if ejemplo:
                ejemplos.append(ejemplo)
        
        return ejemplos
    
    def _construir_comida_ajustada(self, objetivo: Dict, plantilla: Dict, tipo: str, variacion: int) -> Dict:
        """
        Construye una comida ajustando las cantidades para cumplir los macros objetivo
        """
        # Seleccionar alimentos
        alimento_prot = self._seleccionar_alimento(plantilla['prot'], variacion)
        alimento_carb = self._seleccionar_alimento(plantilla['carb'], variacion)
        alimento_grasa = self._seleccionar_alimento(plantilla['grasa'], variacion) if plantilla['grasa'] else None
        
        if not alimento_prot or not alimento_carb:
            return None
        
        # AJUSTAR CANTIDADES PARA CUMPLIR MACROS
        cantidades = self._calcular_cantidades_exactas(
            alimento_prot,
            alimento_carb,
            alimento_grasa,
            objetivo['proteina'],
            objetivo['carbos'],
            objetivo['grasas']
        )
        
        if not cantidades:
            return None
        
        # Construir resultado
        alimentos_finales = []
        macros_totales = {'proteina': 0, 'carbos': 0, 'grasas': 0}
        
        # Añadir proteína
        if cantidades['prot_gramos'] > 0:
            factor_prot = cantidades['prot_gramos'] / alimento_prot['gramos']
            macros_prot = {
                'proteina': alimento_prot['proteina'] * factor_prot,
                'carbos': alimento_prot['carbos'] * factor_prot,
                'grasas': alimento_prot['grasas'] * factor_prot
            }
            
            alimentos_finales.append({
                'nombre': alimento_prot['nombre'],
                'gramos': round(cantidades['prot_gramos'], 0),
                'equivalencia': self._formatear_equivalencia(
                    cantidades['prot_gramos'] / alimento_prot['gramos'],
                    alimento_prot['unidad']
                )
            })
            
            for k in macros_totales:
                macros_totales[k] += macros_prot[k]
        
        # Añadir carbohidrato
        if cantidades['carb_gramos'] > 0:
            factor_carb = cantidades['carb_gramos'] / alimento_carb['gramos']
            macros_carb = {
                'proteina': alimento_carb['proteina'] * factor_carb,
                'carbos': alimento_carb['carbos'] * factor_carb,
                'grasas': alimento_carb['grasas'] * factor_carb
            }
            
            alimentos_finales.append({
                'nombre': alimento_carb['nombre'],
                'gramos': round(cantidades['carb_gramos'], 0),
                'equivalencia': self._formatear_equivalencia(
                    cantidades['carb_gramos'] / alimento_carb['gramos'],
                    alimento_carb['unidad']
                )
            })
            
            for k in macros_totales:
                macros_totales[k] += macros_carb[k]
        
        # Añadir grasa si existe
        if alimento_grasa and cantidades['grasa_gramos'] > 0:
            factor_grasa = cantidades['grasa_gramos'] / alimento_grasa['gramos']
            macros_grasa = {
                'proteina': alimento_grasa['proteina'] * factor_grasa,
                'carbos': alimento_grasa['carbos'] * factor_grasa,
                'grasas': alimento_grasa['grasas'] * factor_grasa
            }
            
            alimentos_finales.append({
                'nombre': alimento_grasa['nombre'],
                'gramos': round(cantidades['grasa_gramos'], 0),
                'equivalencia': self._formatear_equivalencia(
                    cantidades['grasa_gramos'] / alimento_grasa['gramos'],
                    alimento_grasa['unidad']
                )
            })
            
            for k in macros_totales:
                macros_totales[k] += macros_grasa[k]
        
        # Calcular calorías
        calorias = (macros_totales['proteina'] * 4 + 
                   macros_totales['carbos'] * 4 + 
                   macros_totales['grasas'] * 9)
        
        return {
            'tipo': tipo,
            'alimentos': alimentos_finales,
            'macros_totales': {
                'proteina': round(macros_totales['proteina'], 1),
                'carbos': round(macros_totales['carbos'], 1),
                'grasas': round(macros_totales['grasas'], 1),
                'calorias': round(calorias, 0)
            }
        }
    
    def _calcular_cantidades_exactas(self, alim_prot, alim_carb, alim_grasa, prot_obj, carb_obj, grasa_obj):
        """
        Calcula las cantidades exactas de cada alimento para cumplir los macros
        Usa un sistema de ecuaciones simplificado
        """
        # Si no hay alimento de grasa, ajustar solo proteína y carbohidrato
        if not alim_grasa:
            # Sistema 2x2: ajustar proteína y carbos con 2 alimentos
            # Alimento proteico aporta: p1*x proteína, c1*x carbos, g1*x grasas
            # Alimento carbohidrato aporta: p2*y proteína, c2*y carbos, g2*y grasas
            
            # Objetivo: p1*x + p2*y ≈ prot_obj
            #          c1*x + c2*y ≈ carb_obj
            
            # Estrategia: priorizar el macro principal de cada alimento
            # x = gramos de alimento proteico
            # y = gramos de alimento carbohidrato
            
            # Estimar x desde proteína (asumiendo que el alimento carb aporta poca proteína)
            x = (prot_obj / (alim_prot['proteina'] / alim_prot['gramos'])) if alim_prot['proteina'] > 0 else 100
            
            # Calcular carbos que faltan
            carbos_de_prot = (x / alim_prot['gramos']) * alim_prot['carbos']
            carbos_faltantes = carb_obj - carbos_de_prot
            
            # Estimar y desde carbos faltantes
            y = (carbos_faltantes / (alim_carb['carbos'] / alim_carb['gramos'])) if alim_carb['carbos'] > 0 else 100
            
            # Validar y ajustar si es necesario
            if y < 0:
                y = 50  # Mínimo razonable
            
            return {
                'prot_gramos': max(10, min(500, x)),
                'carb_gramos': max(10, min(500, y)),
                'grasa_gramos': 0
            }
        else:
            # Sistema 3x3: ajustar los tres macros con 3 alimentos
            # Estrategia similar pero más compleja
            
            # Estimar cantidad de alimento proteico
            x = (prot_obj / (alim_prot['proteina'] / alim_prot['gramos'])) if alim_prot['proteina'] > 0 else 100
            
            # Estimar cantidad de alimento carbohidrato
            carbos_de_prot = (x / alim_prot['gramos']) * alim_prot['carbos']
            carbos_faltantes = carb_obj - carbos_de_prot
            y = (carbos_faltantes / (alim_carb['carbos'] / alim_carb['gramos'])) if alim_carb['carbos'] > 0 else 100
            
            # Estimar cantidad de alimento graso
            grasas_de_prot = (x / alim_prot['gramos']) * alim_prot['grasas']
            grasas_de_carb = (y / alim_carb['gramos']) * alim_carb['grasas']
            grasas_faltantes = grasa_obj - grasas_de_prot - grasas_de_carb
            z = (grasas_faltantes / (alim_grasa['grasas'] / alim_grasa['gramos'])) if alim_grasa['grasas'] > 0 else 10
            
            # Validar
            if z < 0:
                z = 0
            
            return {
                'prot_gramos': max(10, min(500, x)),
                'carb_gramos': max(10, min(500, y)),
                'grasa_gramos': max(0, min(100, z))
            }
    
    def _seleccionar_alimento(self, categorias: List[str], variacion: int):
        """Selecciona un alimento de las categorías disponibles"""
        if not categorias:
            return None
        
        categoria = categorias[variacion % len(categorias)]
        
        # Buscar en la base de datos correcta
        if categoria in ALIMENTOS_PROTEINAS:
            alimentos = ALIMENTOS_PROTEINAS[categoria]
        elif categoria in ALIMENTOS_CARBOHIDRATOS:
            alimentos = ALIMENTOS_CARBOHIDRATOS[categoria]
        elif categoria in ALIMENTOS_GRASAS:
            alimentos = ALIMENTOS_GRASAS[categoria]
        else:
            return None
        
        if not alimentos:
            return None
        
        return alimentos[variacion % len(alimentos)]
    
    def _formatear_equivalencia(self, unidades: float, unidad_base: str) -> str:
        """Formatea la equivalencia (mismo código que antes)"""
        if unidades < 0.75:
            unidades_red = 0.5
        elif unidades < 1.25:
            unidades_red = 1
        elif unidades < 1.75:
            unidades_red = 1.5
        else:
            unidades_red = round(unidades * 2) / 2
        
        if unidades_red >= 2:
            if unidad_base.endswith('z'):
                unidad_plural = unidad_base[:-1] + 'ces'
            elif unidad_base.endswith('ón'):
                unidad_plural = unidad_base[:-2] + 'ones'
            elif unidad_base.endswith(('a', 'e', 'i', 'o', 'u')):
                unidad_plural = unidad_base + 's'
            else:
                unidad_plural = unidad_base + 'es'
            
            if unidades_red % 1 == 0:
                return f"{int(unidades_red)} {unidad_plural}"
            else:
                return f"{unidades_red} {unidad_plural}".replace('.', ',')
        else:
            if unidades_red == 0.5:
                return f"0,5 {unidad_base}"
            elif unidades_red == 1:
                return f"1 {unidad_base}"
            elif unidades_red == 1.5:
                return f"1,5 {unidad_base}"
        
        return f"{unidades_red} {unidad_base}".replace('.', ',')



# ============================================================================
# FUNCIÓN SIMPLIFICADA PARA PDF
# ============================================================================

def agregar_ejemplos_comidas_pdf(elements, comida, styles=None):
    """Añade ejemplos de comidas simplificados al PDF"""
    
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    
    base_styles = getSampleStyleSheet()
    
    title_style_local = ParagraphStyle(
        'TitleLocal',
        parent=base_styles['Heading1'],
        fontSize=12,
        fontName='Helvetica-Bold',
        spaceAfter=10
    )
    
    heading_style_local = ParagraphStyle(
        'HeadingLocal',
        fontSize=10,
        fontName='Helvetica-Bold',
        spaceAfter=4
    )
    
    try:
        generador = GeneradorComidasOptimas()
        ejemplos = generador.generar_ejemplos_comida(comida)
        
        if not ejemplos:
            elements.append(PageBreak())
            return
        
        elements.append(Spacer(1, 0.8*cm))
        elements.append(Paragraph("EJEMPLOS DE COMIDAS COMPLETAS", title_style_local))
        elements.append(Spacer(1, 0.3*cm))
        
        for idx, ejemplo in enumerate(ejemplos, 1):
            # Título simple
            elements.append(Paragraph(f"{ejemplo['tipo']}", heading_style_local))
            
            # Tabla simple: solo alimento, gramos y equivalencia
            data_ejemplo = [['Alimento', 'Cantidad', 'Equivalencia']]
            
            for alim in ejemplo['alimentos']:
                data_ejemplo.append([
                    alim['nombre'],
                    f"{alim['gramos']:.0f}g",
                    alim['equivalencia']
                ])
            
            # Fila de totales con macros
            data_ejemplo.append([
                'TOTAL',
                f"{ejemplo['macros_totales']['calorias']:.0f} kcal",
                f"P: {ejemplo['macros_totales']['proteina']:.0f}g | "
                f"C: {ejemplo['macros_totales']['carbos']:.0f}g | "
                f"G: {ejemplo['macros_totales']['grasas']:.0f}g"
            ])
            
            tabla_ejemplo = Table(data_ejemplo, colWidths=[6*cm, 3*cm, 5*cm])
            tabla_ejemplo.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.5, 0.5, 0.5)),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 8),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.Color(0.97, 0.97, 0.97)]),
            ]))
            
            elements.append(tabla_ejemplo)
            elements.append(Spacer(1, 0.3*cm))
        
        elements.append(PageBreak())
        
    except Exception as e:
        print(f"❌ ERROR: {e}")
        import traceback
        traceback.print_exc()
        elements.append(PageBreak())


# ============================================================================
# FUNCIÓN GENERAR PDF DE DIETA
# ============================================================================
def generar_pdf_dieta(plan):
    """Genera PDF profesional completo con validación de raciones"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=A4,
        rightMargin=1.5*cm, 
        leftMargin=1.5*cm, 
        topMargin=1.5*cm, 
        bottomMargin=1.5*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ESTILOS
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=22,
        textColor=colors.Color(0.1, 0.1, 0.1),
        spaceAfter=15,
        spaceBefore=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=14,
        textColor=colors.Color(0.2, 0.2, 0.2),
        spaceAfter=10,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        fontSize=11,
        textColor=colors.Color(0.2, 0.2, 0.2),
        spaceAfter=6,
        spaceBefore=8,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=4,
        fontName='Helvetica',
        leading=12,
        alignment=TA_JUSTIFY
    )
    
    cliente_nombre = plan['cliente'] if plan['cliente'] else "Cliente"
    sexo = plan.get('sexo', 'Hombre')
    
    # PORTADA
    elements.append(Spacer(1, 3*cm))
    elements.append(Paragraph("PLAN NUTRICIONAL", title_style))
    elements.append(Paragraph("PERSONALIZADO", subtitle_style))
    elements.append(Spacer(1, 2*cm))
    
    info_portada = f"""
    <b>Cliente:</b> {cliente_nombre}<br/>
    <b>Objetivo:</b> {plan['objetivo']}<br/>
    <b>Calorías diarias:</b> {plan['calorias']:.0f} kcal<br/>
    <b>Fecha:</b> {plan['fecha']}
    """
    
    elements.append(Paragraph(info_portada, info_style))
    elements.append(Spacer(1, 3*cm))
    elements.append(Paragraph("David López Diego", 
                             ParagraphStyle('Footer', parent=info_style, alignment=TA_CENTER, 
                                          fontSize=9, fontName='Helvetica-Bold')))
    
    elements.append(PageBreak())
    
    # PÁGINAS POR COMIDA
    for i, comida in enumerate(plan['comidas']):
        elements.append(Paragraph(f"{comida['nombre'].upper()}", title_style))
        elements.append(Spacer(1, 0.5*cm))
        
        calorias_comida = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
        
        data_comida = [
            ['MACRONUTRIENTE', 'CANTIDAD'],
            ['Proteína', f"{comida['proteina']:.0f}g"],
            ['Carbohidratos', f"{comida['carbos']:.0f}g"],
            ['Grasas', f"{comida['grasas']:.0f}g"],
            ['CALORÍAS TOTALES', f"{calorias_comida:.0f} kcal"]
        ]
        
        tabla_comida = Table(data_comida, colWidths=[7*cm, 5*cm])
        tabla_comida.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.2, 0.2)),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
            ('BACKGROUND', (0, 1), (-1, -2), colors.white),
            ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
            ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.black),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
        ]))
        
        elements.append(tabla_comida)
        elements.append(Spacer(1, 0.7*cm))
        
        # PROTEÍNA - VALIDACIÓN AÑADIDA
        if comida['proteina'] > 5:
            elements.append(Paragraph(f"Para {comida['proteina']:.0f}g de PROTEÍNA, elige UNA opción:", 
                                     ParagraphStyle('SubHead', parent=heading_style, fontSize=10, 
                                                   textColor=colors.Color(0.6, 0, 0))))
            
            opciones_prot = seleccionar_mejores_opciones_comida(
                comida['proteina'], 
                ALIMENTOS_PROTEINAS, 
                'proteina', 
                num_opciones_por_categoria=2
            )
            
            if opciones_prot:  # VALIDACIÓN: Solo si hay opciones
                data_prot = [['Alimento', 'Cantidad', 'Equivalencia']]
                for opcion in opciones_prot:
                    if opcion and opcion.get('racion'):  # VALIDACIÓN adicional
                        data_prot.append([
                            opcion['alimento']['nombre'],
                            f"{opcion['racion']['gramos_alimento']:.0f}g",
                            opcion['racion']['descripcion']
                        ])
                
                if len(data_prot) > 1:  # Solo crear tabla si hay datos
                    tabla_prot = Table(data_prot, colWidths=[6*cm, 3*cm, 4*cm])
                    tabla_prot.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.7, 0.7)),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    
                    elements.append(tabla_prot)
                    elements.append(Spacer(1, 0.5*cm))
        
        # CARBOHIDRATOS - VALIDACIÓN AÑADIDA
        if comida['carbos'] > 5:
            elements.append(Paragraph(f"Para {comida['carbos']:.0f}g de CARBOHIDRATOS, elige UNA opción:", 
                                     ParagraphStyle('SubHead', parent=heading_style, fontSize=10,
                                                   textColor=colors.Color(0, 0, 0.6))))
            
            opciones_carbs = seleccionar_mejores_opciones_comida(
                comida['carbos'], 
                ALIMENTOS_CARBOHIDRATOS, 
                'carbos', 
                num_opciones_por_categoria=2
            )
            
            if opciones_carbs:
                data_carbs = [['Alimento', 'Cantidad', 'Equivalencia']]
                for opcion in opciones_carbs:
                    if opcion and opcion.get('racion'):
                        data_carbs.append([
                            opcion['alimento']['nombre'],
                            f"{opcion['racion']['gramos_alimento']:.0f}g",
                            opcion['racion']['descripcion']
                        ])
                
                if len(data_carbs) > 1:
                    tabla_carbs = Table(data_carbs, colWidths=[6*cm, 3*cm, 4*cm])
                    tabla_carbs.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.9, 1)),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    
                    elements.append(tabla_carbs)
                    elements.append(Spacer(1, 0.5*cm))
        
        # GRASAS - VALIDACIÓN AÑADIDA
        if comida['grasas'] > 3:
            elements.append(Paragraph(f"Para {comida['grasas']:.0f}g de GRASAS, elige UNA opción:", 
                                     ParagraphStyle('SubHead', parent=heading_style, fontSize=10,
                                                   textColor=colors.Color(0.6, 0.5, 0))))
            
            opciones_grasas = seleccionar_mejores_opciones_comida(
                comida['grasas'], 
                ALIMENTOS_GRASAS, 
                'grasas', 
                num_opciones_por_categoria=2
            )
            
            if opciones_grasas:
                data_grasas = [['Alimento', 'Cantidad', 'Equivalencia']]
                for opcion in opciones_grasas:
                    if opcion and opcion.get('racion'):
                        data_grasas.append([
                            opcion['alimento']['nombre'],
                            f"{opcion['racion']['gramos_alimento']:.0f}g",
                            opcion['racion']['descripcion']
                        ])
                
                if len(data_grasas) > 1:
                    tabla_grasas = Table(data_grasas, colWidths=[6*cm, 3*cm, 4*cm])
                    tabla_grasas.setStyle(TableStyle([
                        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(1, 0.95, 0.7)),
                        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                        ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, -1), 8),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                        ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
                        ('TOPPADDING', (0, 0), (-1, -1), 4),
                    ]))
                    
                    elements.append(tabla_grasas)
                    elements.append(Spacer(1, 0.4*cm))
        
        nota_verduras = """
        <b>VERDURAS:</b> Añade la cantidad que desees de verduras sin contabilizar calorías 
        (ejemplos: lechuga, espinacas, brócoli, tomate, pepino, calabacín, pimiento).
        """
        elements.append(Paragraph(nota_verduras, info_style))
        elements.append(Spacer(1, 0.5*cm))
        
        try:
            agregar_ejemplos_comidas_pdf(elements, comida, styles)
        except Exception as e:
            # Si falla, al menos registrar el error y continuar
            print(f"Error generando ejemplos para {comida['nombre']}: {e}")
            elements.append(PageBreak())
    
    # COMIDA PRE-ENTRENAMIENTO
    elements.append(Paragraph("COMIDA PRE-ENTRENAMIENTO", title_style))
    elements.append(Paragraph("(Solo días de entrenamiento - 30-60 min antes)", subtitle_style))
    elements.append(Spacer(1, 0.5*cm))
    
    if sexo == "Mujer":
        proteina_pre = 15
        carbos_pre = 25
    else:
        proteina_pre = 22
        carbos_pre = 35
    
    calorias_pre = (proteina_pre * 4) + (carbos_pre * 4)
    
    data_pre = [
        ['MACRONUTRIENTE', 'CANTIDAD'],
        ['Proteína', f"{proteina_pre}g ({proteina_pre-5}-{proteina_pre+5}g)"],
        ['Carbohidratos', f"{carbos_pre}g ({carbos_pre-5}-{carbos_pre+5}g)"],
        ['Grasas', 'Mínimas (evitar)'],
        ['CALORÍAS', f"~{calorias_pre:.0f} kcal"]
    ]
    
    tabla_pre = Table(data_pre, colWidths=[7*cm, 5*cm])
    tabla_pre.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.3, 0.6, 0.3)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 10),
        ('BACKGROUND', (0, 1), (-1, -2), colors.white),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.9, 0.9, 0.9)),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
    ]))
    
    elements.append(tabla_pre)
    elements.append(Spacer(1, 0.7*cm))
    
    # PROTEÍNA PRE
    elements.append(Paragraph(f"Para {proteina_pre}g de PROTEÍNA:", 
                             ParagraphStyle('SubHead', parent=heading_style, fontSize=10, 
                                           textColor=colors.Color(0.6, 0, 0))))
    
    alimentos_pre_prot = [
        {'nombre': 'Proteína whey', 'proteina': 80, 'carbos': 5, 'grasas': 3, 'unidad': 'cacito', 'gramos': 30}
    ]
    
    data_prot_pre = [['Alimento', 'Cantidad', 'Equivalencia']]
    for alim in alimentos_pre_prot:
        racion = calcular_racion_exacta(proteina_pre, alim, 'proteina')
        if racion:  # VALIDACIÓN
            data_prot_pre.append([
                alim['nombre'],
                f"{racion['gramos_alimento']:.0f}g",
                racion['descripcion']
            ])
    
    if len(data_prot_pre) > 1:
        tabla_prot_pre = Table(data_prot_pre, colWidths=[6*cm, 3*cm, 4*cm])
        tabla_prot_pre.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.9, 0.7, 0.7)),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        elements.append(tabla_prot_pre)
        elements.append(Spacer(1, 0.5*cm))
    
    # CARBOS PRE
    elements.append(Paragraph(f"Para {carbos_pre}g de CARBOHIDRATOS:", 
                             ParagraphStyle('SubHead', parent=heading_style, fontSize=10,
                                           textColor=colors.Color(0, 0, 0.6))))
    
    alimentos_pre_carbs = [
        {'nombre': 'Plátano', 'proteina': 1.1, 'carbos': 23, 'grasas': 0.3, 'unidad': 'plátano', 'gramos': 120},
        {'nombre': 'Copos de maíz sin azúcar', 'proteina': 8, 'carbos': 80, 'grasas': 1, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Dátil', 'proteina': 2, 'carbos': 75, 'grasas': 0.2, 'unidad': 'dátil', 'gramos': 12},
        {'nombre': 'Tortitas de arroz', 'proteina': 8, 'carbos': 81, 'grasas': 3, 'unidad': 'tortita', 'gramos': 9},
        {'nombre': 'Miel', 'proteina': 0.3, 'carbos': 82, 'grasas': 0, 'unidad': 'cucharada', 'gramos': 21},
    ]
    
    data_carbs_pre = [['Alimento', 'Cantidad', 'Equivalencia']]
    for alim in alimentos_pre_carbs:
        racion = calcular_racion_exacta(carbos_pre, alim, 'carbos')
        if racion:  # VALIDACIÓN
            data_carbs_pre.append([
                alim['nombre'],
                f"{racion['gramos_alimento']:.0f}g",
                racion['descripcion']
            ])
    
    if len(data_carbs_pre) > 1:
        tabla_carbs_pre = Table(data_carbs_pre, colWidths=[6*cm, 3*cm, 4*cm])
        tabla_carbs_pre.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.8, 0.9, 1)),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 4),
            ('TOPPADDING', (0, 0), (-1, -1), 4),
        ]))
        
        elements.append(tabla_carbs_pre)
        elements.append(Spacer(1, 0.5*cm))
    
    nota_pre = """
    <b>IMPORTANTE:</b> Esta comida solo se consume los días de entrenamiento, 30-60 minutos antes. 
    Combina 1 opción de proteína + 1 opción de carbohidratos. Evita grasas para facilitar digestión.
    """
    elements.append(Paragraph(nota_pre, info_style))
    
    elements.append(PageBreak())
    
    # RESUMEN DIARIO
    elements.append(Paragraph("RESUMEN DIARIO TOTAL", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    data_resumen = [['Comida', 'Proteína', 'Carbos', 'Grasas', 'Calorías']]
    
    for comida in plan['comidas']:
        cals = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
        data_resumen.append([
            comida['nombre'],
            f"{comida['proteina']:.0f}g",
            f"{comida['carbos']:.0f}g",
            f"{comida['grasas']:.0f}g",
            f"{cals:.0f} kcal"
        ])
    
    data_resumen.append([
        'TOTAL DIARIO',
        f"{plan['macros']['proteina']:.0f}g",
        f"{plan['macros']['carbos']:.0f}g",
        f"{plan['macros']['grasas']:.0f}g",
        f"{plan['calorias']:.0f} kcal"
    ])
    
    tabla_resumen = Table(data_resumen, colWidths=[4*cm, 3*cm, 3*cm, 3*cm, 3*cm])
    tabla_resumen.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.2, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 9),
        ('BACKGROUND', (0, -1), (-1, -1), colors.Color(0.8, 0.8, 0.8)),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('FONTSIZE', (0, 1), (-1, -2), 8),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(tabla_resumen)
    elements.append(PageBreak())
    
    # GUÍA GENERAL
    elements.append(Paragraph("GUÍA GENERAL DE ALIMENTOS", title_style))
    elements.append(Spacer(1, 0.3*cm))
    
    intro_guia = """
    <b>CÓMO USAR ESTA GUÍA:</b> Usa esta sección como referencia rápida para variar tus comidas 
    y combinar diferentes alimentos.
    """
    elements.append(Paragraph(intro_guia, info_style))
    elements.append(Spacer(1, 0.5*cm))
    
    # PROTEÍNAS - GUÍA
    elements.append(Paragraph("FUENTES DE PROTEÍNA", subtitle_style))
    elements.append(Paragraph("Raciones de 20g, 40g y 50g de proteína", heading_style))
    elements.append(Spacer(1, 0.3*cm))
    
    alimentos_prot_guia = seleccionar_alimentos_variados(ALIMENTOS_PROTEINAS, limite_por_categoria=3)
    
    for racion_objetivo in [20, 40, 50]:
        elements.append(Paragraph(f"RACIÓN DE {racion_objetivo}g DE PROTEÍNA", 
                                 ParagraphStyle('RacionHead', parent=heading_style, fontSize=10)))
        
        data_prot = [['Alimento', 'Cantidad', 'Equivalencia']]
        
        for alim in alimentos_prot_guia:
            racion = calcular_racion_exacta(racion_objetivo, alim, 'proteina')
            if racion:  # VALIDACIÓN
                data_prot.append([
                    alim['nombre'],
                    f"{racion['gramos_alimento']:.0f}g",
                    racion['descripcion']
                ])
        
        if len(data_prot) > 1:
            tabla_prot = Table(data_prot, colWidths=[6*cm, 3*cm, 4*cm])
            tabla_prot.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.7, 0.7, 0.7)),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(tabla_prot)
            elements.append(Spacer(1, 0.3*cm))
    
    elements.append(PageBreak())
    
    # CARBOHIDRATOS - GUÍA
    elements.append(Paragraph("FUENTES DE CARBOHIDRATOS", subtitle_style))
    elements.append(Paragraph("Raciones de 20g, 50g y 100g de carbohidratos", heading_style))
    elements.append(Spacer(1, 0.3*cm))
    
    alimentos_carbs_guia = seleccionar_alimentos_variados(ALIMENTOS_CARBOHIDRATOS, limite_por_categoria=3)
    
    for racion_objetivo in [20, 50, 100]:
        elements.append(Paragraph(f"RACIÓN DE {racion_objetivo}g DE CARBOHIDRATOS", 
                                 ParagraphStyle('RacionHead', parent=heading_style, fontSize=10)))
        
        data_carbs = [['Alimento', 'Cantidad', 'Equivalencia']]
        
        for alim in alimentos_carbs_guia:
            racion = calcular_racion_exacta(racion_objetivo, alim, 'carbos')
            if racion:  # VALIDACIÓN
                data_carbs.append([
                    alim['nombre'],
                    f"{racion['gramos_alimento']:.0f}g",
                    racion['descripcion']
                ])
        
        if len(data_carbs) > 1:
            tabla_carbs = Table(data_carbs, colWidths=[6*cm, 3*cm, 4*cm])
            tabla_carbs.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.7, 0.7, 0.7)),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(tabla_carbs)
            elements.append(Spacer(1, 0.3*cm))
    
    elements.append(PageBreak())
    
    # GRASAS - GUÍA
    elements.append(Paragraph("FUENTES DE GRASAS SALUDABLES", subtitle_style))
    elements.append(Paragraph("Raciones de 5g, 10g y 20g de grasas", heading_style))
    elements.append(Spacer(1, 0.3*cm))
    
    alimentos_grasas_guia = seleccionar_alimentos_variados(ALIMENTOS_GRASAS, limite_por_categoria=3)
    
    for racion_objetivo in [5, 10, 20]:
        elements.append(Paragraph(f"RACIÓN DE {racion_objetivo}g DE GRASAS", 
                                 ParagraphStyle('RacionHead', parent=heading_style, fontSize=10)))
        
        data_grasas = [['Alimento', 'Cantidad', 'Equivalencia']]
        
        for alim in alimentos_grasas_guia:
            racion = calcular_racion_exacta(racion_objetivo, alim, 'grasas')
            if racion:  # VALIDACIÓN
                data_grasas.append([
                    alim['nombre'],
                    f"{racion['gramos_alimento']:.0f}g",
                    racion['descripcion']
                ])
        
        if len(data_grasas) > 1:
            tabla_grasas = Table(data_grasas, colWidths=[6*cm, 3*cm, 4*cm])
            tabla_grasas.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.7, 0.7, 0.7)),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('ALIGN', (1, 0), (-1, -1), 'CENTER'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, -1), 7),
                ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
                ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
                ('BOTTOMPADDING', (0, 0), (-1, -1), 3),
                ('TOPPADDING', (0, 0), (-1, -1), 3),
            ]))
            
            elements.append(tabla_grasas)
            elements.append(Spacer(1, 0.3*cm))
    
    elements.append(PageBreak())
    
    # CONSEJOS
    elements.append(Paragraph("CONSEJOS PRÁCTICOS PARA EL ÉXITO", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    consejos = """
    <b>1. Flexibilidad nutricional:</b> No necesitas ser exacto al gramo. Aproximaciones del 5-10% 
    son perfectamente válidas.<br/><br/>
    
    <b>2. Variedad de alimentos:</b> Rota los alimentos semanalmente para obtener un perfil completo 
    de micronutrientes.<br/><br/>
    
    <b>3. Preparación y pesaje:</b> Pesa los alimentos en crudo siempre que sea posible para mayor 
    precisión.<br/><br/>
    
    <b>4. Apps de seguimiento:</b> Usa MyFitnessPal, Cronometer o Yazio los primeros 2-3 semanas 
    para aprender las cantidades visualmente.<br/><br/>
    
    <b>5. Hidratación adecuada:</b> Bebe 30-35ml de agua por kg de peso corporal diariamente.<br/><br/>
    """
    
    elements.append(Paragraph(consejos, info_style))
    
    # Construir PDF
    doc.build(elements, canvasmaker=MarcaAguaCanvas)
    buffer.seek(0)
    return buffer


# ============================================================================
# FUNCIÓN GENERAR EXCEL DE DIETA
# ============================================================================

def generar_excel_dieta(plan):
    """
    Genera Excel profesional con ejemplos de comidas
    """
    
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Colores
    COLOR_HEADER = 'C00000'
    COLOR_SUBHEADER = 'E26B0A'
    COLOR_PROTEINA = 'FFC7CE'
    COLOR_CARBOS = 'C6E0B4'
    COLOR_GRASAS = 'FFE699'
    COLOR_ACCENT = 'F2F2F2'
    COLOR_EJEMPLOS = 'D9E1F2'  # Azul claro para ejemplos
    
    # Estilos
    header_font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
    subheader_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    title_font = Font(name='Calibri', size=11, bold=True)
    normal_font = Font(name='Calibri', size=10)
    
    header_fill = PatternFill(start_color=COLOR_HEADER, end_color=COLOR_HEADER, fill_type='solid')
    subheader_fill = PatternFill(start_color=COLOR_SUBHEADER, end_color=COLOR_SUBHEADER, fill_type='solid')
    accent_fill = PatternFill(start_color=COLOR_ACCENT, end_color=COLOR_ACCENT, fill_type='solid')
    ejemplos_fill = PatternFill(start_color=COLOR_EJEMPLOS, end_color=COLOR_EJEMPLOS, fill_type='solid')
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ========== HOJA 1: RESUMEN GENERAL ==========
    ws_resumen = wb.create_sheet("📊 Resumen General", 0)
    
    # Título
    ws_resumen.merge_cells('A1:F2')
    cell = ws_resumen['A1']
    cell.value = "PLAN NUTRICIONAL PERSONALIZADO"
    cell.font = Font(name='Calibri', size=18, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = center_align
    
    # Información del cliente
    ws_resumen.merge_cells('A4:C4')
    ws_resumen['A4'] = "DATOS DEL CLIENTE"
    ws_resumen['A4'].font = title_font
    ws_resumen['A4'].fill = accent_fill
    ws_resumen['A4'].alignment = center_align
    
    cliente_nombre = plan['cliente'] if plan['cliente'] else "Cliente"
    sexo = plan.get('sexo', 'Hombre')
    
    info_cliente = [
        ['Cliente:', cliente_nombre],
        ['Sexo:', sexo],
        ['Objetivo:', plan['objetivo']],
        ['Fecha:', plan['fecha']],
    ]
    
    row = 5
    for label, value in info_cliente:
        ws_resumen[f'A{row}'] = label
        ws_resumen[f'A{row}'].font = title_font
        ws_resumen[f'B{row}'] = value
        ws_resumen[f'B{row}'].font = normal_font
        row += 1
    
    # Objetivos nutricionales
    row += 2
    ws_resumen.merge_cells(f'A{row}:F{row}')
    ws_resumen[f'A{row}'] = "OBJETIVOS NUTRICIONALES DIARIOS"
    ws_resumen[f'A{row}'].font = title_font
    ws_resumen[f'A{row}'].fill = accent_fill
    ws_resumen[f'A{row}'].alignment = center_align
    
    row += 1
    headers_objetivos = ['Macronutriente', 'Cantidad', 'Calorías', '% del Total']
    for col, header in enumerate(headers_objetivos, 1):
        cell = ws_resumen.cell(row, col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    cals_prot = plan['macros']['proteina'] * 4
    cals_carbs = plan['macros']['carbos'] * 4
    cals_grasas = plan['macros']['grasas'] * 9
    total_cals = cals_prot + cals_carbs + cals_grasas
    
    objetivos_data = [
        ['Proteínas', f"{plan['macros']['proteina']:.0f}g", f"{cals_prot:.0f} kcal", f"{(cals_prot/total_cals)*100:.0f}%"],
        ['Carbohidratos', f"{plan['macros']['carbos']:.0f}g", f"{cals_carbs:.0f} kcal", f"{(cals_carbs/total_cals)*100:.0f}%"],
        ['Grasas', f"{plan['macros']['grasas']:.0f}g", f"{cals_grasas:.0f} kcal", f"{(cals_grasas/total_cals)*100:.0f}%"],
        ['TOTAL', '', f"{total_cals:.0f} kcal", '100%'],
    ]
    
    for data_row in objetivos_data:
        for col, value in enumerate(data_row, 1):
            cell = ws_resumen.cell(row, col)
            cell.value = value
            cell.font = Font(name='Calibri', size=10, bold=(data_row[0]=='TOTAL'))
            cell.alignment = center_align
            cell.border = thin_border
            if data_row[0] == 'TOTAL':
                cell.fill = accent_fill
        row += 1
    
    # Distribución por comidas
    row += 2
    ws_resumen.merge_cells(f'A{row}:F{row}')
    ws_resumen[f'A{row}'] = "DISTRIBUCIÓN POR COMIDAS"
    ws_resumen[f'A{row}'].font = title_font
    ws_resumen[f'A{row}'].fill = accent_fill
    ws_resumen[f'A{row}'].alignment = center_align
    
    row += 1
    headers_comidas = ['Comida', 'Proteína (g)', 'Carbohidratos (g)', 'Grasas (g)', 'Calorías', '% del Día']
    for col, header in enumerate(headers_comidas, 1):
        cell = ws_resumen.cell(row, col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    for comida in plan['comidas']:
        cals_comida = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
        comida_data = [
            comida['nombre'],
            f"{comida['proteina']:.0f}",
            f"{comida['carbos']:.0f}",
            f"{comida['grasas']:.0f}",
            f"{cals_comida:.0f}",
            f"{(cals_comida/total_cals)*100:.0f}%"
        ]
        for col, value in enumerate(comida_data, 1):
            cell = ws_resumen.cell(row, col)
            cell.value = value
            cell.font = normal_font
            cell.alignment = center_align
            cell.border = thin_border
        row += 1
    
    # Total
    total_row = [
        'TOTAL DIARIO',
        f"{plan['macros']['proteina']:.0f}",
        f"{plan['macros']['carbos']:.0f}",
        f"{plan['macros']['grasas']:.0f}",
        f"{total_cals:.0f}",
        '100%'
    ]
    for col, value in enumerate(total_row, 1):
        cell = ws_resumen.cell(row, col)
        cell.value = value
        cell.font = title_font
        cell.alignment = center_align
        cell.border = thin_border
        cell.fill = accent_fill
    
    # Ajustar anchos
    ws_resumen.column_dimensions['A'].width = 25
    ws_resumen.column_dimensions['B'].width = 18
    ws_resumen.column_dimensions['C'].width = 20
    ws_resumen.column_dimensions['D'].width = 15
    ws_resumen.column_dimensions['E'].width = 15
    ws_resumen.column_dimensions['F'].width = 12
    
# ========== HOJAS POR CADA COMIDA CON OPCIONES INDIVIDUALES ==========
    for idx, comida in enumerate(plan['comidas'], 1):
        ws_comida = wb.create_sheet(f"🍽️ {comida['nombre']}", idx)
        
        # Título
        ws_comida.merge_cells('A1:E2')
        cell = ws_comida['A1']
        cell.value = comida['nombre'].upper()
        cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
        cell.fill = header_fill
        cell.alignment = center_align
        
        # Macros objetivo
        row = 4
        ws_comida.merge_cells(f'A{row}:E{row}')
        ws_comida[f'A{row}'] = "OBJETIVOS DE ESTA COMIDA"
        ws_comida[f'A{row}'].font = title_font
        ws_comida[f'A{row}'].fill = accent_fill
        ws_comida[f'A{row}'].alignment = center_align
        
        row += 1
        macros_headers = ['Macronutriente', 'Cantidad Objetivo']
        for col, header in enumerate(macros_headers, 1):
            cell = ws_comida.cell(row, col)
            cell.value = header
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = center_align
            cell.border = thin_border
        
        row += 1
        cals_comida = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
        macros_data = [
            ['Proteína', f"{comida['proteina']:.0f}g"],
            ['Carbohidratos', f"{comida['carbos']:.0f}g"],
            ['Grasas', f"{comida['grasas']:.0f}g"],
            ['Calorías Totales', f"{cals_comida:.0f} kcal"],
        ]
        
        for data_row in macros_data:
            ws_comida.cell(row, 1).value = data_row[0]
            ws_comida.cell(row, 1).font = title_font
            ws_comida.cell(row, 1).border = thin_border
            ws_comida.cell(row, 2).value = data_row[1]
            ws_comida.cell(row, 2).font = normal_font
            ws_comida.cell(row, 2).alignment = center_align
            ws_comida.cell(row, 2).border = thin_border
            row += 1
        
        # ===== PROTEÍNA =====
        if comida['proteina'] > 5:
            row += 2
            ws_comida.merge_cells(f'A{row}:E{row}')
            ws_comida[f'A{row}'] = f"OPCIONES DE PROTEÍNA ({comida['proteina']:.0f}g)"
            ws_comida[f'A{row}'].font = title_font
            ws_comida[f'A{row}'].fill = PatternFill(start_color=COLOR_PROTEINA, end_color=COLOR_PROTEINA, fill_type='solid')
            ws_comida[f'A{row}'].alignment = center_align
            
            row += 1
            prot_headers = ['Alimento', 'Cantidad (g)', 'Equivalencia']
            for col, header in enumerate(prot_headers, 1):
                cell = ws_comida.cell(row, col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            row += 1
            opciones_prot = seleccionar_mejores_opciones_comida(
                comida['proteina'], 
                ALIMENTOS_PROTEINAS, 
                'proteina', 
                num_opciones_por_categoria=2
            )
            
            if opciones_prot:
                for opcion in opciones_prot:
                    if opcion and opcion.get('racion'):
                        ws_comida.cell(row, 1).value = opcion['alimento']['nombre']
                        ws_comida.cell(row, 1).font = normal_font
                        ws_comida.cell(row, 1).border = thin_border
                        
                        ws_comida.cell(row, 2).value = f"{opcion['racion']['gramos_alimento']:.0f}g"
                        ws_comida.cell(row, 2).font = normal_font
                        ws_comida.cell(row, 2).alignment = center_align
                        ws_comida.cell(row, 2).border = thin_border
                        
                        ws_comida.cell(row, 3).value = opcion['racion']['descripcion']
                        ws_comida.cell(row, 3).font = normal_font
                        ws_comida.cell(row, 3).alignment = center_align
                        ws_comida.cell(row, 3).border = thin_border
                        row += 1
        
        # ===== CARBOHIDRATOS =====
        if comida['carbos'] > 5:
            row += 2
            ws_comida.merge_cells(f'A{row}:E{row}')
            ws_comida[f'A{row}'] = f"OPCIONES DE CARBOHIDRATOS ({comida['carbos']:.0f}g)"
            ws_comida[f'A{row}'].font = title_font
            ws_comida[f'A{row}'].fill = PatternFill(start_color=COLOR_CARBOS, end_color=COLOR_CARBOS, fill_type='solid')
            ws_comida[f'A{row}'].alignment = center_align
            
            row += 1
            carbs_headers = ['Alimento', 'Cantidad (g)', 'Equivalencia']
            for col, header in enumerate(carbs_headers, 1):
                cell = ws_comida.cell(row, col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            row += 1
            opciones_carbs = seleccionar_mejores_opciones_comida(
                comida['carbos'], 
                ALIMENTOS_CARBOHIDRATOS, 
                'carbos', 
                num_opciones_por_categoria=2
            )
            
            if opciones_carbs:
                for opcion in opciones_carbs:
                    if opcion and opcion.get('racion'):
                        ws_comida.cell(row, 1).value = opcion['alimento']['nombre']
                        ws_comida.cell(row, 1).font = normal_font
                        ws_comida.cell(row, 1).border = thin_border
                        
                        ws_comida.cell(row, 2).value = f"{opcion['racion']['gramos_alimento']:.0f}g"
                        ws_comida.cell(row, 2).font = normal_font
                        ws_comida.cell(row, 2).alignment = center_align
                        ws_comida.cell(row, 2).border = thin_border
                        
                        ws_comida.cell(row, 3).value = opcion['racion']['descripcion']
                        ws_comida.cell(row, 3).font = normal_font
                        ws_comida.cell(row, 3).alignment = center_align
                        ws_comida.cell(row, 3).border = thin_border
                        row += 1
        
        # ===== GRASAS =====
        if comida['grasas'] > 3:
            row += 2
            ws_comida.merge_cells(f'A{row}:E{row}')
            ws_comida[f'A{row}'] = f"OPCIONES DE GRASAS ({comida['grasas']:.0f}g)"
            ws_comida[f'A{row}'].font = title_font
            ws_comida[f'A{row}'].fill = PatternFill(start_color=COLOR_GRASAS, end_color=COLOR_GRASAS, fill_type='solid')
            ws_comida[f'A{row}'].alignment = center_align
            
            row += 1
            grasas_headers = ['Alimento', 'Cantidad (g)', 'Equivalencia']
            for col, header in enumerate(grasas_headers, 1):
                cell = ws_comida.cell(row, col)
                cell.value = header
                cell.font = subheader_font
                cell.fill = subheader_fill
                cell.alignment = center_align
                cell.border = thin_border
            
            row += 1
            opciones_grasas = seleccionar_mejores_opciones_comida(
                comida['grasas'], 
                ALIMENTOS_GRASAS, 
                'grasas', 
                num_opciones_por_categoria=2
            )
            
            if opciones_grasas:
                for opcion in opciones_grasas:
                    if opcion and opcion.get('racion'):
                        ws_comida.cell(row, 1).value = opcion['alimento']['nombre']
                        ws_comida.cell(row, 1).font = normal_font
                        ws_comida.cell(row, 1).border = thin_border
                        
                        ws_comida.cell(row, 2).value = f"{opcion['racion']['gramos_alimento']:.0f}g"
                        ws_comida.cell(row, 2).font = normal_font
                        ws_comida.cell(row, 2).alignment = center_align
                        ws_comida.cell(row, 2).border = thin_border
                        
                        ws_comida.cell(row, 3).value = opcion['racion']['descripcion']
                        ws_comida.cell(row, 3).font = normal_font
                        ws_comida.cell(row, 3).alignment = center_align
                        ws_comida.cell(row, 3).border = thin_border
                        row += 1
        
        # Nota verduras
        row += 2
        ws_comida.merge_cells(f'A{row}:E{row}')
        ws_comida[f'A{row}'] = "💚 VERDURAS: Añade la cantidad que desees sin contabilizar (lechuga, tomate, pepino, brócoli, espinacas, etc.)"
        ws_comida[f'A{row}'].font = Font(name='Calibri', size=9, italic=True)
        ws_comida[f'A{row}'].alignment = left_align
        ws_comida[f'A{row}'].fill = PatternFill(start_color='E2EFDA', end_color='E2EFDA', fill_type='solid')
        
        # Ajustar anchos
        ws_comida.column_dimensions['A'].width = 30
        ws_comida.column_dimensions['B'].width = 15
        ws_comida.column_dimensions['C'].width = 20
        ws_comida.column_dimensions['D'].width = 15
        ws_comida.column_dimensions['E'].width = 15
    
    # ========== HOJAS DE EJEMPLOS DE COMIDAS COMPLETAS (SEPARADAS) ==========
    generador = GeneradorComidasOptimas()
    
    for idx_comida, comida in enumerate(plan['comidas']):
        # Generar ejemplos
        ejemplos = generador.generar_ejemplos_comida(comida)
        
        if not ejemplos:
            continue
        
        # Crear hoja de ejemplos
        ws_ejemplos = wb.create_sheet(f"📋 Ejemplos {comida['nombre']}")
        
        # Título
        ws_ejemplos.merge_cells('A1:D2')
        cell = ws_ejemplos['A1']
        cell.value = f"EJEMPLOS DE COMIDAS COMPLETAS - {comida['nombre'].upper()}"
        cell.font = Font(name='Calibri', size=14, bold=True, color='FFFFFF')
        cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        cell.alignment = center_align
        
        # Subtítulo con objetivo
        cals_comida = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
        ws_ejemplos.merge_cells('A3:D3')
        ws_ejemplos['A3'] = f"Objetivo: {comida['proteina']:.0f}g P | {comida['carbos']:.0f}g C | {comida['grasas']:.0f}g G | {cals_comida:.0f} kcal"
        ws_ejemplos['A3'].font = Font(name='Calibri', size=10, italic=True)
        ws_ejemplos['A3'].alignment = center_align
        
        row = 5
        
        # Mostrar cada ejemplo
        for idx, ejemplo in enumerate(ejemplos, 1):
            # Título del ejemplo
            ws_ejemplos.merge_cells(f'A{row}:D{row}')
            ws_ejemplos[f'A{row}'] = f"Ejemplo {idx}: {ejemplo['tipo']}"
            ws_ejemplos[f'A{row}'].font = Font(name='Calibri', size=11, bold=True)
            ws_ejemplos[f'A{row}'].fill = PatternFill(start_color=COLOR_EJEMPLOS, end_color=COLOR_EJEMPLOS, fill_type='solid')
            ws_ejemplos[f'A{row}'].alignment = left_align
            row += 1
            
            # Encabezados
            headers = ['Alimento', 'Cantidad', 'Equivalencia']
            for col, header in enumerate(headers, 1):
                cell = ws_ejemplos.cell(row, col)
                cell.value = header
                cell.font = Font(name='Calibri', size=10, bold=True, color='FFFFFF')
                cell.fill = subheader_fill
                cell.alignment = center_align
                cell.border = thin_border
            row += 1
            
            # Alimentos
            for alim in ejemplo['alimentos']:
                ws_ejemplos.cell(row, 1).value = alim['nombre']
                ws_ejemplos.cell(row, 1).font = normal_font
                ws_ejemplos.cell(row, 1).border = thin_border
                
                ws_ejemplos.cell(row, 2).value = f"{alim['gramos']:.0f}g"
                ws_ejemplos.cell(row, 2).font = normal_font
                ws_ejemplos.cell(row, 2).alignment = center_align
                ws_ejemplos.cell(row, 2).border = thin_border
                
                ws_ejemplos.cell(row, 3).value = alim['equivalencia']
                ws_ejemplos.cell(row, 3).font = normal_font
                ws_ejemplos.cell(row, 3).alignment = center_align
                ws_ejemplos.cell(row, 3).border = thin_border
                
                row += 1
            
            # Totales
            ws_ejemplos.cell(row, 1).value = 'TOTAL'
            ws_ejemplos.cell(row, 1).font = Font(name='Calibri', size=10, bold=True)
            ws_ejemplos.cell(row, 1).fill = accent_fill
            ws_ejemplos.cell(row, 1).alignment = center_align
            ws_ejemplos.cell(row, 1).border = thin_border
            
            ws_ejemplos.cell(row, 2).value = f"{ejemplo['macros_totales']['calorias']:.0f} kcal"
            ws_ejemplos.cell(row, 2).font = Font(name='Calibri', size=10, bold=True)
            ws_ejemplos.cell(row, 2).fill = accent_fill
            ws_ejemplos.cell(row, 2).alignment = center_align
            ws_ejemplos.cell(row, 2).border = thin_border
            
            ws_ejemplos.cell(row, 3).value = (
                f"P: {ejemplo['macros_totales']['proteina']:.0f}g | "
                f"C: {ejemplo['macros_totales']['carbos']:.0f}g | "
                f"G: {ejemplo['macros_totales']['grasas']:.0f}g"
            )
            ws_ejemplos.cell(row, 3).font = Font(name='Calibri', size=9, bold=True)
            ws_ejemplos.cell(row, 3).fill = accent_fill
            ws_ejemplos.cell(row, 3).alignment = center_align
            ws_ejemplos.cell(row, 3).border = thin_border
            
            row += 2  # Espacio entre ejemplos
        
        # Ajustar anchos
        ws_ejemplos.column_dimensions['A'].width = 30
        ws_ejemplos.column_dimensions['B'].width = 15
        ws_ejemplos.column_dimensions['C'].width = 25
        ws_ejemplos.column_dimensions['D'].width = 15

    # ========== HOJA: PRE-ENTRENAMIENTO ==========
    ws_pre = wb.create_sheet("⚡ Pre-Entrenamiento")
    
    # Título principal
    ws_pre.merge_cells('A1:E2')
    cell = ws_pre['A1']
    cell.value = "COMIDA PRE-ENTRENAMIENTO"
    cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    cell.fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    cell.alignment = center_align
    
    # Subtítulo
    ws_pre.merge_cells('A3:E3')
    ws_pre['A3'] = "Solo días de entrenamiento - 30-60 minutos antes"
    ws_pre['A3'].font = Font(name='Calibri', size=11, italic=True)
    ws_pre['A3'].alignment = center_align
    
    # Determinar macros según sexo
    if sexo == "Mujer":
        proteina_pre = 15
        carbos_pre = 25
    else:
        proteina_pre = 22
        carbos_pre = 35
    
    calorias_pre = (proteina_pre * 4) + (carbos_pre * 4)
    
    # Objetivos
    row = 5
    ws_pre.merge_cells(f'A{row}:E{row}')
    ws_pre[f'A{row}'] = "OBJETIVOS"
    ws_pre[f'A{row}'].font = title_font
    ws_pre[f'A{row}'].fill = accent_fill
    ws_pre[f'A{row}'].alignment = center_align
    
    row += 1
    macros_headers = ['Macronutriente', 'Cantidad Objetivo']
    for col, header in enumerate(macros_headers, 1):
        cell = ws_pre.cell(row, col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    macros_pre_data = [
        ['Proteína', f"{proteina_pre}g ({proteina_pre-5}-{proteina_pre+5}g)"],
        ['Carbohidratos', f"{carbos_pre}g ({carbos_pre-5}-{carbos_pre+5}g)"],
        ['Grasas', 'Mínimas (evitar)'],
        ['Calorías', f"~{calorias_pre:.0f} kcal"],
    ]
    
    for data_row in macros_pre_data:
        ws_pre.cell(row, 1).value = data_row[0]
        ws_pre.cell(row, 1).font = title_font
        ws_pre.cell(row, 1).border = thin_border
        ws_pre.cell(row, 2).value = data_row[1]
        ws_pre.cell(row, 2).font = normal_font
        ws_pre.cell(row, 2).alignment = center_align
        ws_pre.cell(row, 2).border = thin_border
        row += 1
    
    # PROTEÍNAS PRE-ENTRENAMIENTO
    row += 2
    ws_pre.merge_cells(f'A{row}:E{row}')
    ws_pre[f'A{row}'] = f"PARA {proteina_pre}g DE PROTEÍNA:"
    ws_pre[f'A{row}'].font = title_font
    ws_pre[f'A{row}'].fill = PatternFill(start_color=COLOR_PROTEINA, end_color=COLOR_PROTEINA, fill_type='solid')
    ws_pre[f'A{row}'].alignment = center_align
    
    row += 1
    prot_headers = ['Alimento', 'Cantidad (g)', 'Equivalencia']
    for col, header in enumerate(prot_headers, 1):
        cell = ws_pre.cell(row, col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    alimentos_pre_prot = [
        {'nombre': 'Proteína whey', 'proteina': 80, 'carbos': 5, 'grasas': 3, 'unidad': 'cacito', 'gramos': 30}
    ]
    
    for alim in alimentos_pre_prot:
        racion = calcular_racion_exacta(proteina_pre, alim, 'proteina')
        if racion:
            ws_pre.cell(row, 1).value = alim['nombre']
            ws_pre.cell(row, 1).font = normal_font
            ws_pre.cell(row, 1).border = thin_border
            
            ws_pre.cell(row, 2).value = f"{racion['gramos_alimento']:.0f}g"
            ws_pre.cell(row, 2).font = normal_font
            ws_pre.cell(row, 2).alignment = center_align
            ws_pre.cell(row, 2).border = thin_border
            
            ws_pre.cell(row, 3).value = racion['descripcion']
            ws_pre.cell(row, 3).font = normal_font
            ws_pre.cell(row, 3).alignment = center_align
            ws_pre.cell(row, 3).border = thin_border
            row += 1
    
    # CARBOHIDRATOS PRE-ENTRENAMIENTO
    row += 2
    ws_pre.merge_cells(f'A{row}:E{row}')
    ws_pre[f'A{row}'] = f"PARA {carbos_pre}g DE CARBOHIDRATOS:"
    ws_pre[f'A{row}'].font = title_font
    ws_pre[f'A{row}'].fill = PatternFill(start_color=COLOR_CARBOS, end_color=COLOR_CARBOS, fill_type='solid')
    ws_pre[f'A{row}'].alignment = center_align
    
    row += 1
    carbs_headers = ['Alimento', 'Cantidad (g)', 'Equivalencia']
    for col, header in enumerate(carbs_headers, 1):
        cell = ws_pre.cell(row, col)
        cell.value = header
        cell.font = subheader_font
        cell.fill = subheader_fill
        cell.alignment = center_align
        cell.border = thin_border
    
    row += 1
    alimentos_pre_carbs = [
        {'nombre': 'Plátano', 'proteina': 1.1, 'carbos': 23, 'grasas': 0.3, 'unidad': 'plátano', 'gramos': 120},
        {'nombre': 'Copos de maíz sin azúcar', 'proteina': 8, 'carbos': 80, 'grasas': 1, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Dátil', 'proteina': 2, 'carbos': 75, 'grasas': 0.2, 'unidad': 'dátil', 'gramos': 12},
        {'nombre': 'Tortitas de arroz', 'proteina': 8, 'carbos': 81, 'grasas': 3, 'unidad': 'tortita', 'gramos': 9},
        {'nombre': 'Miel', 'proteina': 0.3, 'carbos': 82, 'grasas': 0, 'unidad': 'cucharada', 'gramos': 21},
    ]
    
    for alim in alimentos_pre_carbs:
        racion = calcular_racion_exacta(carbos_pre, alim, 'carbos')
        if racion:
            ws_pre.cell(row, 1).value = alim['nombre']
            ws_pre.cell(row, 1).font = normal_font
            ws_pre.cell(row, 1).border = thin_border
            
            ws_pre.cell(row, 2).value = f"{racion['gramos_alimento']:.0f}g"
            ws_pre.cell(row, 2).font = normal_font
            ws_pre.cell(row, 2).alignment = center_align
            ws_pre.cell(row, 2).border = thin_border
            
            ws_pre.cell(row, 3).value = racion['descripcion']
            ws_pre.cell(row, 3).font = normal_font
            ws_pre.cell(row, 3).alignment = center_align
            ws_pre.cell(row, 3).border = thin_border
            row += 1
    
    # NOTA IMPORTANTE
    row += 2
    ws_pre.merge_cells(f'A{row}:E{row+2}')
    nota_texto = "⚠️ IMPORTANTE: Esta comida solo se consume los días de entrenamiento, 30-60 minutos antes. Combina 1 opción de proteína + 1 opción de carbohidratos. Evita grasas para facilitar digestión."
    ws_pre[f'A{row}'] = nota_texto
    ws_pre[f'A{row}'].font = Font(name='Calibri', size=10, bold=True, color='FF0000')
    ws_pre[f'A{row}'].alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    ws_pre[f'A{row}'].fill = PatternFill(start_color='FFF2CC', end_color='FFF2CC', fill_type='solid')
    ws_pre[f'A{row}'].border = thin_border
    
    # Ajustar anchos de columna
    ws_pre.column_dimensions['A'].width = 30
    ws_pre.column_dimensions['B'].width = 15
    ws_pre.column_dimensions['C'].width = 20
    ws_pre.column_dimensions['D'].width = 15
    ws_pre.column_dimensions['E'].width = 15
    
    # Ajustar alturas de filas para el texto largo
    ws_pre.row_dimensions[row].height = 45

    # ========== HOJA: CONSEJOS ==========
    ws_consejos = wb.create_sheet("📝 Consejos")
    
    ws_consejos.merge_cells('A1:D2')
    cell = ws_consejos['A1']
    cell.value = "CONSEJOS PRÁCTICOS PARA EL ÉXITO"
    cell.font = Font(name='Calibri', size=16, bold=True, color='FFFFFF')
    cell.fill = header_fill
    cell.alignment = center_align
    
    consejos_list = [
        "1. Flexibilidad: Aproximaciones del 5-10% son válidas",
        "2. Variedad: Rota los alimentos semanalmente",
        "3. Pesaje: Pesa en crudo para mayor precisión",
        "4. Hidratación: 30-35ml de agua por kg de peso",
        "5. Apps: Usa MyFitnessPal o Cronometer las primeras semanas"
    ]
    
    row = 4
    for consejo in consejos_list:
        ws_consejos[f'A{row}'] = consejo
        ws_consejos[f'A{row}'].font = normal_font
        ws_consejos[f'A{row}'].alignment = left_align
        ws_consejos.merge_cells(f'A{row}:D{row}')
        row += 1
    
    ws_consejos.column_dimensions['A'].width = 80
    
    # Guardar
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# CONFIGURACIÓN INICIAL
# ============================================================================
st.set_page_config(
    page_title="Constructor de Dieta Pro", 
    page_icon="🥗", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# BASE DE DATOS AMPLIADA 
# Valores por porción indicada (basados en USDA FoodData Central y BEDCA)
# ============================================================================

ALIMENTOS_PROTEINAS = {
    'Carnes magras': [
        {'nombre': 'Pechuga de pollo', 'proteina': 23, 'carbos': 0, 'grasas': 1.2, 'unidad': 'pechuga', 'gramos': 150},
        {'nombre': 'Pavo', 'proteina': 24, 'carbos': 0, 'grasas': 1, 'unidad': 'filete', 'gramos': 120},
        {'nombre': 'Ternera magra', 'proteina': 22, 'carbos': 0, 'grasas': 3, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Solomillo de cerdo', 'proteina': 22, 'carbos': 0, 'grasas': 3.5, 'unidad': 'filete', 'gramos': 120},
        {'nombre': 'Conejo', 'proteina': 22, 'carbos': 0, 'grasas': 3.5, 'unidad': 'muslo', 'gramos': 100},
        {'nombre': 'Lomo embuchado', 'proteina': 50, 'carbos': 1, 'grasas': 15, 'unidad': 'loncha', 'gramos': 10},
    ],
    'Pescados blancos': [
        {'nombre': 'Merluza', 'proteina': 17, 'carbos': 0, 'grasas': 2, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Bacalao', 'proteina': 18, 'carbos': 0, 'grasas': 0.7, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Lenguado', 'proteina': 17, 'carbos': 0, 'grasas': 1.5, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Rape', 'proteina': 16, 'carbos': 0, 'grasas': 1.3, 'unidad': 'rodaja', 'gramos': 150},
        {'nombre': 'Lubina', 'proteina': 18, 'carbos': 0, 'grasas': 1.5, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Dorada', 'proteina': 18, 'carbos': 0, 'grasas': 1, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Gallo', 'proteina': 16, 'carbos': 0, 'grasas': 0.8, 'unidad': 'filete', 'gramos': 150},
    ],
    'Pescados azules': [
        {'nombre': 'Salmón', 'proteina': 20, 'carbos': 0, 'grasas': 13, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Atún en lata al natural', 'proteina': 26, 'carbos': 0, 'grasas': 1, 'unidad': 'lata', 'gramos': 80},
        {'nombre': 'Atún fresco', 'proteina': 23, 'carbos': 0, 'grasas': 6, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Sardinas', 'proteina': 25, 'carbos': 0, 'grasas': 11, 'unidad': 'sardina', 'gramos': 35},
        {'nombre': 'Caballa', 'proteina': 19, 'carbos': 0, 'grasas': 14, 'unidad': 'filete', 'gramos': 120},
        {'nombre': 'Boquerones', 'proteina': 20, 'carbos': 0, 'grasas': 6, 'unidad': 'unidad', 'gramos': 20},
        {'nombre': 'Trucha', 'proteina': 16, 'carbos': 0, 'grasas': 3, 'unidad': 'filete', 'gramos': 150},
        {'nombre': 'Salmón ahumado', 'proteina': 18, 'carbos': 0, 'grasas': 4, 'unidad': 'loncha', 'gramos': 18},
    ],
    'Mariscos': [
        {'nombre': 'Calamar', 'proteina': 18, 'carbos': 3, 'grasas': 1.4, 'unidad': 'ración', 'gramos': 100},
        {'nombre': 'Pulpo', 'proteina': 16, 'carbos': 2.2, 'grasas': 1, 'unidad': 'ración', 'gramos': 100},
        {'nombre': 'Gambas', 'proteina': 20, 'carbos': 0, 'grasas': 1.4, 'unidad': 'ración', 'gramos': 100},
        {'nombre': 'Langostinos', 'proteina': 20, 'carbos': 0, 'grasas': 1.4, 'unidad': 'unidad', 'gramos': 15},
        {'nombre': 'Mejillones', 'proteina': 17, 'carbos': 7, 'grasas': 2.2, 'unidad': 'ración', 'gramos': 100},
        {'nombre': 'Sepia', 'proteina': 16, 'carbos': 0.7, 'grasas': 1.5, 'unidad': 'ración', 'gramos': 100},
        
    ],
    'Lácteos': [
        {'nombre': 'Yogur natural', 'proteina': 4, 'carbos': 6, 'grasas': 3.5, 'unidad': 'yogur', 'gramos': 125},
        {'nombre': 'Queso cottage', 'proteina': 11, 'carbos': 3.4, 'grasas': 4.3, 'unidad': 'tarrina', 'gramos': 200},
        {'nombre': 'Queso fresco batido 0%', 'proteina': 8, 'carbos': 4, 'grasas': 0.2, 'unidad': 'tarrina', 'gramos': 250},
        {'nombre': 'Yogur griego 0%', 'proteina': 10, 'carbos': 4, 'grasas': 0.2, 'unidad': 'yogur', 'gramos': 150},
        {'nombre': 'Yogur desnatado', 'proteina': 4, 'carbos': 6, 'grasas': 0.2, 'unidad': 'yogur', 'gramos': 125},
        {'nombre': 'Leche desnatada', 'proteina': 3.4, 'carbos': 5, 'grasas': 0.1, 'unidad': 'vaso', 'gramos': 200},
        {'nombre': 'Leche entera', 'proteina': 3.2, 'carbos': 4.8, 'grasas': 3.6, 'unidad': 'vaso', 'gramos': 200},
        {'nombre': 'Kéfir', 'proteina': 3.3, 'carbos': 4.5, 'grasas': 0.5, 'unidad': 'vaso', 'gramos': 200},
        {'nombre': 'Skyr', 'proteina': 10, 'carbos': 4, 'grasas': 0.2, 'unidad': 'yogur', 'gramos': 125},
        {'nombre': 'Yogur proteico', 'proteina': 10, 'carbos': 4.7, 'grasas': 0.2, 'unidad': 'yogur', 'gramos': 125},
    ],
    'Huevos': [
        {'nombre': 'Huevo entero', 'proteina': 13, 'carbos': 1.1, 'grasas': 11, 'unidad': 'huevo', 'gramos': 60},
        {'nombre': 'Clara de huevo', 'proteina': 11, 'carbos': 0.7, 'grasas': 0.2, 'unidad': 'clara', 'gramos': 33},
        {'nombre': 'Claras pasteurizadas', 'proteina': 11, 'carbos': 0.7, 'grasas': 0.1, 'unidad': 'vaso', 'gramos': 200},
    ],
    'Legumbres': [
        {'nombre': 'Lentejas cocidas', 'proteina': 9, 'carbos': 20, 'grasas': 0.4, 'unidad': 'plato', 'gramos': 200},
        {'nombre': 'Garbanzos cocidos', 'proteina': 8.9, 'carbos': 27, 'grasas': 2.6, 'unidad': 'plato', 'gramos': 200},
        {'nombre': 'Judías pintas cocidas', 'proteina': 8.7, 'carbos': 22, 'grasas': 0.5, 'unidad': 'plato', 'gramos': 200},
        {'nombre': 'Judías blancas cocidas', 'proteina': 8, 'carbos': 22, 'grasas': 0.7, 'unidad': 'plato', 'gramos': 200},
        {'nombre': 'Soja texturizada', 'proteina': 50, 'carbos': 30, 'grasas': 1, 'unidad': 'taza', 'gramos': 80},
        {'nombre': 'Edamame', 'proteina': 11, 'carbos': 10, 'grasas': 5, 'unidad': 'bol', 'gramos': 150},
        {'nombre': 'Guisantes cocidos', 'proteina': 6, 'carbos': 15, 'grasas': 0.5, 'unidad': 'bol', 'gramos': 160},
    ],
    'Derivados de soja': [
        {'nombre': 'Tofu firme', 'proteina': 8, 'carbos': 2, 'grasas': 4.8, 'unidad': 'bloque', 'gramos': 200},
        {'nombre': 'Tempeh', 'proteina': 19, 'carbos': 9, 'grasas': 11, 'unidad': 'porción', 'gramos': 100},
        {'nombre': 'Bebida de soja', 'proteina': 3.3, 'carbos': 2.5, 'grasas': 1.8, 'unidad': 'vaso', 'gramos': 200},
        {'nombre': 'Seitan', 'proteina': 25, 'carbos': 7, 'grasas': 2, 'unidad': 'porción', 'gramos': 100},
    ],
    'Suplementos': [
        {'nombre': 'Proteína whey', 'proteina': 80, 'carbos': 5, 'grasas': 3, 'unidad': 'cacito', 'gramos': 30},
        {'nombre': 'Proteína caseína', 'proteina': 78, 'carbos': 6, 'grasas': 2, 'unidad': 'cacito', 'gramos': 30},
        {'nombre': 'Proteína vegetal', 'proteina': 75, 'carbos': 8, 'grasas': 4, 'unidad': 'cacito', 'gramos': 30},
        {'nombre': 'Proteína de huevo', 'proteina': 80, 'carbos': 3, 'grasas': 1, 'unidad': 'cacito', 'gramos': 30},
    ],
    'Quesos': [
        {'nombre': 'Queso mozzarella light', 'proteina': 20, 'carbos': 2, 'grasas': 16, 'unidad': 'loncha', 'gramos': 30},
        {'nombre': 'Queso fresco de Burgos', 'proteina': 15, 'carbos': 1.3, 'grasas': 2.5, 'unidad': 'porción', 'gramos': 80},
        {'nombre': 'Queso curado', 'proteina': 25, 'carbos': 1.3, 'grasas': 33, 'unidad': 'loncha', 'gramos': 30},
        {'nombre': 'Queso tierno', 'proteina': 20, 'carbos': 2, 'grasas': 23, 'unidad': 'loncha', 'gramos': 40},
        {'nombre': 'Requesón', 'proteina': 11, 'carbos': 3, 'grasas': 4, 'unidad': 'tarrina', 'gramos': 200},
        {'nombre': 'Queso parmesano', 'proteina': 36, 'carbos': 1, 'grasas': 28, 'unidad': 'cucharada', 'gramos': 15},
        {'nombre': 'Queso feta', 'proteina': 14, 'carbos': 4, 'grasas': 20, 'unidad': 'porción', 'gramos': 80},
    ],
    'Embutidos magros': [
        {'nombre': 'Jamón cocido', 'proteina': 19, 'carbos': 1, 'grasas': 3, 'unidad': 'loncha', 'gramos': 25},
        {'nombre': 'Pechuga de pavo', 'proteina': 22, 'carbos': 1, 'grasas': 1.5, 'unidad': 'loncha', 'gramos': 25},
        {'nombre': 'Jamón serrano', 'proteina': 30, 'carbos': 0, 'grasas': 16, 'unidad': 'loncha', 'gramos': 20},
        {'nombre': 'Jamón ibérico', 'proteina': 33, 'carbos': 0, 'grasas': 25, 'unidad': 'loncha', 'gramos': 20},
        {'nombre': 'Cecina', 'proteina': 35, 'carbos': 0.8, 'grasas': 6, 'unidad': 'loncha', 'gramos': 15},
    ]
}

ALIMENTOS_CARBOHIDRATOS = {
    'Cereales integrales': [
        {'nombre': 'Avena en copos', 'proteina': 13.5, 'carbos': 58, 'grasas': 7, 'unidad': 'vaso', 'gramos': 80, 'cocido': False},
        {'nombre': 'Arroz integral cocido', 'proteina': 2.6, 'carbos': 23, 'grasas': 0.9, 'unidad': 'plato', 'gramos': 200, 'cocido': True},
        {'nombre': 'Arroz integral crudo', 'proteina': 7.5, 'carbos': 77, 'grasas': 2.7, 'unidad': 'vaso', 'gramos': 180, 'cocido': False},
        {'nombre': 'Quinoa cocida', 'proteina': 4, 'carbos': 21, 'grasas': 1.9, 'unidad': '1 plato', 'gramos': 185, 'cocido': True},
        {'nombre': 'Pasta integral cocida', 'proteina': 5, 'carbos': 26, 'grasas': 1.1, 'unidad': 'plato', 'gramos': 200, 'cocido': True},
        {'nombre': 'Pan integral', 'proteina': 9, 'carbos': 49, 'grasas': 3.5, 'unidad': '1 rebanada', 'gramos': 30, 'cocido': False},
        {'nombre': 'Trigo sarraceno cocido', 'proteina': 3.4, 'carbos': 20, 'grasas': 0.6, 'unidad': 'plato', 'gramos': 170, 'cocido': True},
        {'nombre': 'Cuscús integral cocido', 'proteina': 3.8, 'carbos': 23, 'grasas': 0.4, 'unidad': 'plato', 'gramos': 160, 'cocido': True},
        {'nombre': 'Pan de centeno', 'proteina': 8.5, 'carbos': 48, 'grasas': 1.7, 'unidad': 'rebanada', 'gramos': 30, 'cocido': False},
        {'nombre': 'Tortitas de arroz', 'proteina': 8, 'carbos': 81, 'grasas': 3, 'unidad': 'tortita', 'gramos': 9, 'cocido': False},
        {'nombre': 'Copos de maíz sin azúcar', 'proteina': 8, 'carbos': 80, 'grasas': 1, 'unidad': 'bol', 'gramos': 75, 'cocido': False},
    ],
    'Cereales refinados': [
        {'nombre': 'Arroz blanco cocido', 'proteina': 2.7, 'carbos': 28, 'grasas': 0.3, 'unidad': 'plato', 'gramos': 200, 'cocido': True},
        {'nombre': 'Pasta blanca cocida', 'proteina': 5, 'carbos': 31, 'grasas': 0.9, 'unidad': 'plato', 'gramos': 200, 'cocido': True},
        {'nombre': 'Arroz blanco crudo', 'proteina': 7, 'carbos': 80, 'grasas': 0.9, 'unidad': 'vaso', 'gramos': 180, 'cocido': False},
        {'nombre': 'Cuscús cocido', 'proteina': 3.8, 'carbos': 24, 'grasas': 0.2, 'unidad': 'plato', 'gramos': 160, 'cocido': True},
    ],
    'Pan': [
        {'nombre': 'Pan blanco', 'proteina': 9, 'carbos': 51, 'grasas': 3.2, 'unidad': 'rebanada', 'gramos': 30, 'cocido': False},
        {'nombre': 'Baguette', 'proteina': 8, 'carbos': 55, 'grasas': 1.5, 'unidad': 'barra', 'gramos': 240, 'cocido': False},
        {'nombre': 'Pan de molde blanco', 'proteina': 9, 'carbos': 52, 'grasas': 3, 'unidad': 'rebanada', 'gramos': 25, 'cocido': False},      
    ],
    'Tubérculos': [
        {'nombre': 'Patata hervida', 'proteina': 2, 'carbos': 17, 'grasas': 0.1, 'unidad': 'mediana', 'gramos': 150, 'cocido': True},
        {'nombre': 'Boniato', 'proteina': 1.6, 'carbos': 20, 'grasas': 0.1, 'unidad': '1 unidad', 'gramos': 200, 'cocido': True},
        {'nombre': 'Patata asada', 'proteina': 2.5, 'carbos': 21, 'grasas': 0.1, 'unidad': 'mediana', 'gramos': 150, 'cocido': True},
        {'nombre': 'Yuca cocida', 'proteina': 0.8, 'carbos': 38, 'grasas': 0.3, 'unidad': '1 porción', 'gramos': 200, 'cocido': True},
    ],
    'Frutas': [
        {'nombre': 'Plátano', 'proteina': 1.1, 'carbos': 23, 'grasas': 0.3, 'unidad': 'plátano', 'gramos': 120, 'cocido': False},
        {'nombre': 'Manzana', 'proteina': 0.3, 'carbos': 14, 'grasas': 0.2, 'unidad': 'manzana', 'gramos': 180, 'cocido': False},
        {'nombre': 'Pera', 'proteina': 0.4, 'carbos': 15, 'grasas': 0.1, 'unidad': 'pera', 'gramos': 180, 'cocido': False},
        {'nombre': 'Naranja', 'proteina': 0.9, 'carbos': 12, 'grasas': 0.1, 'unidad': 'naranja', 'gramos': 150, 'cocido': False},
        {'nombre': 'Mandarina', 'proteina': 0.8, 'carbos': 13, 'grasas': 0.3, 'unidad': 'mandarina', 'gramos': 100, 'cocido': False},
        {'nombre': 'Kiwi', 'proteina': 1.1, 'carbos': 15, 'grasas': 0.5, 'unidad': 'kiwi', 'gramos': 100, 'cocido': False},
        {'nombre': 'Fresas', 'proteina': 0.7, 'carbos': 8, 'grasas': 0.3, 'unidad': 'bol', 'gramos': 150, 'cocido': False},
        {'nombre': 'Arándanos', 'proteina': 0.7, 'carbos': 14, 'grasas': 0.3, 'unidad': 'bol', 'gramos': 150, 'cocido': False},
        {'nombre': 'Frambuesas', 'proteina': 0.7, 'carbos': 8, 'grasas': 0.5, 'unidad': 'bol', 'gramos': 150, 'cocido': False},
        {'nombre': 'Sandía', 'proteina': 0.6, 'carbos': 8, 'grasas': 0.2, 'unidad': 'rodaja', 'gramos': 200, 'cocido': False},
        {'nombre': 'Melón', 'proteina': 0.8, 'carbos': 8, 'grasas': 0.2, 'unidad': 'rodaja', 'gramos': 200, 'cocido': False},
        {'nombre': 'Piña', 'proteina': 0.5, 'carbos': 13, 'grasas': 0.1, 'unidad': 'rodaja', 'gramos': 150, 'cocido': False},
        {'nombre': 'Uvas', 'proteina': 0.7, 'carbos': 18, 'grasas': 0.2, 'unidad': 'racimo pequeño', 'gramos': 100, 'cocido': False},
        {'nombre': 'Cerezas', 'proteina': 1, 'carbos': 16, 'grasas': 0.2, 'unidad': 'bol', 'gramos': 100, 'cocido': False},
        {'nombre': 'Melocotón', 'proteina': 0.6, 'carbos': 13, 'grasas': 0.2, 'unidad': 'melocotón', 'gramos': 150, 'cocido': False},
        {'nombre': 'Higos', 'proteina': 0.8, 'carbos': 19, 'grasas': 0.3, 'unidad': 'higo', 'gramos': 100, 'cocido': False},
    ],
    'Frutas deshidratadas': [
        {'nombre': 'Dátiles', 'proteina': 2, 'carbos': 75, 'grasas': 0.2, 'unidad': 'dátil', 'gramos': 15, 'cocido': False},
        {'nombre': 'Pasas', 'proteina': 3, 'carbos': 79, 'grasas': 0.5, 'unidad': 'puñado', 'gramos': 35, 'cocido': False},
        {'nombre': 'Ciruelas pasas', 'proteina': 2.2, 'carbos': 63, 'grasas': 0.4, 'unidad': 'ciruela', 'gramos': 15, 'cocido': False},
    ],
    'Azúcares simples': [
        {'nombre': 'Miel', 'proteina': 0.3, 'carbos': 82, 'grasas': 0, 'unidad': 'cucharada', 'gramos': 21, 'cocido': False},
        {'nombre': 'Mermelada', 'proteina': 0.4, 'carbos': 70, 'grasas': 0.1, 'unidad': 'cucharada', 'gramos': 20, 'cocido': False},
        {'nombre': 'Azúcar', 'proteina': 0, 'carbos': 100, 'grasas': 0, 'unidad': 'cucharadita', 'gramos': 10, 'cocido': False},
    ]
}

ALIMENTOS_GRASAS = {
    'Aceites': [
        {'nombre': 'Aceite de oliva virgen extra', 'proteina': 0, 'carbos': 0, 'grasas': 100, 'unidad': 'cucharada', 'gramos': 14},
        {'nombre': 'Aceite de coco', 'proteina': 0, 'carbos': 0, 'grasas': 100, 'unidad': 'cucharada', 'gramos': 14},
    ],
    'Frutos secos': [
        {'nombre': 'Almendras', 'proteina': 21, 'carbos': 22, 'grasas': 50, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Nueces', 'proteina': 15, 'carbos': 14, 'grasas': 65, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Anacardos', 'proteina': 18, 'carbos': 30, 'grasas': 44, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Avellanas', 'proteina': 15, 'carbos': 17, 'grasas': 61, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Pistachos', 'proteina': 20, 'carbos': 28, 'grasas': 45, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Cacahuetes', 'proteina': 26, 'carbos': 16, 'grasas': 49, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Pecanas', 'proteina': 9, 'carbos': 14, 'grasas': 72, 'unidad': 'puñado', 'gramos': 30},
    ],
    'Semillas': [
        {'nombre': 'Semillas de chía', 'proteina': 17, 'carbos': 42, 'grasas': 31, 'unidad': 'cucharada', 'gramos': 15},
        {'nombre': 'Semillas de lino', 'proteina': 18, 'carbos': 29, 'grasas': 42, 'unidad': 'cucharada', 'gramos': 15},
        {'nombre': 'Semillas de calabaza', 'proteina': 30, 'carbos': 14, 'grasas': 49, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Semillas de girasol', 'proteina': 21, 'carbos': 20, 'grasas': 51, 'unidad': 'puñado', 'gramos': 30},
        {'nombre': 'Semillas de sésamo', 'proteina': 18, 'carbos': 23, 'grasas': 50, 'unidad': 'cucharada', 'gramos': 15},
        {'nombre': 'Semillas de amapola', 'proteina': 17, 'carbos': 27, 'grasas': 42, 'unidad': 'cucharada', 'gramos': 15},
    ],
    'Aguacates y aceitunas': [
        {'nombre': 'Aguacate', 'proteina': 2, 'carbos': 9, 'grasas': 15, 'unidad': 'aguacate', 'gramos': 200},
        {'nombre': 'Aceitunas negras', 'proteina': 1, 'carbos': 6, 'grasas': 11, 'unidad': 'aceituna', 'gramos': 5},
        {'nombre': 'Aceitunas verdes', 'proteina': 0.8, 'carbos': 3.8, 'grasas': 10, 'unidad': 'aceituna', 'gramos': 5},
    ],
    'Cremas de frutos secos': [
        {'nombre': 'Mantequilla de cacahuete', 'proteina': 25, 'carbos': 20, 'grasas': 50, 'unidad': 'cucharada', 'gramos': 20},
        {'nombre': 'Crema de almendras', 'proteina': 21, 'carbos': 19, 'grasas': 55, 'unidad': 'cucharada', 'gramos': 20},
        {'nombre': 'Tahini (pasta de sésamo)', 'proteina': 17, 'carbos': 21, 'grasas': 54, 'unidad': 'cucharada', 'gramos': 20},
        {'nombre': 'Crema de anacardos', 'proteina': 18, 'carbos': 30, 'grasas': 44, 'unidad': 'cucharada', 'gramos': 20},
        {'nombre': 'Crema de avellanas', 'proteina': 15, 'carbos': 17, 'grasas': 61, 'unidad': 'cucharada', 'gramos': 20},
    ],
    'Lácteos grasos': [
        {'nombre': 'Queso curado', 'proteina': 25, 'carbos': 1.3, 'grasas': 33, 'unidad': 'loncha', 'gramos': 30},
        {'nombre': 'Queso crema', 'proteina': 6, 'carbos': 4, 'grasas': 33, 'unidad': 'cucharada', 'gramos': 30},
        {'nombre': 'Mantequilla', 'proteina': 0.9, 'carbos': 0.1, 'grasas': 81, 'unidad': 'cucharada', 'gramos': 10},
        {'nombre': 'Nata líquida', 'proteina': 2.1, 'carbos': 3.3, 'grasas': 19, 'unidad': 'cucharada', 'gramos': 15},
    ],
    'Otros': [
        {'nombre': 'Chocolate negro >85%', 'proteina': 8, 'carbos': 20, 'grasas': 52, 'unidad': 'onza', 'gramos': 10},
        {'nombre': 'Chocolate negro 70-85%', 'proteina': 7.5, 'carbos': 30, 'grasas': 43, 'unidad': 'onza', 'gramos': 10},
        {'nombre': 'Coco rallado', 'proteina': 3.3, 'carbos': 6.2, 'grasas': 65, 'unidad': 'cucharada', 'gramos': 10},
        {'nombre': 'Leche de coco', 'proteina': 1.9, 'carbos': 5, 'grasas': 20, 'unidad': 'vaso', 'gramos': 200},
    ]
}

VERDURAS_VEGETALES = {
    'Verduras ricas en vitamina C': [
        {'nombre': 'Pimiento rojo', 'proteina': 1, 'carbos': 6, 'grasas': 0.3, 'vitamina_c': 'alta'},
        {'nombre': 'Brócoli', 'proteina': 3, 'carbos': 7, 'grasas': 0.4, 'vitamina_c': 'alta'},
        {'nombre': 'Tomate', 'proteina': 1, 'carbos': 4, 'grasas': 0.2, 'vitamina_c': 'media'},
    ],
    'Verduras de hoja verde': [
        {'nombre': 'Espinacas', 'proteina': 3, 'carbos': 4, 'grasas': 0.4, 'hierro': 'alta'},
        {'nombre': 'Lechuga', 'proteina': 1, 'carbos': 3, 'grasas': 0.2},
        {'nombre': 'Acelgas', 'proteina': 2, 'carbos': 4, 'grasas': 0.2},
    ],
    'Verduras crucíferas': [
        {'nombre': 'Coliflor', 'proteina': 2, 'carbos': 5, 'grasas': 0.3},
        {'nombre': 'Calabacín', 'proteina': 1.2, 'carbos': 3, 'grasas': 0.3},
    ]
}


# Base de datos de micronutrientes ampliada y actualizada
MICRONUTRIENTES_DB = {
    'Vitamina A': {
        'funcion': 'Visión, sistema inmune, salud de la piel',
        'fuentes': ['Zanahorias', 'Boniato', 'Espinacas', 'Hígado', 'Huevos', 'Lácteos'],
        'RDA_hombre': '900 mcg/día',
        'RDA_mujer': '700 mcg/día'
    },
    'Vitamina D': {
        'funcion': 'Absorción de calcio, salud ósea, función inmune',
        'fuentes': ['Salmón', 'Atún', 'Yema de huevo', 'Lácteos fortificados', 'Exposición solar'],
        'RDA_hombre': '15-20 mcg/día',
        'RDA_mujer': '15-20 mcg/día'
    },
    'Vitamina E': {
        'funcion': 'Antioxidante, protección celular',
        'fuentes': ['Almendras', 'Semillas de girasol', 'Aguacate', 'Espinacas', 'Aceite de oliva'],
        'RDA_hombre': '15 mg/día',
        'RDA_mujer': '15 mg/día'
    },
    'Vitamina K': {
        'funcion': 'Coagulación sanguínea, salud ósea',
        'fuentes': ['Espinacas', 'Brócoli', 'Col rizada', 'Perejil', 'Lechuga'],
        'RDA_hombre': '120 mcg/día',
        'RDA_mujer': '90 mcg/día'
    },
    'Vitamina C': {
        'funcion': 'Antioxidante, síntesis de colágeno, absorción de hierro',
        'fuentes': ['Naranja', 'Kiwi', 'Fresas', 'Pimiento rojo', 'Brócoli', 'Tomate'],
        'RDA_hombre': '90 mg/día',
        'RDA_mujer': '75 mg/día'
    },
    'Vitamina B1 (Tiamina)': {
        'funcion': 'Metabolismo energético, función nerviosa',
        'fuentes': ['Carne de cerdo', 'Semillas de girasol', 'Legumbres', 'Cereales integrales'],
        'RDA_hombre': '1.2 mg/día',
        'RDA_mujer': '1.1 mg/día'
    },
    'Vitamina B2 (Riboflavina)': {
        'funcion': 'Metabolismo energético, salud ocular',
        'fuentes': ['Lácteos', 'Huevos', 'Carne', 'Almendras', 'Espinacas'],
        'RDA_hombre': '1.3 mg/día',
        'RDA_mujer': '1.1 mg/día'
    },
    'Vitamina B3 (Niacina)': {
        'funcion': 'Metabolismo energético, salud de la piel',
        'fuentes': ['Pollo', 'Atún', 'Salmón', 'Cacahuetes', 'Legumbres'],
        'RDA_hombre': '16 mg/día',
        'RDA_mujer': '14 mg/día'
    },
    'Vitamina B6': {
        'funcion': 'Metabolismo de proteínas, función cerebral',
        'fuentes': ['Pollo', 'Salmón', 'Patatas', 'Plátanos', 'Garbanzos'],
        'RDA_hombre': '1.3-1.7 mg/día',
        'RDA_mujer': '1.3-1.5 mg/día'
    },
    'Vitamina B12': {
        'funcion': 'Formación de glóbulos rojos, función nerviosa',
        'fuentes': ['Carne', 'Pescado', 'Huevos', 'Lácteos', 'Alimentos fortificados'],
        'RDA_hombre': '2.4 mcg/día',
        'RDA_mujer': '2.4 mcg/día'
    },
    'Ácido Fólico (B9)': {
        'funcion': 'Síntesis de ADN, formación de glóbulos rojos',
        'fuentes': ['Espinacas', 'Brócoli', 'Legumbres', 'Espárragos', 'Aguacate'],
        'RDA_hombre': '400 mcg/día',
        'RDA_mujer': '400 mcg/día (600 embarazo)'
    },
    'Colina': {
        'funcion': 'Síntesis de neurotransmisores, función hepática, metabolismo lipídico',
        'fuentes': ['Huevos', 'Hígado', 'Salmón', 'Pollo', 'Brócoli'],
        'RDA_hombre': '550 mg/día',
        'RDA_mujer': '425 mg/día'
    },
    'Calcio': {
        'funcion': 'Salud ósea, contracción muscular, función nerviosa',
        'fuentes': ['Lácteos', 'Sardinas con espinas', 'Tofu', 'Almendras', 'Brócoli'],
        'RDA_hombre': '1000-1200 mg/día',
        'RDA_mujer': '1000-1200 mg/día'
    },
    'Hierro': {
        'funcion': 'Transporte de oxígeno, producción energética',
        'fuentes': ['Carne roja', 'Hígado', 'Espinacas', 'Lentejas', 'Tofu'],
        'RDA_hombre': '8 mg/día',
        'RDA_mujer': '18 mg/día (8 post-menopausia)'
    },
    'Magnesio': {
        'funcion': 'Función muscular y nerviosa, producción energética',
        'fuentes': ['Espinacas', 'Almendras', 'Aguacate', 'Plátano', 'Chocolate negro'],
        'RDA_hombre': '400-420 mg/día',
        'RDA_mujer': '310-320 mg/día'
    },
    'Zinc': {
        'funcion': 'Sistema inmune, cicatrización, síntesis proteica',
        'fuentes': ['Carne roja', 'Ostras', 'Legumbres', 'Semillas de calabaza', 'Frutos secos'],
        'RDA_hombre': '11 mg/día',
        'RDA_mujer': '8 mg/día'
    },
    'Potasio': {
        'funcion': 'Equilibrio de fluidos, función muscular y nerviosa',
        'fuentes': ['Plátano', 'Patata', 'Aguacate', 'Espinacas', 'Salmón'],
        'RDA_hombre': '3400 mg/día',
        'RDA_mujer': '2600 mg/día'
    },
    'Sodio': {
        'funcion': 'Equilibrio de fluidos, transmisión nerviosa',
        'fuentes': ['Sal de mesa', 'Alimentos procesados', 'Queso', 'Pan'],
        'RDA_hombre': '1500 mg/día (máx 2300)',
        'RDA_mujer': '1500 mg/día (máx 2300)'
    },
    'Fósforo': {
        'funcion': 'Salud ósea, producción energética (ATP), membranas celulares',
        'fuentes': ['Lácteos', 'Carne', 'Pescado', 'Legumbres', 'Frutos secos'],
        'RDA_hombre': '700 mg/día',
        'RDA_mujer': '700 mg/día'
    },
    'Yodo': {
        'funcion': 'Función tiroidea, metabolismo',
        'fuentes': ['Sal yodada', 'Pescado', 'Lácteos', 'Huevos'],
        'RDA_hombre': '150 mcg/día',
        'RDA_mujer': '150 mcg/día'
    },
    'Selenio': {
        'funcion': 'Antioxidante, función tiroidea',
        'fuentes': ['Nueces de Brasil', 'Atún', 'Huevos', 'Pollo'],
        'RDA_hombre': '55 mcg/día',
        'RDA_mujer': '55 mcg/día'
    },
    'Cobre': {
        'funcion': 'Formación de glóbulos rojos, absorción de hierro, función inmune',
        'fuentes': ['Hígado', 'Ostras', 'Semillas de sésamo', 'Anacardos', 'Cacao'],
        'RDA_hombre': '900 mcg/día',
        'RDA_mujer': '900 mcg/día'
    },
    'Manganeso': {
        'funcion': 'Metabolismo de carbohidratos, formación ósea, antioxidante',
        'fuentes': ['Avena', 'Frutos secos', 'Té', 'Legumbres', 'Piña'],
        'RDA_hombre': '2.3 mg/día',
        'RDA_mujer': '1.8 mg/día'
    },
    'Cromo': {
        'funcion': 'Metabolismo de glucosa e insulina',
        'fuentes': ['Brócoli', 'Judías verdes', 'Carne', 'Cereales integrales'],
        'RDA_hombre': '35 mcg/día',
        'RDA_mujer': '25 mcg/día'
    }
}

# ============================================================================
# INFORMACIÓN ADICIONAL SOBRE MICRONUTRIENTES
# ============================================================================

NOTAS_MICRONUTRIENTES = {
    
    # 1. INTERACCIONES NUTRIENTE-NUTRIENTE
    'interacciones': {
        'descripcion': 'Combinaciones de nutrientes que potencian o reducen su absorción',
        'sinergicas': [
            {
                'combinacion': 'Vitamina C + Hierro no-hemo',
                'efecto': 'La vitamina C aumenta la absorción de hierro no-hemo (vegetales) hasta 3-4 veces',
                'ejemplo': 'Lentejas con pimiento rojo, espinacas con limón'
            },
            {
                'combinacion': 'Vitamina D + Calcio + Vitamina K2 + Magnesio',
                'efecto': 'Sinergia para salud ósea. K2 dirige el calcio a los huesos (no arterias). Se ha observado que una insuficiencia de magnesio puede desencadenar niveles bajos de vitamina D (no se produce correcta síntesis)',
                'ejemplo': 'Lácteos + exposición solar + verduras de hoja verde'
            },
            {
                'combinacion': 'Vitamina E + Grasas',
                'efecto': 'La vitamina E es liposoluble, requiere grasa para su absorción',
                'ejemplo': 'Almendras, aguacate, aceite de oliva'
            },
            {
                'combinacion': 'Vitaminas A, D, E, K + Grasas',
                'efecto': 'Todas las vitaminas liposolubles requieren grasa para absorción óptima',
                'ejemplo': 'Ensalada con aceite de oliva, zanahoria con frutos secos'
            },
            {
                'combinacion': 'Proteína + Leucina',
                'efecto': 'La leucina potencia la síntesis proteica muscular (3g por comida)',
                'ejemplo': 'Presentes en lácteos, carne, huevos'
            },
        ],
        'antagonicas': [
            {
                'combinacion': 'Calcio + Hierro',
                'efecto': 'Compiten por absorción. Calcio reduce absorción de hierro hasta 50-60%',
                'recomendacion': 'Separar lácteos de carnes rojas/suplementos de hierro 2-3 horas'
            },
            {
                'combinacion': 'Fitatos + Hierro/Zinc/Calcio',
                'efecto': 'Fitatos (cereales integrales, legumbres) se unen a minerales reduciendo absorción',
                'solucion': 'Remojar legumbres, fermentar cereales, germinar semillas mejora biodisponibilidad'
            },
            {
                'combinacion': 'Taninos (té, café) + Hierro',
                'efecto': 'Taninos reducen absorción de hierro no-hemo hasta 60-70%',
                'recomendacion': 'No tomar té/café en comidas ricas en hierro. Esperar 1-2 horas'
            },
            {
                'combinacion': 'Oxalatos + Calcio',
                'efecto': 'Oxalatos (espinacas, acelgas) se unen al calcio impidiendo absorción',
                'nota': 'Las espinacas tienen calcio pero baja biodisponibilidad. Preferir brócoli/lácteos'
            },
            {
                'combinacion': 'Zinc + Cobre',
                'efecto': 'Exceso de zinc (>40mg/día) interfiere con absorción de cobre',
                'recomendacion': 'Evitar suplementación prolongada alta en zinc sin supervisión o combinar suplementos de Zinc + Cobre'
            },
            {
                'combinacion': 'Alcohol + Vitaminas B (especialmente B1, B9, B12)',
                'efecto': 'El alcohol reduce absorción y aumenta excreción de vitaminas B',
                'nota': 'El consumo crónico puede causar deficiencias severas'
            },
        ]
    },
    
    # 2. BIODISPONIBILIDAD Y ABSORCIÓN
    'biodisponibilidad': {
        'descripcion': 'Porcentaje de nutriente absorbido y utilizado por el organismo',
        'proteinas': [
            {'alimento': 'Huevo', 'digestibilidad': 97, 'nota': 'Proteína de mayor valor biológico'},
            {'alimento': 'Suero de leche (whey)', 'digestibilidad': 95, 'nota': 'Absorción rápida, rico en leucina'},
            {'alimento': 'Pescado', 'digestibilidad': 94, 'nota': 'Alta digestibilidad'},
            {'alimento': 'Carne vacuna', 'digestibilidad': 92, 'nota': 'Fuente completa de aminoácidos'},
            {'alimento': 'Pollo', 'digestibilidad': 92, 'nota': 'Proteína magra de alta calidad'},
            {'alimento': 'Soja', 'digestibilidad': 91, 'nota': 'Mejor proteína vegetal completa'},
            {'alimento': 'Garbanzos', 'digestibilidad': 78, 'nota': 'Mejorar combinando con cereales'},
            {'alimento': 'Lentejas', 'digestibilidad': 76, 'nota': 'Remojar 8-12h mejora digestibilidad'},
        ],
        'hierro': [
            {
                'tipo': 'Hierro hemo (carne, pescado)',
                'absorcion': '15-35%',
                'nota': 'No afectado por fitatos/taninos. Mejor fuente'
            },
            {
                'tipo': 'Hierro no-hemo (vegetales, legumbres)',
                'absorcion': '2-20%',
                'nota': 'Muy afectado por inhibidores. Combinar con vitamina C'
            },
        ],
        'calcio': [
            {
                'fuente': 'Lácteos',
                'absorcion': '30-35%',
                'nota': 'Mejor fuente. Lactosa facilita absorción'
            },
            {
                'fuente': 'Brócoli, col rizada',
                'absorcion': '40-60%',
                'nota': 'Alta biodisponibilidad pero menor contenido total'
            },
            {
                'fuente': 'Espinacas',
                'absorcion': '5%',
                'nota': 'Baja por oxalatos. No es buena fuente de calcio'
            },
            {
                'fuente': 'Sardinas con espinas',
                'absorcion': '30%',
                'nota': 'Excelente fuente alternativa'
            },
        ],
        'omega3': [
            {
                'tipo': 'EPA/DHA (pescado azul)',
                'biodisponibilidad': 'Directa',
                'nota': 'Forma activa, lista para usar. 2-3 raciones/semana'
            },
            {
                'tipo': 'ALA (nueces, lino, chía)',
                'conversion': '<5-10%',
                'nota': 'Conversión a EPA/DHA muy limitada. No sustituye al pescado'
            },
        ],
        'factores_mejoran_absorcion': [
            'Vitamina D: necesaria para absorción de calcio',
            'Grasa: mejora absorción de vitaminas liposolubles (A, D, E, K)',
            'Vitamina C: triplica absorción de hierro no-hemo',
            'Probióticos/fermentados: mejoran salud intestinal y absorción de nutrientes',
        ],
        'factores_reducen_absorcion': [
            'Fitatos (cereales integrales, legumbres sin remojar)',
            'Oxalatos (espinacas, acelgas, ruibarbo)',
            'Taninos (té negro, café, vino tinto)',
            'Exceso de fibra (>50g/día) puede reducir absorción mineral',
            'Medicamentos: antiácidos (minerales), antibióticos (probióticos)',
            'Alcohol: reduce absorción de tiamina, folato, B12',
        ]
    },
    
    # 3. DEFICIENCIAS MÁS COMUNES EN POBLACIÓN ACTUAL (2024-2025)
    'deficiencias_comunes': {
        'descripcion': 'Micronutrientes con mayor prevalencia de déficit en población occidental',
        'criticos': [
            {
                'nutriente': 'Vitamina D',
                'prevalencia': '40-50% de adultos con niveles subóptimos',
                'grupos_riesgo': ['Vida en interior/oficina', 'Latitudes norte', 'Piel oscura', '>65 años'],
                'consecuencias': 'Salud ósea, sistema inmune debilitado, mayor riesgo cardiovascular',
                'solucion': 'Exposición solar 15-20 min/día, pescado azul, suplementación (1000-2000 UI/día)'
            },
            {
                'nutriente': 'Hierro',
                'prevalencia': '10-20% mujeres edad reproductiva con anemia',
                'grupos_riesgo': ['Mujeres premenopáusicas', 'Vegetarianos/veganos', 'Deportistas de resistencia'],
                'consecuencias': 'Fatiga crónica, rendimiento físico reducido, función cognitiva afectada',
                'solucion': 'Combinar hierro vegetal con vitamina C, suplementar si analítica baja'
            },
            {
                'nutriente': 'Magnesio',
                'prevalencia': '50-60% no alcanza requerimientos diarios',
                'grupos_riesgo': ['Dietas procesadas', 'Estrés crónico', 'Deportistas', 'Diabéticos'],
                'consecuencias': 'Calambres musculares, insomnio, fatiga, ansiedad, arritmias',
                'solucion': 'Frutos secos, espinacas, chocolate negro, aguacate, legumbres, suplementos: citrato de magnesio (buena biodisponibilidad), bisglicinato de magnesio (excelente biodisponibilidad)'
            },
            {
                'nutriente': 'Omega-3 (EPA/DHA)',
                'prevalencia': '70-80% con índice omega-3 bajo (<4%)',
                'grupos_riesgo': ['Bajo consumo de pescado', 'Dieta occidental típica'],
                'consecuencias': 'Mayor riesgo cardiovascular, inflamación crónica, salud cerebral',
                'solucion': 'Pescado azul 2-3x/semana (salmón, sardinas, caballa) o suplementación'
            },
        ],
        'moderados': [
            {
                'nutriente': 'Vitamina E',
                'prevalencia': '85-90% ingesta inadecuada',
                'nota': 'Deficiencia clínica rara pero ingesta subóptima común'
            },
            {
                'nutriente': 'Yodo',
                'prevalencia': '30-40% ingesta subóptima',
                'nota': 'Usar sal yodada. Déficit leve puede afectar función tiroidea'
            },
            {
                'nutriente': 'Vitamina C',
                'prevalencia': '40-50% ingesta inadecuada',
                'nota': 'Común en dietas bajas en frutas/verduras frescas'
            },
            {
                'nutriente': 'Folato (B9)',
                'prevalencia': '20-30% (mejorado por fortificación)',
                'nota': 'Crítico en mujeres embarazo/planificación. Fortificación obligatoria en muchos países'
            },
            {
                'nutriente': 'Vitamina B12',
                'prevalencia': '10-15% >60 años, 50-90% veganos',
                'nota': 'Absorción disminuye con edad. Veganos deben suplementar obligatoriamente'
            },
            {
                'nutriente': 'Zinc',
                'prevalencia': '15-20% ingesta inadecuada',
                'nota': 'Importante para sistema inmune. Vegetarianos en mayor riesgo'
            },
        ]
    },
    
    # 4. ANALÍTICAS: QUÉ VALORES SOLICITAR Y RANGOS ÓPTIMOS
    'analitica_completa': {
        'descripcion': 'Marcadores en sangre para evaluar estado nutricional completo',
        'panel_basico': [
            {
                'marcador': 'Hemograma completo',
                'evalua': 'Anemia (hierro, B12, folato)',
                'frecuencia': 'Anual o si síntomas de fatiga'
            },
            {
                'marcador': 'Glucosa en ayunas',
                'rango_optimo': '70-100 mg/dL',
                'nota': 'Pre-diabetes: 100-125 mg/dL. Diabetes: ≥126 mg/dL'
            },
            {
                'marcador': 'HbA1c (hemoglobina glicosilada)',
                'rango_optimo': '<5.7%',
                'nota': 'Refleja control glucémico últimos 3 meses. Pre-diabetes: 5.7-6.4%'
            },
        ],
        'vitaminas': [
            {
                'vitamina': 'Vitamina D - 25(OH)D',
                'rango_deficiencia': '<20 ng/mL (<50 nmol/L)',
                'rango_insuficiente': '20-30 ng/mL (50-75 nmol/L)',
                'rango_optimo': '30-50 ng/mL (75-125 nmol/L)',
                'rango_toxicidad': '>100 ng/mL (>250 nmol/L)',
                'notas': 'Solicitar específicamente 25(OH)D, no 1,25(OH)2D. Déficit muy común'
            },
            {
                'vitamina': 'Vitamina B12',
                'rango_deficiencia': '<200 pg/mL (<148 pmol/L)',
                'rango_optimo': '>400 pg/mL (>295 pmol/L)',
                'notas': 'Síntomas pueden aparecer con valores "normales-bajos". Solicitar también homocisteína y ácido metilmalónico si sospecha'
            },
            {
                'vitamina': 'Folato (B9) sérico',
                'rango_deficiencia': '<3 ng/mL (<6.8 nmol/L)',
                'rango_optimo': '5-25 ng/mL (11-57 nmol/L)',
                'notas': 'Importante en mujeres edad fértil. Déficit causa anemia megaloblástica'
            },
            {
                'vitamina': 'Vitamina A',
                'rango_optimo': '32.5-78 μg/dL (1.1-2.7 μmol/L)',
                'notas': 'Rara vez deficiente en países desarrollados salvo malabsorción'
            },
            {
                'vitamina': 'Vitamina E',
                'rango_optimo': '5.5-17 mg/L',
                'notas': 'Deficiencia clínica rara. Solicitar si sospecha malabsorción grasas'
            },
        ],
        'minerales': [
            {
                'mineral': 'Hierro sérico',
                'rango_optimo_hombre': '60-170 μg/dL',
                'rango_optimo_mujer': '50-150 μg/dL',
                'nota': 'Varía durante el día. Solicitar en ayunas'
            },
            {
                'mineral': 'Ferritina',
                'rango_deficiencia': '<30 ng/mL',
                'rango_optimo_hombre': '50-300 ng/mL',
                'rango_optimo_mujer': '30-200 ng/mL',
                'nota': 'MEJOR indicador de reservas de hierro. Valores <30 indican depósitos agotados aunque no haya anemia aún'
            },
            {
                'mineral': 'Saturación de transferrina',
                'rango_optimo': '20-45%',
                'rango_deficiencia': '<20%',
                'nota': 'Complementa estudio del hierro. <15% sugiere déficit'
            },
            {
                'mineral': 'Magnesio sérico',
                'rango_optimo': '1.7-2.2 mg/dL (0.7-0.9 mmol/L)',
                'nota': 'Poco sensible (solo 1% Mg está en sangre). Déficit intracelular puede existir con valores normales'
            },
            {
                'mineral': 'Magnesio eritrocitario',
                'nota': 'Mejor reflejo de magnesio intracelular pero menos disponible'
            },
            {
                'mineral': 'Calcio sérico',
                'rango_optimo': '8.9-10.1 mg/dL (2.2-2.5 mmol/L)',
                'nota': 'Fuertemente regulado. Niveles bajos raros salvo enfermedad'
            },
            {
                'mineral': 'Zinc sérico',
                'rango_optimo': '70-120 μg/dL',
                'nota': 'Déficit afecta inmunidad, cicatrización, sentido del gusto'
            },
            {
                'mineral': 'Selenio',
                'rango_optimo': '70-150 ng/mL',
                'nota': 'Importante para función tiroidea y antioxidante'
            },
        ],
        'marcadores_funcionales': [
            {
                'marcador': 'Homocisteína',
                'rango_optimo': '<10 μmol/L',
                'evalua': 'Metabolismo B12, B6, folato. Valores altos: riesgo cardiovascular',
                'nota': 'Aumenta con déficit de vitaminas B. Objetivo: <8 μmol/L'
            },
            {
                'marcador': 'Índice Omega-3',
                'rango_optimo': '>8%',
                'rango_moderado': '4-8%',
                'rango_bajo': '<4%',
                'nota': 'Mide EPA+DHA en membrana eritrocitos. Refleja consumo pescado/suplementos'
            },
            {
                'marcador': 'Proteína C Reactiva (PCR-us)',
                'rango_optimo': '<1 mg/L',
                'nota': 'Marcador inflamación. Omega-3, antioxidantes, magnesio la reducen'
            },
            {
                'marcador': 'TSH (hormona estimulante tiroides)',
                'rango_optimo': '0.5-2.5 mUI/L',
                'nota': 'Evalúa función tiroidea. Requiere yodo, selenio, zinc adecuados'
            },
            {
                'marcador': 'T3 libre y T4 libre',
                'nota': 'Solicitar si TSH alterada. Conversión T4→T3 requiere selenio, zinc, hierro'
            },
        ],
        'frecuencia_recomendada': {
            'poblacion_general_sana': 'Analítica completa cada 1-2 años',
            'vegetarianos_veganos': 'B12, hierro, ferritina, zinc cada 6-12 meses',
            'deportistas': 'Hierro, ferritina, magnesio, vitamina D cada 6-12 meses',
            'mayores_60años': 'Anual: B12, vitamina D, calcio, función renal',
            'mujeres_edad_fertil': 'Hierro, ferritina anual. Folato si planificación embarazo',
            'con_sintomas': 'Según sospecha clínica: fatiga→hierro/B12/D; calambres→magnesio/potasio',
        },
        'cuando_solicitar': [
            'Fatiga crónica sin causa aparente → hierro, ferritina, B12, vitamina D, TSH',
            'Calambres musculares frecuentes → magnesio, potasio, calcio, vitamina D',
            'Dieta vegetariana/vegana → B12, hierro, ferritina, zinc, vitamina D',
            'Pérdida cabello excesiva → hierro, ferritina, zinc, biotina, proteína total',
            'Inmunidad baja (resfriados frecuentes) → vitamina D, zinc, vitamina C',
            'Cicatrización lenta → zinc, vitamina C, proteína total',
            'Entumecimiento/hormigueo → B12, ácido metilmalónico, homocisteína',
            'Depresión/ansiedad → vitamina D, magnesio, B12, folato, omega-3',
            'Atletas de alto rendimiento → panel completo incluyendo hierro, magnesio, vitamina D',
        ]
    },
    
    # 5. CONSIDERACIONES ESPECIALES
    'notas_importantes': [
        'Los rangos "normales" de laboratorio no siempre son "óptimos". Consulta con profesional',
        'La suplementación debe ser individualizada según analítica, no preventiva indiscriminada',
        'Megadosis de vitaminas pueden ser tóxicas (A, D, E) o interferir con otros nutrientes',
        'Los suplementos no sustituyen una dieta equilibrada, solo complementan déficits específicos',
        'La biodisponibilidad de suplementos varía: citrato de magnesio > óxido de magnesio',
    ]
}


# ============================================================================
# FUNCIONES DE CÁLCULO 
# ============================================================================

def calcular_tmb_mifflin(peso, altura, edad, sexo):
    if sexo == "Hombre":
        tmb = (10 * peso) + (6.25 * altura) - (5 * edad) + 5
    else:
        tmb = (10 * peso) + (6.25 * altura) - (5 * edad) - 161
    return round(tmb, 1)

def calcular_factor_actividad_detallado(trabajo, entreno_pesas_horas, cardio_dias, cardio_minutos, pasos_diarios):
    factor_base = {
        'Sedentario': 1.2,
        'Ligeramente activo': 1.3,
        'Moderadamente activo': 1.4,
        'Muy activo': 1.6,
        'Extremadamente activo': 1.8
    }
    
    factor = factor_base.get(trabajo, 1.2)
    
    if entreno_pesas_horas > 0:
        factor += entreno_pesas_horas * 0.025
    
    if cardio_dias > 0 and cardio_minutos > 0:
        sesiones_equiv = (cardio_dias * cardio_minutos) / 30
        factor += sesiones_equiv * 0.015
    
    if pasos_diarios >= 15000:
        factor += 0.1
    elif pasos_diarios >= 10000:
        factor += 0.05
    elif pasos_diarios >= 7000:
        factor += 0.025
    
    return round(factor, 3)

def calcular_tdee(tmb, factor_actividad):
    return round(tmb * factor_actividad, 0)

def calcular_calorias_objetivo(tdee, objetivo, intensidad):
    if objetivo == "Déficit":
        if intensidad == "Conservador":
            ajuste = -0.075
        elif intensidad == "Moderado":
            ajuste = -0.175
        else:
            ajuste = -0.25
    elif objetivo == "Mantenimiento":
        ajuste = 0
    else:  # Superávit
        if intensidad == "Conservador":
            ajuste = 0.075
        elif intensidad == "Moderado":
            ajuste = 0.125
        else:
            ajuste = 0.175
    
    calorias = tdee * (1 + ajuste)
    return round(calorias, 0)

def calcular_macros(calorias_objetivo, peso, sexo):
    if sexo == "Hombre":
        proteina_min = peso * 1.8
        proteina_max = peso * 2.2
    else:
        proteina_min = peso * 1.6
        proteina_max = peso * 1.8
    
    proteina = round((proteina_min + proteina_max) / 2, 1)
    grasas = round(peso * 0.9, 1)
    
    cal_proteina = proteina * 4
    cal_grasas = grasas * 9
    cal_carbos = calorias_objetivo - cal_proteina - cal_grasas
    carbos = round(cal_carbos / 4, 1)
    
    return {
        'proteina': proteina,
        'proteina_min': round(proteina_min, 1),
        'proteina_max': round(proteina_max, 1),
        'carbos': carbos,
        'grasas': grasas,
        'grasas_min': round(peso * 0.8, 1),
        'grasas_max': round(peso * 1.0, 1)
    }

def distribuir_macros_comidas(macros_totales, num_comidas):
    comidas = []
    for i in range(num_comidas):
        comidas.append({
            'nombre': f'Comida {i+1}',
            'proteina': round(macros_totales['proteina'] / num_comidas, 1),
            'carbos': round(macros_totales['carbos'] / num_comidas, 1),
            'grasas': round(macros_totales['grasas'] / num_comidas, 1)
        })
    
    # Ajustar última comida para que cuadre
    suma_prot = sum(c['proteina'] for c in comidas[:-1])
    suma_carbs = sum(c['carbos'] for c in comidas[:-1])
    suma_grasas = sum(c['grasas'] for c in comidas[:-1])
    
    comidas[-1]['proteina'] = round(macros_totales['proteina'] - suma_prot, 1)
    comidas[-1]['carbos'] = round(macros_totales['carbos'] - suma_carbs, 1)
    comidas[-1]['grasas'] = round(macros_totales['grasas'] - suma_grasas, 1)
    
    return comidas

# ============================================================================
# INICIALIZACIÓN SESSION STATE
# ============================================================================

if 'plan' not in st.session_state:
    st.session_state.plan = {
        'cliente': '',
        'sexo': 'Hombre',
        'edad': 30,
        'peso': 75,
        'altura': 175,
        'objetivo': 'Mantenimiento',
        'intensidad': 'Moderado',
        'tmb': 0,
        'tdee': 0,
        'calorias': 0,
        'macros': {},
        'num_comidas': 3,
        'comidas': [],
        'calculado': False,
        'vista': 'calculadora'
    }

# ============================================================================
# NAVEGACIÓN SUPERIOR
# ============================================================================

col_nav1, col_nav2, col_nav3, col_nav4 = st.columns([1, 1, 1, 2])

with col_nav1:
    if st.button("🧮 Calculadora", use_container_width=True,
                 type="primary" if st.session_state.plan['vista'] == 'calculadora' else "secondary"):
        st.session_state.plan['vista'] = 'calculadora'
        st.rerun()

with col_nav2:
    if st.button("🍽️ Distribución", use_container_width=True,
                 type="primary" if st.session_state.plan['vista'] == 'distribucion' else "secondary"):
        if st.session_state.plan['calculado']:
            st.session_state.plan['vista'] = 'distribucion'
            st.rerun()
        else:
            st.warning("Primero calcula los macros")

with col_nav3:
    if st.button("📚 Base de Datos", use_container_width=True,
                 type="primary" if st.session_state.plan['vista'] == 'base_datos' else "secondary"):
        st.session_state.plan['vista'] = 'base_datos'
        st.rerun()

st.markdown("---")

# ============================================================================
# SIDEBAR: RESUMEN
# ============================================================================

with st.sidebar:
    st.header("📋 Resumen del Plan")
    
    if st.session_state.plan['cliente']:
        st.metric("Cliente", st.session_state.plan['cliente'])
    else:
        st.info("Sin nombre")
    
    if st.session_state.plan['calculado']:
        st.success("✅ Plan calculado")
        st.metric("Calorías", f"{st.session_state.plan['calorias']:.0f} kcal")
        st.metric("Proteína", f"{st.session_state.plan['macros']['proteina']:.1f}g")
        st.metric("Carbohidratos", f"{st.session_state.plan['macros']['carbos']:.1f}g")
        st.metric("Grasas", f"{st.session_state.plan['macros']['grasas']:.1f}g")
    else:
        st.info("ℹ️ Aún no calculado")

# ============================================================================
# VISTA: CALCULADORA (TODO EN UNA PÁGINA CON DESPLEGABLES)
# ============================================================================

if st.session_state.plan['vista'] == 'calculadora':
    st.title("🧮 Calculadora Nutricional Completa")
    
    # SECCIÓN 1: DATOS DEL CLIENTE
    with st.expander("👤 DATOS ", expanded=True):
        col1, col2, col3, col4, col5 = st.columns(5)
        
        with col1:
            st.session_state.plan['cliente'] = st.text_input("Nombre", st.session_state.plan['cliente'])
        
        with col2:
            st.session_state.plan['sexo'] = st.selectbox("Sexo", ["Hombre", "Mujer"])
        
        with col3:
            st.session_state.plan['edad'] = st.number_input("Edad", 15, 80, st.session_state.plan['edad'])
        
        with col4:
            st.session_state.plan['peso'] = st.number_input("Peso (kg)", 40, 200, st.session_state.plan['peso'])
        
        with col5:
            st.session_state.plan['altura'] = st.number_input("Altura (cm)", 140, 220, st.session_state.plan['altura'])
    
    # SECCIÓN 2: ACTIVIDAD FÍSICA
    with st.expander("⚡ ACTIVIDAD FÍSICA", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            st.session_state.plan['trabajo'] = st.selectbox(
                "Nivel de actividad diaria",
                ['Sedentario', 'Ligeramente activo', 'Moderadamente activo', 'Muy activo', 'Extremadamente activo']
            )
            
            st.session_state.plan['entreno_pesas'] = st.number_input(
                "Horas entrenamiento pesas/semana",
                0.0, 20.0, 4.0, 0.5
            )
        
        with col2:
            st.session_state.plan['pasos_diarios'] = st.number_input(
                "Pasos diarios promedio",
                0, 30000, 8000, 1000
            )
            
            col_cardio1, col_cardio2 = st.columns(2)
            with col_cardio1:
                st.session_state.plan['cardio_dias'] = st.number_input("Días cardio/sem", 0, 7, 0)
            with col_cardio2:
                st.session_state.plan['cardio_minutos'] = st.number_input("Minutos/sesión", 0, 120, 30)
    
    # SECCIÓN 3: OBJETIVO
    with st.expander("🎯 OBJETIVO NUTRICIONAL", expanded=True):
        col1, col2 = st.columns(2)
        
        with col1:
            st.session_state.plan['objetivo'] = st.selectbox(
                "Selecciona objetivo",
                ["Déficit", "Mantenimiento", "Superávit"]
            )
        
        with col2:
            if st.session_state.plan['objetivo'] != "Mantenimiento":
                st.session_state.plan['intensidad'] = st.selectbox(
                    "Intensidad",
                    ["Conservador", "Moderado", "Agresivo"]
                )
                
                if st.session_state.plan['objetivo'] == "Déficit":
                    porcentajes = {"Conservador": "5-10%", "Moderado": "15-20%", "Agresivo": "25%"}
                else:
                    porcentajes = {"Conservador": "5-10%", "Moderado": "10-15%", "Agresivo": "15-20%"}
                
                st.caption(f"📊 {porcentajes[st.session_state.plan['intensidad']]} de ajuste calórico")
            else:
                st.session_state.plan['intensidad'] = "Mantenimiento"
                st.info("Calorías = TDEE para recomposición")
    
    # SECCIÓN 4: RESULTADOS
    with st.expander("📊 RESULTADOS", expanded=True):
        if st.button("🔄 CALCULAR PLAN NUTRICIONAL", type="primary", use_container_width=True):
            # Calcular
            tmb = calcular_tmb_mifflin(
                st.session_state.plan['peso'],
                st.session_state.plan['altura'],
                st.session_state.plan['edad'],
                st.session_state.plan['sexo']
            )
            st.session_state.plan['tmb'] = tmb
            
            factor = calcular_factor_actividad_detallado(
                st.session_state.plan['trabajo'],
                st.session_state.plan['entreno_pesas'],
                st.session_state.plan['cardio_dias'],
                st.session_state.plan['cardio_minutos'],
                st.session_state.plan['pasos_diarios']
            )
            st.session_state.plan['factor_actividad'] = factor
            
            tdee = calcular_tdee(tmb, factor)
            st.session_state.plan['tdee'] = tdee
            
            calorias = calcular_calorias_objetivo(
                tdee,
                st.session_state.plan['objetivo'],
                st.session_state.plan['intensidad']
            )
            st.session_state.plan['calorias'] = calorias
            
            macros = calcular_macros(
                calorias,
                st.session_state.plan['peso'],
                st.session_state.plan['sexo']
            )
            st.session_state.plan['macros'] = macros
            
            comidas = distribuir_macros_comidas(macros, st.session_state.plan['num_comidas'])
            st.session_state.plan['comidas'] = comidas
            
            st.session_state.plan['calculado'] = True
            st.session_state.plan['fecha'] = datetime.now().strftime("%Y-%m-%d")
            
            st.success("✅ ¡Plan calculado con éxito!")
            st.rerun()
        
        if st.session_state.plan['calculado']:
            st.markdown("---")
            
            # Métricas principales
            col1, col2, col3 = st.columns(3)
            
            with col1:
                st.metric("🔥 TMB", f"{st.session_state.plan['tmb']:.0f} kcal")
                st.caption("Metabolismo basal")
            
            with col2:
                st.metric("⚡ Factor", f"{st.session_state.plan['factor_actividad']:.2f}x")
                st.caption("Multiplicador actividad")
            
            with col3:
                st.metric("📊 TDEE", f"{st.session_state.plan['tdee']:.0f} kcal")
                st.caption("Gasto total diario")
            
            st.markdown("---")
            
            # Macros
            col4, col5, col6, col7 = st.columns(4)
            
            with col4:
                st.metric("🎯 Calorías", f"{st.session_state.plan['calorias']:.0f} kcal")
            
            with col5:
                macros = st.session_state.plan['macros']
                st.metric("🥩 Proteína", f"{macros['proteina']:.1f}g")
                st.caption(f"{macros['proteina_min']:.1f}-{macros['proteina_max']:.1f}g")
            
            with col6:
                st.metric("🍚 Carbos", f"{macros['carbos']:.1f}g")
            
            with col7:
                st.metric("🥑 Grasas", f"{macros['grasas']:.1f}g")
                st.caption(f"{macros['grasas_min']:.1f}-{macros['grasas_max']:.1f}g")
            
            # Gráfico
            cal_prot = macros['proteina'] * 4
            cal_carbs = macros['carbos'] * 4
            cal_grasas = macros['grasas'] * 9
            
            fig = go.Figure(data=[go.Pie(
                labels=['Proteína', 'Carbohidratos', 'Grasas'],
                values=[cal_prot, cal_carbs, cal_grasas],
                hole=.4,
                marker_colors=['#FF6B6B', '#4ECDC4', '#FFE66D']
            )])
            
            fig.update_layout(height=350)
            st.plotly_chart(fig, use_container_width=True)

# ============================================================================
# VISTA: DISTRIBUCIÓN DE COMIDAS (SIN REDISTRIBUCIÓN AUTOMÁTICA)
# ============================================================================

elif st.session_state.plan['vista'] == 'distribucion':
    st.title("🍽️ Distribución de Macros por Comida")
    
    if not st.session_state.plan['calculado']:
        st.warning("⚠️ Primero calcula el plan en la Calculadora")
        st.stop()
    
    # Selector de comidas
    col1, col2 = st.columns([1, 3])
    
    with col1:
        nuevo_num_comidas = st.number_input("Número de comidas", 2, 6, st.session_state.plan['num_comidas'])
        
        if nuevo_num_comidas != st.session_state.plan['num_comidas']:
            st.session_state.plan['num_comidas'] = nuevo_num_comidas
            comidas = distribuir_macros_comidas(
                st.session_state.plan['macros'],
                nuevo_num_comidas
            )
            st.session_state.plan['comidas'] = comidas
            st.rerun()
    
    with col2:
        st.info("💡 Valores iniciales por defecto. Modifícalos libremente sin límites.")
    
    st.markdown("---")
    
    # Editar comidas SIN redistribución automática
    for i, comida in enumerate(st.session_state.plan['comidas']):
        with st.expander(f"🍽️ {comida['nombre']}", expanded=True):
            col1, col2, col3, col4 = st.columns(4)
            
            with col1:
                comida['proteina'] = st.number_input(
                    "Proteína (g)",
                    min_value=0.0,
                    max_value=500.0,
                    value=float(comida['proteina']),
                    step=0.5,
                    key=f"prot_{i}"
                )
            
            with col2:
                comida['carbos'] = st.number_input(
                    "Carbohidratos (g)",
                    min_value=0.0,
                    max_value=500.0,
                    value=float(comida['carbos']),
                    step=0.5,
                    key=f"carbs_{i}"
                )
            
            with col3:
                comida['grasas'] = st.number_input(
                    "Grasas (g)",
                    min_value=0.0,
                    max_value=200.0,
                    value=float(comida['grasas']),
                    step=0.5,
                    key=f"grasas_{i}"
                )
            
            with col4:
                calorias_comida = (comida['proteina'] * 4) + (comida['carbos'] * 4) + (comida['grasas'] * 9)
                st.metric("Calorías", f"{calorias_comida:.0f} kcal")
    
    st.markdown("---")
    
    # Resumen total con diferencias
    st.header("📊 Resumen Total")
    
    suma_prot = sum(c['proteina'] for c in st.session_state.plan['comidas'])
    suma_carbs = sum(c['carbos'] for c in st.session_state.plan['comidas'])
    suma_grasas = sum(c['grasas'] for c in st.session_state.plan['comidas'])
    suma_cals = (suma_prot * 4) + (suma_carbs * 4) + (suma_grasas * 9)
    
    col1, col2, col3, col4 = st.columns(4)
    
    objetivo_prot = st.session_state.plan['macros']['proteina']
    objetivo_carbs = st.session_state.plan['macros']['carbos']
    objetivo_grasas = st.session_state.plan['macros']['grasas']
    objetivo_cals = st.session_state.plan['calorias']
    
    with col1:
        diff_prot = suma_prot - objetivo_prot
        st.metric("Proteína Total", f"{suma_prot:.1f}g", 
                 delta=f"{diff_prot:+.1f}g del objetivo")
    
    with col2:
        diff_carbs = suma_carbs - objetivo_carbs
        st.metric("Carbos Total", f"{suma_carbs:.1f}g",
                 delta=f"{diff_carbs:+.1f}g del objetivo")
    
    with col3:
        diff_grasas = suma_grasas - objetivo_grasas
        st.metric("Grasas Total", f"{suma_grasas:.1f}g",
                 delta=f"{diff_grasas:+.1f}g del objetivo")
    
    with col4:
        diff_cals = suma_cals - objetivo_cals
        st.metric("Calorías Total", f"{suma_cals:.0f} kcal",
                 delta=f"{diff_cals:+.0f} kcal del objetivo")
    
    # Indicador de precisión
    if abs(diff_prot) < 1 and abs(diff_carbs) < 1 and abs(diff_grasas) < 1:
        st.success("✅ Los macros cuadran perfectamente con el objetivo!")
    elif abs(diff_prot) < 5 and abs(diff_carbs) < 5 and abs(diff_grasas) < 2:
        st.info("ℹ️ Los macros están muy cerca del objetivo")
    else:
        st.warning("⚠️ Ajusta las comidas para acercarte al objetivo")
    
    st.markdown("---")
    
    # CALCULADORA DE EQUIVALENCIAS AMPLIADA
    st.header("🔍 Calculadora de Equivalencias")
    
    tab_prot, tab_carbs, tab_grasas = st.tabs(["🥩 PROTEÍNAS", "🍚 CARBOHIDRATOS", "🥑 GRASAS"])
    
    # TAB PROTEÍNAS
    with tab_prot:
        col_input, col_result = st.columns([1, 1])
        
        with col_input:
            gramos_prot = st.number_input("Gramos de proteína objetivo", 10, 100, 30, key="calc_prot")
            categoria_prot = st.selectbox("Categoría", list(ALIMENTOS_PROTEINAS.keys()), key="cat_prot")
            alimento_prot_nombre = st.selectbox("Alimento", 
                                               [a['nombre'] for a in ALIMENTOS_PROTEINAS[categoria_prot]],
                                               key="alim_prot")
        
        with col_result:
            alimento = next(a for a in ALIMENTOS_PROTEINAS[categoria_prot] if a['nombre'] == alimento_prot_nombre)
            
            gramos_necesarios = (gramos_prot / alimento['proteina']) * 100
            unidades = gramos_necesarios / alimento['gramos']
            
            st.success(f"**Para {gramos_prot}g de proteína necesitas:**")
            st.metric("Gramos totales", f"{gramos_necesarios:.0f}g")
            st.markdown(f"### ≈ {unidades:.1f} × {alimento['unidad']}")
            
            # Macros totales de esa cantidad
            prot_total = (gramos_necesarios / 100) * alimento['proteina']
            carbs_total = (gramos_necesarios / 100) * alimento['carbos']
            grasas_total = (gramos_necesarios / 100) * alimento['grasas']
            cals_total = (prot_total * 4) + (carbs_total * 4) + (grasas_total * 9)
            
            st.caption(f"📊 Macros totales: P:{prot_total:.1f}g | C:{carbs_total:.1f}g | G:{grasas_total:.1f}g | {cals_total:.0f} kcal")
    
    # TAB CARBOHIDRATOS
    with tab_carbs:
        col_input2, col_result2 = st.columns([1, 1])
        
        with col_input2:
            gramos_carbs = st.number_input("Gramos de carbohidratos objetivo", 10, 150, 50, key="calc_carbs")
            categoria_carbs = st.selectbox("Categoría", list(ALIMENTOS_CARBOHIDRATOS.keys()), key="cat_carbs")
            alimento_carbs_nombre = st.selectbox("Alimento", 
                                                [a['nombre'] for a in ALIMENTOS_CARBOHIDRATOS[categoria_carbs]],
                                                key="alim_carbs")
        
        with col_result2:
            alimento_c = next(a for a in ALIMENTOS_CARBOHIDRATOS[categoria_carbs] if a['nombre'] == alimento_carbs_nombre)
            
            gramos_nec_c = (gramos_carbs / alimento_c['carbos']) * 100
            unidades_c = gramos_nec_c / alimento_c['gramos']
            
            st.success(f"**Para {gramos_carbs}g de carbohidratos necesitas:**")
            st.metric("Gramos totales", f"{gramos_nec_c:.0f}g")
            st.markdown(f"### ≈ {unidades_c:.1f} × {alimento_c['unidad']}")
            
            prot_c = (gramos_nec_c / 100) * alimento_c['proteina']
            carbs_c = (gramos_nec_c / 100) * alimento_c['carbos']
            grasas_c = (gramos_nec_c / 100) * alimento_c['grasas']
            cals_c = (prot_c * 4) + (carbs_c * 4) + (grasas_c * 9)
            
            st.caption(f"📊 Macros totales: P:{prot_c:.1f}g | C:{carbs_c:.1f}g | G:{grasas_c:.1f}g | {cals_c:.0f} kcal")
    
    # TAB GRASAS
    with tab_grasas:
        col_input3, col_result3 = st.columns([1, 1])
        
        with col_input3:
            gramos_grasas = st.number_input("Gramos de grasas objetivo", 5, 80, 20, key="calc_grasas")
            categoria_grasas = st.selectbox("Categoría", list(ALIMENTOS_GRASAS.keys()), key="cat_grasas")
            alimento_grasas_nombre = st.selectbox("Alimento", 
                                                  [a['nombre'] for a in ALIMENTOS_GRASAS[categoria_grasas]],
                                                  key="alim_grasas")
        
        with col_result3:
            alimento_g = next(a for a in ALIMENTOS_GRASAS[categoria_grasas] if a['nombre'] == alimento_grasas_nombre)
            
            gramos_nec_g = (gramos_grasas / alimento_g['grasas']) * 100
            unidades_g = gramos_nec_g / alimento_g['gramos']
            
            st.success(f"**Para {gramos_grasas}g de grasas necesitas:**")
            st.metric("Gramos totales", f"{gramos_nec_g:.0f}g")
            st.markdown(f"### ≈ {unidades_g:.1f} × {alimento_g['unidad']}")
            
            prot_g = (gramos_nec_g / 100) * alimento_g['proteina']
            carbs_g = (gramos_nec_g / 100) * alimento_g['carbos']
            grasas_g = (gramos_nec_g / 100) * alimento_g['grasas']
            cals_g = (prot_g * 4) + (carbs_g * 4) + (grasas_g * 9)
            
            st.caption(f"📊 Macros totales: P:{prot_g:.1f}g | C:{carbs_g:.1f}g | G:{grasas_g:.1f}g | {cals_g:.0f} kcal")
    
    st.markdown("---")
    
    # EXPORTACIÓN (mantener funciones PDF/Excel del código anterior)

    st.header("📤 Exportar Plan")
    
    col_btn1, col_btn2 = st.columns(2)
    
    with col_btn1:
        if st.button("📄 Generar PDF", use_container_width=True, type="primary"):
            with st.spinner("Generando PDF profesional..."):
                try:
                    pdf_buffer = generar_pdf_dieta(st.session_state.plan)
                    st.success("✅ PDF generado correctamente")
                    
                    st.download_button(
                        label="⬇️ Descargar PDF",
                        data=pdf_buffer,
                        file_name=f"Plan_Nutricional_{st.session_state.plan['cliente'].replace(' ', '_') if st.session_state.plan['cliente'] else 'Cliente'}_{datetime.now().strftime('%Y%m%d')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Error al generar PDF: {str(e)}")
    
    with col_btn2:
        if st.button("📊 Generar Excel", use_container_width=True, type="primary"):
            with st.spinner("Generando Excel profesional..."):
                try:
                    xlsx_buffer = generar_excel_dieta(st.session_state.plan)
                    st.success("✅ Excel generado correctamente")
                    
                    st.download_button(
                        label="⬇️ Descargar Excel",
                        data=xlsx_buffer,
                        file_name=f"Plan_Nutricional_{st.session_state.plan['cliente'].replace(' ', '_') if st.session_state.plan['cliente'] else 'Cliente'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        use_container_width=True
                    )
                except Exception as e:
                    st.error(f"Error al generar Excel: {str(e)}")

# ============================================================================
# VISTA: BASE DE DATOS DE ALIMENTOS Y MICRONUTRIENTES
# ============================================================================

elif st.session_state.plan['vista'] == 'base_datos':
    st.title("📚 Base de Datos Nutricional")
    
    # Pestañas principales
    tab_macros, tab_micros, tab_guias = st.tabs([
        "🥗 MACRONUTRIENTES", 
        "💊 MICRONUTRIENTES",
        "📖 GUÍAS Y RECOMENDACIONES"
    ])
    
    # ========================================================================
    # TAB MACRONUTRIENTES
    # ========================================================================
    with tab_macros:
        st.header("Guía Completa de Macronutrientes")
        
        # Info box con recomendaciones generales
        st.info("""
        **💡 Recomendaciones generales de distribución:**
        - **Proteína**: 1.6-2.2 g/kg peso corporal (deportistas)
        - **Carbohidratos**: 3-7 g/kg (según nivel actividad)
        - **Grasas**: 20-35% calorías totales (priorizar insaturadas)
        """)
        
        # Subtabs por macro
        subtab_prot, subtab_carbs, subtab_grasas = st.tabs([
            "🥩 Proteínas", 
            "🍚 Carbohidratos", 
            "🥑 Grasas"
        ])
        
        # ====================================================================
        # SUBTAB PROTEÍNAS
        # ====================================================================
        with subtab_prot:
            st.subheader("FUENTES DE PROTEÍNA")
            
            # Notas importantes sobre proteína
            with st.expander("📌 NOTAS IMPORTANTES SOBRE PROTEÍNAS", expanded=False):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**Digestibilidad y Calidad Proteica:**")
                    digest_data = [
                        {'Fuente': 'Huevo', 'Digestibilidad': '97%', 'Calidad': '⭐⭐⭐⭐⭐'},
                        {'Fuente': 'Whey (suero)', 'Digestibilidad': '95%', 'Calidad': '⭐⭐⭐⭐⭐'},
                        {'Fuente': 'Pescado', 'Digestibilidad': '94%', 'Calidad': '⭐⭐⭐⭐⭐'},
                        {'Fuente': 'Carne/Pollo', 'Digestibilidad': '92%', 'Calidad': '⭐⭐⭐⭐⭐'},
                        {'Fuente': 'Soja', 'Digestibilidad': '91%', 'Calidad': '⭐⭐⭐⭐'},
                        {'Fuente': 'Garbanzos', 'Digestibilidad': '78%', 'Calidad': '⭐⭐⭐'},
                        {'Fuente': 'Lentejas', 'Digestibilidad': '76%', 'Calidad': '⭐⭐⭐'},
                    ]
                    st.dataframe(pd.DataFrame(digest_data), hide_index=True, use_container_width=True)
                
                with col2:
                    st.markdown("**Timing Proteico Óptimo:**")
                    st.markdown("""
                    - **Pre-entreno**: 20-40g en 1h pre-ejercicio (estimula síntesis proteica)
                    - **Post-entreno**: 20-40g en 2h post-ejercicio
                    - **Comida/Cena**: 25-40g cada una
                    - **Pre-sueño**: 20-40g caseína (liberación lenta)
                    - **Total diario**: Repartir en 3-5 tomas """)
                    
                
                st.markdown("---")
                st.markdown("**🔬 Mejora biodisponibilidad vegetal:**")
                st.markdown("""
                - Combinar legumbres + cereales (aminoácidos complementarios)
                - Remojar legumbres 8-12h (reduce fitatos)
                - Fermentar (tempeh mejor que soja cocida)
                - Germinar semillas/legumbres
                """)
            
            # Listado de alimentos
            for categoria, alimentos in ALIMENTOS_PROTEINAS.items():
                with st.expander(f"📌 {categoria} ({len(alimentos)} alimentos)", expanded=False):
                    data = []
                    for alim in alimentos:
                        rac_comun = f"{alim['unidad']} = {alim['gramos']}g"
                        prots = round((alim['gramos'] / 100) * alim['proteina'], 1)
                        carbs = round((alim['gramos'] / 100) * alim['carbos'], 1)
                        fats = round((alim['gramos'] / 100) * alim['grasas'], 1)
                        cals = round((prots * 4) + (carbs * 4) + (fats * 9), 0)
                        
                        data.append({
                            'Alimento': alim['nombre'],
                            'Proteína/100g': f"{alim['proteina']}g",
                            'Ración común': rac_comun,
                            'Proteína': f"{prots}g",
                            'Carbos': f"{carbs}g",
                            'Grasas': f"{fats}g",
                            'Calorías': f"{int(cals)} kcal"
                        })
                    
                    df = pd.DataFrame(data)
                    st.dataframe(df, use_container_width=True, hide_index=True)
        
        # ====================================================================
        # SUBTAB CARBOHIDRATOS
        # ====================================================================
        with subtab_carbs:
            st.subheader("FUENTES DE CARBOHIDRATOS")
            
            # Notas importantes sobre carbohidratos
            with st.expander("📌 NOTAS IMPORTANTES SOBRE CARBOHIDRATOS", expanded=False):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**⚡ Timing de Carbohidratos:**")
                    st.markdown("""
                    **Deportistas:**
                    - **Pre-entreno 2h-3h antes**: 1-4 g/kg carbos complejos
                    - **Pre-entreno 30 minutos antes**: 0.5 g/kg carbos simples
                    - **Durante ejercicio >90min**: 30-60g/hora (amilopectina  / ciclodextrina + palatinosa (fibra para regular carga glucémica))
                    - **Post-entreno (0-2h)**: 1-1.2 g/kg (ventana anabólica)   
                    """)
                
                with col2:
                    st.markdown("**🔢 Índice Glucémico (IG):**")
                    st.markdown("""
                    **IG Bajo (<55)** - Libera energía lenta:
                    - Avena, legumbres, boniato, frutas
                    
                    **IG Medio (56-69)**:
                    - Arroz integral, pasta integral
                    
                    **IG Alto (>70)** - Energía rápida:
                    - Arroz blanco, pan blanco, patata
                    - Ideal: pre o post-entreno 
                    """)
                
                st.markdown("---")
                st.warning("""
                ⚠️ **IMPORTANTE**: Los valores cambian drásticamente entre crudo/cocido:
                - Arroz/pasta absorben ~200-300% agua al cocinar
                - Siempre especificar estado en tu planificación
                """)
            
            # Listado de alimentos
            for categoria, alimentos in ALIMENTOS_CARBOHIDRATOS.items():
                with st.expander(f"📌 {categoria} ({len(alimentos)} alimentos)", expanded=False):
                    data = []
                    for alim in alimentos:
                        estado = " (cocido)" if alim.get('cocido', False) else " (crudo)" if 'cocido' in alim else ""
                        rac_comun = f"{alim['unidad']}{estado} = {alim['gramos']}g"
                        prots = round((alim['gramos'] / 100) * alim['proteina'], 1)
                        carbs = round((alim['gramos'] / 100) * alim['carbos'], 1)
                        fats = round((alim['gramos'] / 100) * alim['grasas'], 1)
                        cals = round((prots * 4) + (carbs * 4) + (fats * 9), 0)
                        
                        data.append({
                            'Alimento': alim['nombre'],
                            'Carbos/100g': f"{alim['carbos']}g",
                            'Ración común': rac_comun,
                            'Proteína': f"{prots}g",
                            'Carbos': f"{carbs}g",
                            'Grasas': f"{fats}g",
                            'Calorías': f"{int(cals)} kcal"
                        })
                    
                    df = pd.DataFrame(data)
                    st.dataframe(df, use_container_width=True, hide_index=True)
        
        # ====================================================================
        # SUBTAB GRASAS
        # ====================================================================
        with subtab_grasas:
            st.subheader("FUENTES DE GRASAS SALUDABLES")
            
            # Notas importantes sobre grasas
            with st.expander("📌 NOTAS IMPORTANTES SOBRE GRASAS", expanded=False):
                col1, col2 = st.columns(2)
                
                with col1:
                    st.markdown("**🥑 Tipos de Grasas:**")
                    st.markdown("""
                    **Grasas Insaturadas (Priorizar):**
                    - **Omega-3 (EPA/DHA)**: Pescado azul 🐟
                    - **Omega-9**: Aceite oliva, aguacate, almendras
                    - **Omega-6**: Semillas, frutos secos
                      - Importante pero no exceder
                      
                    **Grasas Saturadas (Moderar):**
                    - Carnes grasas, lácteos enteros, coco
                    
                    **Grasas Trans (Evitar):**
                    - Bollería, fritos, procesados
                    """)
                
                with col2:
                    st.markdown("**⚠️ Omega-3**")
                    st.markdown("""
                    **EPA/DHA (pescado azul):**
                    - Ejemplos: salmón, sardinas, caballa
                    
                    **ALA (vegetales: nueces, lino, chía):**
                    - ⚠️ Conversión a EPA/DHA: <5-10%
                    
                    💊 **Suplementación**: Solo si no consumes pescado
                    - 1-2g EPA+DHA combinados/día
                    - Verificar certificación pureza (IFOS)
                    """)
                
                st.markdown("---")
                st.success("""
                **✅ Consejo práctico**: Las grasas son **vitaminas liposolubles** (A, D, E, K) 
                requieren grasa para absorción. Siempre añade grasa saludable a ensaladas/verduras.
                """)
            
            # Listado de alimentos
            for categoria, alimentos in ALIMENTOS_GRASAS.items():
                with st.expander(f"📌 {categoria} ({len(alimentos)} alimentos)", expanded=False):
                    data = []
                    for alim in alimentos:
                        rac_comun = f"{alim['unidad']} = {alim['gramos']}g"
                        prots = round((alim['gramos'] / 100) * alim['proteina'], 1)
                        carbs = round((alim['gramos'] / 100) * alim['carbos'], 1)
                        fats = round((alim['gramos'] / 100) * alim['grasas'], 1)
                        cals = round((prots * 4) + (carbs * 4) + (fats * 9), 0)
                        
                        data.append({
                            'Alimento': alim['nombre'],
                            'Grasas/100g': f"{alim['grasas']}g",
                            'Ración común': rac_comun,
                            'Proteína': f"{prots}g",
                            'Carbos': f"{carbs}g",
                            'Grasas': f"{fats}g",
                            'Calorías': f"{int(cals)} kcal"
                        })
                    
                    df = pd.DataFrame(data)
                    st.dataframe(df, use_container_width=True, hide_index=True)
    
    # ========================================================================
    # TAB MICRONUTRIENTES
    # ========================================================================
    with tab_micros:
        st.header("💊 Guía de Micronutrientes")
        
        # Subtabs
        subtab_vitaminas, subtab_minerales, subtab_info = st.tabs([
            "🧪 Vitaminas",
            "⚗️ Minerales",
            "📊 Información Adicional"
        ])
        
        # Vitaminas
        with subtab_vitaminas:
            vitaminas = {k: v for k, v in MICRONUTRIENTES_DB.items() 
                        if 'Vitamina' in k or 'Ácido' in k or 'Colina' in k}
            
            for nombre, info in vitaminas.items():
                with st.expander(f"💊 {nombre}", expanded=False):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Función**: {info['funcion']}")
                        st.markdown(f"**Mejores fuentes**: {', '.join(info['fuentes'][:3])}")
                    
                    with col2:
                        st.markdown(f"**RDA Hombre**: {info['RDA_hombre']}")
                        st.markdown(f"**RDA Mujer**: {info['RDA_mujer']}")
        
        # Minerales
        with subtab_minerales:
            minerales = {k: v for k, v in MICRONUTRIENTES_DB.items() 
                        if 'Vitamina' not in k and 'Ácido' not in k and 'Colina' not in k}
            
            for nombre, info in minerales.items():
                with st.expander(f"⚗️ {nombre}", expanded=False):
                    col1, col2 = st.columns([2, 1])
                    
                    with col1:
                        st.markdown(f"**Función**: {info['funcion']}")
                        st.markdown(f"**Mejores fuentes**: {', '.join(info['fuentes'][:3])}")
                    
                    with col2:
                        st.markdown(f"**RDA Hombre**: {info['RDA_hombre']}")
                        st.markdown(f"**RDA Mujer**: {info['RDA_mujer']}")
        
        # Información adicional
        with subtab_info:
            st.subheader("📊 Información Clave sobre Micronutrientes")
            
            # Interacciones
            with st.expander("🔄 INTERACCIONES ENTRE NUTRIENTES", expanded=True):
                st.markdown("### ✅ Interacciones Sinérgicas (Potencian absorción)")
                
                for inter in NOTAS_MICRONUTRIENTES['interacciones']['sinergicas']:
                    st.markdown(f"""
                    **{inter['combinacion']}**  
                    - Efecto: {inter['efecto']}  
                    - Ejemplo: {inter['ejemplo']}
                    """)
                    st.markdown("---")
                
                st.markdown("### ⚠️ Interacciones Antagónicas (Reducen absorción)")
                
                for inter in NOTAS_MICRONUTRIENTES['interacciones']['antagonicas']:
                    recomendacion = inter.get('recomendacion', inter.get('solucion', inter.get('nota', 'Ver información detallada')))
                    st.markdown(f"""
                    **{inter['combinacion']}**  
                    - Efecto: {inter['efecto']}  
                     - Recomendación: {recomendacion}""")
                    st.markdown("---")
            
            # Biodisponibilidad
            with st.expander("🔬 BIODISPONIBILIDAD Y ABSORCIÓN", expanded=False):
                st.markdown("### Digestibilidad de Proteínas")
                prot_bio = pd.DataFrame(NOTAS_MICRONUTRIENTES['biodisponibilidad']['proteinas'])
                st.dataframe(prot_bio, hide_index=True, use_container_width=True)
                
                st.markdown("### Absorción de Hierro")
                for item in NOTAS_MICRONUTRIENTES['biodisponibilidad']['hierro']:
                    st.markdown(f"**{item['tipo']}**: {item['absorcion']} - {item['nota']}")
                
                st.markdown("### Absorción de Calcio")
                for item in NOTAS_MICRONUTRIENTES['biodisponibilidad']['calcio']:
                    st.markdown(f"**{item['fuente']}**: {item['absorcion']} - {item['nota']}")
                
                st.markdown("### Factores que Mejoran Absorción ✅")
                for factor in NOTAS_MICRONUTRIENTES['biodisponibilidad']['factores_mejoran_absorcion']:
                    st.markdown(f"- {factor}")
                
                st.markdown("### Factores que Reducen Absorción ⚠️")
                for factor in NOTAS_MICRONUTRIENTES['biodisponibilidad']['factores_reducen_absorcion']:
                    st.markdown(f"- {factor}")
            
            # Deficiencias comunes
            with st.expander("🚨 DEFICIENCIAS MÁS COMUNES", expanded=False):
                st.markdown("### Déficits Críticos")
                
                for def_item in NOTAS_MICRONUTRIENTES['deficiencias_comunes']['criticos']:
                    st.error(f"""
                    **{def_item['nutriente']}**
                    - Prevalencia: {def_item['prevalencia']}
                    - Grupos de riesgo: {', '.join(def_item['grupos_riesgo'])}
                    - Consecuencias: {def_item['consecuencias']}
                    - Solución: {def_item['solucion']}
                    """)
                
                st.markdown("### Déficits Moderados")
                for def_item in NOTAS_MICRONUTRIENTES['deficiencias_comunes']['moderados']:
                    st.warning(f"""
                    **{def_item['nutriente']}**  
                    Prevalencia: {def_item['prevalencia']}  
                    Nota: {def_item['nota']}
                    """)
    
    # ========================================================================
    # TAB GUÍAS Y RECOMENDACIONES
    # ========================================================================
    with tab_guias:
        st.header("📖 Guías y Recomendaciones Prácticas")
        
        # Sub-tabs
        subtab_analitica, subtab_hidra = st.tabs([
            "🩸 Analíticas Recomendadas",
            "💧 Hidratación"
            
        ])
        
        # ====================================================================
        # ANALÍTICAS
        # ====================================================================
        with subtab_analitica:
            st.subheader("🩸 Analíticas: Qué Solicitar y Rangos Óptimos")
            
            st.info("""
            **💡 Importante**: Los rangos "normales" de laboratorio no siempre son "óptimos".  
            Esta guía proporciona valores objetivo para rendimiento y salud óptimos. Consulta con un médico.
            """)
            
            # Panel básico
            with st.expander("📋 PANEL BÁSICO (Anual)", expanded=True):
                st.markdown("### Análisis Básicos Recomendados")
                for item in NOTAS_MICRONUTRIENTES['analitica_completa']['panel_basico']:
                    st.markdown(f"""
                    **{item['marcador']}**
                    - Evalúa: {item.get('evalua', '-')}
                    - Rango óptimo: {item.get('rango_optimo', item.get('nota', '-'))}
                    - Frecuencia: {item.get('frecuencia', 'Anual')}
                    """)
                    st.markdown("---")
            
            # Vitaminas
            with st.expander("🧪 VITAMINAS", expanded=False):
                for vit in NOTAS_MICRONUTRIENTES['analitica_completa']['vitaminas']:
                    st.markdown(f"### {vit['vitamina']}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if 'rango_deficiencia' in vit:
                            st.metric("❌ Deficiencia", vit['rango_deficiencia'])
                        if 'rango_insuficiente' in vit:
                            st.metric("⚠️ Insuficiente", vit['rango_insuficiente'])
                    
                    with col2:
                        st.metric("✅ Óptimo", vit['rango_optimo'])
                        if 'rango_toxicidad' in vit:
                            st.metric("☠️ Toxicidad", vit['rango_toxicidad'])
                    
                    st.markdown(f"**Notas**: {vit['notas']}")
                    st.markdown("---")
            
            # Minerales
            with st.expander("⚗️ MINERALES", expanded=False):
                for min_item in NOTAS_MICRONUTRIENTES['analitica_completa']['minerales']:
                    st.markdown(f"### {min_item['mineral']}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if 'rango_deficiencia' in min_item:
                            st.metric("❌ Deficiencia", min_item['rango_deficiencia'])
                        if 'rango_optimo_hombre' in min_item:
                            st.metric("✅ Óptimo Hombre", min_item['rango_optimo_hombre'])
                    
                    with col2:
                        if 'rango_optimo' in min_item:
                            st.metric("✅ Óptimo", min_item['rango_optimo'])
                        if 'rango_optimo_mujer' in min_item:
                            st.metric("✅ Óptimo Mujer", min_item['rango_optimo_mujer'])
                    
                    st.markdown(f"**Nota**: {min_item['nota']}")
                    st.markdown("---")
            
            # Marcadores funcionales
            with st.expander("📊 MARCADORES FUNCIONALES", expanded=False):
                for marc in NOTAS_MICRONUTRIENTES['analitica_completa']['marcadores_funcionales']:
                    st.markdown(f"### {marc['marcador']}")
                    
                    col1, col2 = st.columns(2)
                    with col1:
                        if 'rango_optimo' in marc:
                            st.metric("✅ Óptimo", marc['rango_optimo'])
                        if 'rango_moderado' in marc:
                            st.metric("⚠️ Moderado", marc['rango_moderado'])
                    
                    with col2:
                        if 'rango_bajo' in marc:
                            st.metric("❌ Bajo", marc['rango_bajo'])
                        if 'evalua' in marc:
                            st.markdown(f"**Evalúa**: {marc['evalua']}")
                    
                    if 'nota' in marc:
                        st.markdown(f"**Nota**: {marc['nota']}")
                    st.markdown("---")
            
            # Cuándo solicitar
            with st.expander("🔍 CUÁNDO SOLICITAR ANALÍTICAS", expanded=False):
                st.markdown("### 🚨 Por Síntomas")
                for situacion in NOTAS_MICRONUTRIENTES['analitica_completa']['cuando_solicitar']:
                    st.markdown(f"- {situacion}")
                
                st.markdown("---")
                st.markdown("### 📅 Frecuencia Recomendada por Perfil")
                for perfil, frec in NOTAS_MICRONUTRIENTES['analitica_completa']['frecuencia_recomendada'].items():
                    perfil_formateado = perfil.replace('_', ' ').title()
                    st.markdown(f"**{perfil_formateado}**: {frec}")

        
        # ====================================================================
        # HIDRATACIÓN
        # ====================================================================
        with subtab_hidra:
            st.subheader("💧 Guía Completa de Hidratación")
            
            col1, col2 = st.columns(2)
            
            with col1:
                st.markdown("### 📏 Cálculo de Necesidades")
                
                peso_hidra = st.number_input(
                    "Tu peso corporal (kg)",
                    min_value=40,
                    max_value=150,
                    value=70,
                    key="peso_hidra"
                )
                
                agua_base = peso_hidra * 0.033  # 33ml por kg
                st.metric("Agua Base Diaria", f"{agua_base:.1f} litros")
                
                st.markdown("### 🏃 Ajuste por Actividad")
                horas_ejercicio = st.slider(
                    "Horas de ejercicio/día",
                    0.0, 4.0, 1.0, 0.5,
                    key="horas_ejercicio_hidra"
                )
                
                agua_ejercicio = horas_ejercicio * 0.5  # 500ml por hora
                agua_total = agua_base + agua_ejercicio
                
                st.metric("💧 Total Recomendado", f"{agua_total:.1f} litros/día")
                st.caption(f"≈ {int(agua_total * 5)} vasos de 200ml")
            
            with col2:
                st.markdown("### 🔥 Factores que Aumentan Necesidades")
                st.markdown("""
                - **Calor/humedad**: +500-1000ml
                - **Altitud >2500m**: +500ml
                - **Fiebre**: +200ml por cada °C >37°C
                - **Lactancia**: +700ml
                - **Diarrea/vómitos**: Reponer pérdidas
                - **Alcohol**: +250ml por bebida alcohólica
                - **Cafeína**: +100ml por café
                """)
                
                st.markdown("### ⚠️ Señales de Deshidratación")
                st.error("""
                - Orina oscura (color miel)
                - Fatiga inusual
                - Dolor de cabeza
                - Mareos/confusión
                - Calambres musculares
                - Sed intensa
                """)
            
            st.markdown("---")
            
            # Timing de hidratación
            st.markdown("### ⏰ Timing de Hidratación para Deportistas")
            
            timing_hidra = pd.DataFrame([
                {
                    'Momento': 'Al despertar',
                    'Cantidad': '400-500ml',
                    'Objetivo': 'Rehidratar tras ayuno nocturno'
                },
                {
                    'Momento': '2-3h pre-ejercicio',
                    'Cantidad': '400-600ml',
                    'Objetivo': 'Asegurar estado hidratado'
                },
                {
                    'Momento': '15min pre-ejercicio',
                    'Cantidad': '200-300ml',
                    'Objetivo': 'Último aporte antes de empezar'
                },
                {
                    'Momento': 'Durante (<60min)',
                    'Cantidad': '150-200ml cada 15-20min',
                    'Objetivo': 'Solo agua necesaria'
                },
                {
                    'Momento': 'Durante (>60min)',
                    'Cantidad': '200-300ml cada 15-20min',
                    'Objetivo': 'Añadir electrolitos/carbos'
                },
                {
                    'Momento': 'Post-ejercicio',
                    'Cantidad': '150% del peso perdido',
                    'Objetivo': 'Pesarse antes y después. Reponer 1.5L por kg perdido'
                }
            ])
            
            st.dataframe(timing_hidra, hide_index=True, use_container_width=True)
            
            st.info("""
            **💡 Consejo práctico**: Pésate antes y después del ejercicio.  
            Por cada kg perdido, necesitas 1.5 litros de líquido (se pierde más de lo absorbido).
            """)
        
            
            st.markdown("---")
            st.info("""
            **💡 Nota importante**: Estas son recomendaciones generales.  
            La individualización es clave. Prueba estrategias en entrenamientos antes de aplicar en competición.
            """)
        
        # Notas finales importantes
        st.markdown("---")
        st.markdown("### ⚠️ Consideraciones Finales Importantes")
        
        for nota in NOTAS_MICRONUTRIENTES['notas_importantes']:
            st.markdown(f"- {nota}")


# ============================================================================
# FOOTER
# ============================================================================
st.markdown("---")
st.caption("🥗 **Constructor de Dieta Pro** | David López | Basado en evidencia científica 2024-2025")
