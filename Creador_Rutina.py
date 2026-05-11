# app.py
import streamlit as st
import pandas as pd
from datetime import datetime
import plotly.graph_objects as go
from reportlab.lib import colors
from reportlab.lib.pagesizes import A4, landscape
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import cm, mm
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT, TA_JUSTIFY
from reportlab.pdfgen import canvas
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

# ============================================================================
# CONFIGURACIÓN INICIAL
# ============================================================================
st.set_page_config(
    page_title="Constructor de Rutina", 
    page_icon="🏋️", 
    layout="wide",
    initial_sidebar_state="expanded"
)

# ============================================================================
# BASE DE DATOS DE EJERCICIOS
# ============================================================================
GRUPOS_MUSCULARES = [
    "Pectoral", "Espalda", "Hombro", "Cuádriceps", "Femoral", 
    "Glúteo", "Bíceps", "Tríceps", "Core", "Gemelos"
]

REGIONES_MAP = {
    "Pecho": ["Superior", "Medio", "Inferior"],
    "Espalda": ["Dorsal","Trapecio","Erectores espinares"],
    "Trapecio": ["General"],
    "Hombro": ["Anterior", "Lateral", "Posterior"],
    "Cuádriceps": ["General", "Recto femoral"],
    "Femoral": ["General"],
    "Glúteo": ["Glúteo mayor", "Glúteo medio"],
    "Bíceps": ["General"],
    "Tríceps": ["General"],
    "Core": ["Recto abdominal", "Oblicuos", "Transverso"],
    "Gemelos": ["Gastrocnemio", "Sóleo"]
}

EJERCICIOS_DB = {
    "Pecho": {
        "Superior": [
            "Press inclinado barra 45°", 
            "Press inclinado máquina",
            "Press inclinado mancuernas 45°",
            "Press inclinado multipower",
            "Press landmine", 
            "Cruces cable polea baja", 
            "Flexiones declinadas",
            "Peck deck inclinado 30°", 
            "Press máquina convergente inclinado", 
            "Aperturas cable inclinado 30°"
        ],
        "Medio": [
            "Press banca plano barra", 
            "Press banca plano máquina",
            "Press mancuernas plano", 
            "Press plano multipower",
            "Fondos paralelas (inclinar tronco hacia adelnte 30º)", 
            "Flexiones planas", 
            "Aperturas mancuernas plano", 
            "Cruces cable altura media", 
            "Flexiones estándar",
            "Peck deck máquina / contractora", 
            "Flexiones anillas", 
            "Press máquina convergente"
        ],
        "Inferior": [
            "Press declinado barra", 
            "Press declinado máquina",
            "Dips paralelas con lastre", 
            "Cruces cable polea alta", 
            "Press declinado mancuernas",
            "Fondos pectorales inclinación adelante", 
            "Cruces cable de alto a bajo arrodillado"
        ],
    }, 
    "Espalda": {
        "Dorsal": [
            "Dominadas pronación agarre ancho", 
            "Dominadas supinación", 
            "Dominadas agarre neutro",
            "Dominadas asistidas máquina",
            "Jalón unilateral en máquina", 
            "Jalón unilateral en banco", 
            "Jalón al pecho prono agarre ancho", 
            "Jalón al pecho agarre neutro", 
            "Jalón al pecho supino", 
            "Remo máquina unilateral brazo pegado al cuerpo (codo a cadera)",
            "Seal row", 
            "Remo unilateral en polea apoyado en banco", 
            "Remo gironda agarre neutro",
            "Remo con mancuerna", 
            "Pullover polea brazo extendido unilateral", 
            "Pullover mancuerna",
            "Dominadas agarre cerrado supino", 
            "Jalón agarre neutro cerrado", 


        ],
        "Trapecio": [
            "Remo mancuerna",
            "Remo barra",
            "Remo T agarre ancho",
            "Remo T agarre estrecho",
            "Remo alto máquina",
            "Remo alto Gironda agarre ancho",
            "Jalón espalda alta",
            "Remo cable agarre neutro", 
            "Face pulls", 
            "Remo invertido", 
            "Chest supported row", 
            "Encogimientos trapecio barra",
            "I-Y-T raises", 
            "Encogimientos con mancuerna", 
            "Encogimientos en máquina remo T",
            "Encogimientos en máquina", 
            "Remo invertido anillas", 
            "Peso muerto", 
            "Remo mancuerna inclinado 45°", 
            "Remo Kroc"
        ],
        "Erectores espinales": [
            "Peso muerto convencional", 
            "Peso muerto rumano", 
            "Buenos días", 
            "Extensiones lumbares 45°", 
            "Hiperextensiones espalda",
            "Peso muerto trap bar", 
            "Good mornings en caja", 
            "Back extension 90°"
        ],
    },
    "Hombro": {
        "Anterior": [
            "Press militar barra", 
            "Press militar sentado multipower",
            "Press mancuernas sentado", 
            "Press militar máquina",
            "Press Arnold", 
            "Press landmine unilateral", 
            "Elevaciones frontales barra", 
            "Elevaciones frontales disco", 
            "Pike push-ups",
            "Press máquina convergente hombro", 
            "Elevaciones frontales cable polea baja", 
            "Press Bradford"
        ],
        "Lateral": [
            "Elevaciones laterales mancuernas", 
            "Elevaciones laterales cable unilateral", 
            "Elevaciones laterales máquina", 
            "Press mancuernas sentado", 
            "Elevaciones laterales cable tumbado", 
            "Upright row agarre ancho", 
            "Elevaciones laterales en banco inclinado", 
            "Lu raises",
            "Press mancuernas neutral grip", 
            "Upright row cable agarre ancho", 
            "Elevaciones laterales máquina", 
            "Cable lateral raises inclinado 30°"
        ],
        "Posterior": [
            "Face pulls", 
            "Elevaciones posteriores 45° inclinado", 
            "Remo alto cable agarre amplio", 
            "Reverse peck deck", 
            "Elevaciones posteriores cable", 
            "Prone Y raises",
            "Remo al cuello cable cuerda", 
            "Reverse flyes banco inclinado 45°", 
            "Cable posterior deltoid fly", 
            "Band pull-aparts"
        ],
    },
    "Cuádriceps": {
        "General": [
            "Sentadilla back", 
            "Sentadilla frontal", 
            "Prensa 45°", 
            "Hack squat", 
            "Sentadilla multipower",
            "Sentadilla goblet", 
            "Extensión cuádriceps máquina", 
            "Sentadilla búlgara",
            "Sentadilla Zercher", 
            "Belt squat", 
            "Sissy squat asistida", 
            "Sentadilla Smith inclinada",
            "Sentadilla búlgara mancuerna pie hacia hacia atrás", 
            "Sentadilla búlgara multipower pie hacia atrás", 
        ],
        "Recto femoral": [

            "Extensión cuádriceps máquina", 
            "Sissy squat", 
            "Step-ups con mancuernas",
            "Leg extension unilateral", 
        ],
    },
    "Isquiosurales": {
        "General": [
            "Peso muerto rumano",
            "Peso muerto mancuernas",
            "Peso muerto piernas semirígidas",  
            "Curl femoral tumbado", 
            "Curl femoral sentado", 
            "Buenos días", 
            "Peso muerto piernas rígidas", 
            "Deslizamientos nórdicos",
            "Curl Nordico", 
            "Peso muerto unilateral mancuerna", 
            "Curl femoral nórdico inverso", 
            "Deslizamientos isquios TRX",
            "Glute-ham raise", 
            "Stiff-leg deadlift elevado", 
            "Single-leg RDL"
            "Sentadilla búlgara mancuerna pie hacia adelante", 
            "Sentadilla búlgara multipower pie hacia adelante", 
        ],
    },
    "Glúteo": {
        "Glúteo mayor": [
            "Hip thrust barra", 
            "Hip thrust máquina", 
            "Sentadilla búlgara mancuerna pie hacia adelante", 
            "Sentadilla búlgara multipower pie hacia adelante", 
            "Peso muerto mancuernas",
            "Peso muerto sumo", 
            "Patada glúteo cable", 
            "Prensa unilateral", 
            "Hip thrust unilateral", 
            "Puente glúteo",
            "Hip thrust pies elevados", 
            "Peso muerto rumano unilateral", 
            "Frog pump", 
            "Reverse hyper", 
            "Patada glúteo máquina"
        ],
        "Glúteo medio": [
            "Abducción cadera cable", 
            "Clamshells con banda", 
            "Abducción cadera máquina", 
            "Zancadas laterales", 
            "Step-ups laterales", 
            "Hip thrust unilateral", 
            "Abducción cadera máquina de pie", 
            "Monster walks banda diagonal", 
            "Fire hydrants con banda"
        ],
    },
    "Aductor": {
        "General": [
            "Aductor máquina", 
            "Aductor polea"
        ],
    },
    "Bíceps": {
        "General": [
            "Curl barra", 
            "Curl mancuernas",
            "Curl barra polea", 
            "Curl predicador / banco scott", 
            "Curl cable",
            "Curlo cable hombro atrás",
            "Curl concentrado mancuerna", 
            "Curl inclinado 45°",
            "Curl concentrado polea",
            "Curl araña", 
            "Drag curl"
        ],
        "Braquial": [
            "Curl martillo mancuernas", 
            "Curl martillo polea agarre neutro", 

                   ],

    },
    "Tríceps": {
        "General": [
            "Press banca agarre cerrado", 
            "Press agarre cerrado multipower",
            "Extensión tríceps polea cuerda unilateral", 
            "Extensión tríceps barra Z", 
            "Extensión tríceps polea cuerda", 
            "Press francés mancuernas",
            "Press francés barra",
            "Extensión francesa", 
            "Patada tríceps", 
            "Dips tríceps", 
            "Extensión cable polea sobre la cabeza (Katana)",
            "Extensión mancuerna sobre la cabeza (Katana)",
            "Fondos",
            "Cable pushdown agarre inverso", 
            "Kaz press multipower", 
            "Diamond push-ups"
        ],
    },
    "Core": {
        "Recto abdominal": [
            "Crunch abdominal", 
            "Elevaciones piernas colgado", 
            "Elevaciones piernas máquina",
            "Crunch cable polea alta", 
            "Rueda abdominal", 
            "Plancha",
            "Dragon flag", 
            "Ab wheel de pie", 
            "Decline crunch con peso", 
            "Cable crunch arrodillado"
        ],
        "Oblicuos": [
            "Plancha lateral", 
            "Russian twist", 
            "Oblicuos cable", 
            "Bicicleta abdominal",
            "Landmine rotation", 
            "Copenhagen plank", 
            "Cable woodchop de alto a bajo", 
            "Suitcase carry"]
    },
    "Gemelos": {
        "Gastrocnemio": [
            "Elevación gemelos de pie", 
            "Elevación gemelos prensa", 
            "Saltos cuerda",
            "Elevación gemelos unilateral de pie", 
            "Elevación gemelos Smith", 
            "Donkey calf raise"
        ],
        "Sóleo": [
            "Elevación gemelos sentado", 
            "Elevación gemelos sentado mancuernas",
            "Elevación gemelos sentado máquina", 
            "Calf press en prensa 45°", 
        ],
    },
}


VOLUMEN_OPTIMO = {
    "Pectoral": {"min": 10, "max": 20},
    "Aductor": {"min": 6, "max": 12},
    "Dorsal": {"min": 10, "max": 20},
    "Trapecio": {"min": 8, "max": 16},
    "Hombro Anterior": {"min": 6, "max": 10},
    "Hombro Lateral": {"min": 8, "max": 14},
    "Hombro Posterior": {"min": 8, "max": 12},
    "Cuádriceps": {"min": 12, "max": 20},
    "Femoral": {"min": 10, "max": 16},
    "Glúteo": {"min": 6, "max": 14},
    "Bíceps": {"min": 10, "max": 18},
    "Tríceps": {"min": 10, "max": 18},
    "Core": {"min": 12, "max": 20},
    "Gemelos": {"min": 8, "max": 16},
}

def get_todos_ejercicios():
    todos = []
    for grupo, regiones in EJERCICIOS_DB.items():
        for region, ejercicios in regiones.items():
            for ej in ejercicios:
                todos.append({'nombre': ej, 'grupo': grupo, 'region': region})
    return todos

TODOS_EJERCICIOS = get_todos_ejercicios()

# ============================================================================
# FUNCIONES AUXILIARES
# ============================================================================
def get_grupo_directo(grupo, region):
    if grupo == "Hombro":
        return f"Hombro {region}"
    else:
        return grupo

def segundos_a_minutos(segundos):
    minutos = segundos / 60
    if minutos == int(minutos):
        return f"{int(minutos)}'"
    else:
        return f"{minutos:.1f}'".replace('.', "'")

def calcular_volumen_semanal(routine):
    volumen = {k: 0 for k in VOLUMEN_OPTIMO.keys()}
    for dia, data in routine['dias'].items():
        for ej in data['ejercicios']:
            grupo_directo = ej['grupo_directo']
            if grupo_directo in volumen:
                volumen[grupo_directo] += ej['series']
    return volumen

def calcular_volumen_por_dia(routine, dia):
    volumen_dia = {k: 0 for k in VOLUMEN_OPTIMO.keys()}
    for ej in routine['dias'].get(dia, {}).get('ejercicios', []):
        grupo_directo = ej['grupo_directo']
        if grupo_directo in volumen_dia:
            volumen_dia[grupo_directo] += ej['series']
    return volumen_dia

def validar_volumen(volumen):
    deficit = []
    exceso = []
    optimo = []
    for grupo, series in volumen.items():
        opt = VOLUMEN_OPTIMO[grupo]
        if series < opt['min'] and series > 0:
            deficit.append(grupo)
        elif series > opt['max']:
            exceso.append(grupo)
        elif series >= opt['min'] and series <= opt['max']:
            optimo.append(grupo)
    return deficit, exceso, optimo

# ============================================================================
# CLASE PARA MARCA DE AGUA
# ============================================================================
class MarcaAguaCanvas(canvas.Canvas):
    def __init__(self, *args, **kwargs):
        canvas.Canvas.__init__(self, *args, **kwargs)
        self.pages = []
        
    def showPage(self):
        self.pages.append(dict(self.__dict__))
        self._startPage()
        
    def save(self):
        for page_dict in self.pages:
            self.__dict__.update(page_dict)
            self.draw_watermark()
            canvas.Canvas.showPage(self)
        canvas.Canvas.save(self)
        
    def draw_watermark(self):
        self.saveState()
        self.setFont('Helvetica', 8)
        self.setFillColor(colors.Color(0.7, 0.7, 0.7, alpha=0.3))
        self.drawRightString(self._pagesize[0] - 1.5*cm, 1*cm, "David López - Rutina Personalizada")
        self.restoreState()

# ============================================================================
# FUNCIÓN GENERAR PDF 
# ============================================================================


def generar_pdf_rutina(routine):
    """Genera PDF profesional con todas las páginas en horizontal"""
    buffer = BytesIO()
    
    doc = SimpleDocTemplate(
        buffer, 
        pagesize=landscape(A4),
        rightMargin=2*cm, 
        leftMargin=2*cm, 
        topMargin=1.2*cm, 
        bottomMargin=1.2*cm
    )
    
    elements = []
    styles = getSampleStyleSheet()
    
    # ========== ESTILOS MEJORADOS ==========
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=24,
        textColor=colors.Color(0.1, 0.1, 0.1),
        spaceAfter=20,
        spaceBefore=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    subtitle_style = ParagraphStyle(
        'CustomSubtitle',
        parent=styles['Heading2'],
        fontSize=16,
        textColor=colors.Color(0.2, 0.2, 0.2),
        spaceAfter=12,
        spaceBefore=8,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    info_style = ParagraphStyle(
        'InfoStyle',
        parent=styles['Normal'],
        fontSize=11,
        spaceAfter=8,
        fontName='Helvetica',
        leading=16,
        alignment=TA_JUSTIFY
    )
    
    heading_style = ParagraphStyle(
        'HeadingStyle',
        fontSize=12,
        textColor=colors.Color(0.2, 0.2, 0.2),
        spaceAfter=8,
        spaceBefore=10,
        fontName='Helvetica-Bold'
    )
    
    info_label_style = ParagraphStyle(
        'InfoLabel',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica-Bold',
        textColor=colors.Color(0.3, 0.3, 0.3)
    )
    
    info_value_style = ParagraphStyle(
        'InfoValue',
        parent=styles['Normal'],
        fontSize=11,
        fontName='Helvetica',
        textColor=colors.Color(0.1, 0.1, 0.1)
    )
    
    cliente_nombre = routine['cliente'] if routine['cliente'] else "Cliente"
    
    # ========== PÁGINA 1: PROGRAMA DE ENTRENAMIENTO ==========
    elements.append(Spacer(1, 2.5*cm))
    elements.append(Paragraph("PROGRAMA DE ENTRENAMIENTO", title_style))
    elements.append(Paragraph("PERSONALIZADO", subtitle_style))
    elements.append(Spacer(1, 1.5*cm))
    
    # Información del cliente en tabla limpia
    data_info = [
        [Paragraph('Nombre:', info_label_style), Paragraph(cliente_nombre, info_value_style)],
        [Paragraph('Duración:', info_label_style), Paragraph(f"{routine['semanas']} semanas", info_value_style)],
        [Paragraph('Frecuencia:', info_label_style), Paragraph(f"{routine['num_dias']} días por semana", info_value_style)],
        [Paragraph('Sexo:', info_label_style), Paragraph(routine['metadata']['sexo'], info_value_style)],
        [Paragraph('Fecha de creación:', info_label_style), Paragraph(routine['metadata']['fecha_creacion'], info_value_style)]
    ]
    
    tabla_info = Table(data_info, colWidths=[4*cm, 10*cm])
    tabla_info.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, -1), 'RIGHT'),
        ('ALIGN', (1, 0), (1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 10),
        ('LEFTPADDING', (0, 0), (-1, -1), 15),
        ('RIGHTPADDING', (0, 0), (-1, -1), 15),
    ]))
    
    # Centrar la tabla
    tabla_centrada = Table([[tabla_info]], colWidths=[14*cm])
    tabla_centrada.setStyle(TableStyle([
        ('ALIGN', (0, 0), (0, 0), 'CENTER'),
    ]))
    
    elements.append(tabla_centrada)
    elements.append(Spacer(1, 3*cm))
    
    elements.append(Paragraph("David López", 
                             ParagraphStyle('Footer1', parent=info_style, alignment=TA_CENTER, 
                                          fontSize=12, fontName='Helvetica-Bold')))
    elements.append(Paragraph("Entrenador Personal", 
                             ParagraphStyle('Footer2', parent=info_style, alignment=TA_CENTER, 
                                          textColor=colors.Color(0.5, 0.5, 0.5), fontSize=10)))
    
    elements.append(PageBreak())
    
    # ========== PÁGINA 2: INSTRUCCIONES ==========
    elements.append(Paragraph("INSTRUCCIONES DE USO", title_style))
    elements.append(Spacer(1, 0.5*cm))
    
    elements.append(Paragraph("1. Estructura del programa", heading_style))
    elements.append(Paragraph(
        "Este documento contiene tu rutina semanal completa con hojas de seguimiento para cada semana del programa. ", info_style))
    
    elements.append(Paragraph("2. Cómo registrar los datos", heading_style))
    instrucciones_registro = """
    Para cada ejercicio y cada serie, debes anotar los siguientes datos en las casillas correspondientes:<br/><br/>
    • <b>Peso:</b> Carga utilizada en kilogramos (kg). Anota el peso exacto que levantaste en cada serie.<br/><br/>
    • <b>Repeticiones (Reps):</b> Número total de repeticiones completadas con técnica correcta. 
    No cuentes repeticiones con mala ejecución.<br/><br/>
    • <b>RIR (Repeticiones en Reserva):</b> Cuántas repeticiones más podrías haber hecho al terminar la serie. 
    RIR 0 significa que llegaste al fallo muscular completo, RIR 1 significa que podrías haber hecho 1 repetición más, 
    RIR 2 significa que podrías haber hecho 2 repeticiones más, y así sucesivamente.<br/><br/>
    • <b>Notas:</b> Espacio para anotar sensaciones, dificultades, aspectos técnicos o cualquier observación relevante sobre la ejecución.
    """
    elements.append(Paragraph(instrucciones_registro, info_style))
    
    elements.append(Paragraph("3. Sistema de progresión y sobrecarga progresiva", heading_style))
    progresion = """
    <b>Cuando el ejercicio indica un rango de repeticiones (ejemplo: 6-8 repeticiones):</b><br/>
    Comienza con un peso que te permita realizar el número mínimo de repeticiones del rango (6 repeticiones en este ejemplo) 
    alcanzando el RIR objetivo marcado. En las siguientes sesiones de entrenamiento, mantén ese mismo peso y aumenta 
    progresivamente el número de repeticiones hasta alcanzar el máximo del rango (8 repeticiones), siempre controlando 
    completamente el movimiento y manteniendo una técnica perfecta. Una vez que logres realizar el máximo de repeticiones 
    del rango (8 repeticiones) con buena técnica al RIR objetivo durante al menos dos sesiones consecutivas, incrementa 
    la carga en la siguiente sesión (generalmente entre 2.5-5 kg dependiendo del ejercicio) y vuelve a comenzar la progresión 
    desde el número mínimo de repeticiones del rango (6 repeticiones). Repite este ciclo de forma continua.<br/><br/>
    <b>Cuando el ejercicio indica un número fijo de repeticiones (ejemplo: 10 repeticiones):</b><br/>
    La sobrecarga progresiva se realizará únicamente aumentando el peso. Mantén el peso hasta que puedas completar 
    todas las series programadas al RIR objetivo con técnica adecuada durante al menos dos sesiones. Entonces, 
    aumenta el peso y continúa.
    """
    elements.append(Paragraph(progresion, info_style))
    
    elements.append(PageBreak())
    
    # ========== PÁGINA 3: PLANNING ==========
    elements.append(Paragraph("PLANNING DE ENTRENAMIENTO", title_style))
    elements.append(Spacer(1, 0.3*cm))
    
    elements.append(Paragraph(
        f"Duración total: {routine['semanas']} semanas  |  Estructura semanal: {routine['num_dias']} días de entrenamiento",
        ParagraphStyle('InfoPlanning', parent=info_style, alignment=TA_CENTER, fontSize=11)
    ))
    elements.append(Spacer(1, 0.5*cm))
    
    dias_necesarios = [f"Día {i+1}" for i in range(routine['num_dias'])]
    max_ejercicios = max([len(routine['dias'].get(dia, {}).get('ejercicios', [])) for dia in dias_necesarios])
    
    # Headers sin <b>
    headers = []
    for dia in dias_necesarios:
        dia_nombre = routine['dias'][dia].get('nombre', '')
        if dia_nombre:
            headers.append(f"{dia} - {dia_nombre}")
        else:
            headers.append(dia)
    
    data_planning = [headers]
    
    for i in range(max_ejercicios):
        fila = []
        for dia in dias_necesarios:
            ejercicios = routine['dias'].get(dia, {}).get('ejercicios', [])
            if i < len(ejercicios):
                fila.append(ejercicios[i]['ejercicio'])
            else:
                fila.append("")
        data_planning.append(fila)
    
    ancho_columna = (landscape(A4)[0] - 4*cm) / len(dias_necesarios)
    tabla_planning = Table(data_planning, colWidths=[ancho_columna] * len(dias_necesarios))
    tabla_planning.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.2, 0.2, 0.2)),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        ('TOPPADDING', (0, 0), (-1, 0), 12),
        ('GRID', (0, 0), (-1, -1), 1, colors.Color(0.6, 0.6, 0.6)),
        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -1), 8),
        ('LEFTPADDING', (0, 0), (-1, -1), 5),
        ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.95, 0.95, 0.95)]),
    ]))
    
    elements.append(tabla_planning)
    elements.append(Spacer(1, 0.5*cm))
    
    # Nota calentamiento
    elements.append(Paragraph(
        "IMPORTANTE: EL CALENTAMIENTO Y LA MOVILIDAD SIEMPRE SE HARÁ CON GOMAS O PESAS DE BAJO PESO E INTENSIDAD BAJA.",
        ParagraphStyle('Aviso', parent=info_style, fontName='Helvetica-Bold', alignment=TA_CENTER, 
                      textColor=colors.Color(0.8, 0, 0), fontSize=10)
    ))
    
    elements.append(PageBreak())
    
    # ========== RUTINAS POR SEMANA Y DÍA - MÁS COMPACTAS ==========
    for semana in range(1, routine['semanas'] + 1):
        for dia in dias_necesarios:
            if routine['dias'].get(dia, {}).get('ejercicios', []):
                dia_nombre = routine['dias'][dia].get('nombre', '')
                
                titulo_pagina = f"RUTINA {cliente_nombre.upper()} - SEMANA {semana}"
                elements.append(Paragraph(titulo_pagina, 
                    ParagraphStyle('TituloRutina', parent=title_style, fontSize=18, spaceAfter=8)))
                
                titulo_dia = f"{dia.upper()}" + (f" - {dia_nombre.upper()}" if dia_nombre else "")
                elements.append(Paragraph(titulo_dia, 
                    ParagraphStyle('TituloDia', parent=subtitle_style, fontSize=13, spaceAfter=10)))
                
                # Tabla compacta con títulos correctos
                data_ejercicios = [[
                    'Ejercicio',
                    'Series',
                    'Repeticiones',
                    'Peso (kg)',
                    'RIR',
                    'Descanso',
                    'Notas'
                ]]
                
                for ej in routine['dias'][dia]['ejercicios']:
                    descanso_min = segundos_a_minutos(ej['descanso'])
                    
                    # Formato compacto: objetivo arriba, espacio abajo más reducido
                    data_ejercicios.append([
                        ej['ejercicio'],
                        f"{ej['series']}\n\n",
                        f"{ej['reps']}\n\n",
                        "\n\n",
                        f"RIR {ej['rir']}\n\n",
                        descanso_min,
                        ej['notas'][:20] if ej['notas'] else ""
                    ])
                
                # Anchos optimizados para caber todo en una página
                tabla_ejercicios = Table(
                    data_ejercicios, 
                    colWidths=[6.5*cm, 1.8*cm, 2.2*cm, 1.8*cm, 1.8*cm, 1.5*cm, 3*cm]
                )
                
                tabla_ejercicios.setStyle(TableStyle([
                    # Header más compacto
                    ('BACKGROUND', (0, 0), (-1, 0), colors.Color(0.25, 0.25, 0.25)),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                    ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
                    ('VALIGN', (0, 0), (-1, 0), 'MIDDLE'),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 9),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
                    ('TOPPADDING', (0, 0), (-1, 0), 8),
                    # Body más compacto
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.Color(0.5, 0.5, 0.5)),
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 8),
                    ('LEFTPADDING', (0, 0), (-1, -1), 5),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 5),
                    # Espaciado reducido pero suficiente para escribir
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 18),
                    ('TOPPADDING', (0, 1), (-1, -1), 6),
                    ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.Color(0.97, 0.97, 0.97)]),
                    # Alineaciones
                    ('ALIGN', (0, 1), (0, -1), 'LEFT'),
                    ('ALIGN', (6, 1), (6, -1), 'LEFT'),
                    ('VALIGN', (0, 1), (-1, -1), 'TOP'),
                ]))
                
                elements.append(tabla_ejercicios)
                elements.append(PageBreak())
    
    doc.build(elements, canvasmaker=MarcaAguaCanvas)
    buffer.seek(0)
    return buffer
    
# ============================================================================
# FUNCIÓN GENERAR EXCEL 
# ============================================================================

def generar_excel_rutina(routine):
    """Genera Excel profesional con planning y hojas por semana"""
    
    wb = Workbook()
    
    # Estilos predefinidos
    header_fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    header_font = Font(name='Calibri', size=11, bold=True, color="FFFFFF")
    
    subheader_fill = PatternFill(start_color="D9D9D9", end_color="D9D9D9", fill_type="solid")
    subheader_font = Font(name='Calibri', size=10, bold=True)
    
    title_font = Font(name='Calibri', size=16, bold=True)
    subtitle_font = Font(name='Calibri', size=12, bold=True)
    normal_font = Font(name='Calibri', size=10)
    small_font = Font(name='Calibri', size=9)
    
    center_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_align = Alignment(horizontal='left', vertical='center', wrap_text=True)
    top_align = Alignment(horizontal='center', vertical='top', wrap_text=True)
    
    thin_border = Border(
        left=Side(style='thin', color='000000'),
        right=Side(style='thin', color='000000'),
        top=Side(style='thin', color='000000'),
        bottom=Side(style='thin', color='000000')
    )
    
    # ========== HOJA 1: PLANNING E INSTRUCCIONES ==========
    ws_planning = wb.active
    ws_planning.title = "Planning y Guía"
    
    cliente_nombre = routine['cliente'] if routine['cliente'] else "Cliente"
    dias_necesarios = [f'Día {i+1}' for i in range(routine['num_dias'])]
    
    # Título
    ws_planning['A1'] = "PROGRAMA DE ENTRENAMIENTO"
    ws_planning['A1'].font = title_font
    ws_planning['A1'].alignment = center_align
    ws_planning.merge_cells('A1:H1')
    ws_planning.row_dimensions[1].height = 25
    
    # Información del cliente
    row = 3
    info_data = [
        ['Cliente:', cliente_nombre],
        ['Duración:', f"{routine['semanas']} semanas"],
        ['Frecuencia:', f"{routine['num_dias']} días por semana"],
        ['Sexo:', routine['metadata']['sexo']],
        ['Fecha creación:', routine['metadata']['fecha_creacion']]
    ]
    
    for label, value in info_data:
        ws_planning[f'B{row}'] = label
        ws_planning[f'B{row}'].font = Font(name='Calibri', size=11, bold=True)
        ws_planning[f'B{row}'].alignment = Alignment(horizontal='right', vertical='center')
        
        ws_planning[f'C{row}'] = value
        ws_planning[f'C{row}'].font = normal_font
        ws_planning[f'C{row}'].alignment = left_align
        ws_planning.merge_cells(f'C{row}:E{row}')
        
        row += 1
    
    row += 2
    
    # PLANNING DE ENTRENAMIENTO
    ws_planning[f'A{row}'] = "PLANNING DE ENTRENAMIENTO"
    ws_planning[f'A{row}'].font = subtitle_font
    ws_planning[f'A{row}'].alignment = center_align
    ws_planning.merge_cells(f'A{row}:H{row}')
    ws_planning.row_dimensions[row].height = 20
    
    row += 1
    
    # Tabla de planning
    max_ejercicios = max([len(routine['dias'].get(dia, {}).get('ejercicios', [])) for dia in dias_necesarios])
    
    # Headers
    for i, dia in enumerate(dias_necesarios):
        col = get_column_letter(i + 1)
        dia_nombre = routine['dias'][dia].get('nombre', '')
        header_text = f"{dia}" + (f" - {dia_nombre}" if dia_nombre else "")
        
        ws_planning[f'{col}{row}'] = header_text
        ws_planning[f'{col}{row}'].font = header_font
        ws_planning[f'{col}{row}'].fill = header_fill
        ws_planning[f'{col}{row}'].alignment = center_align
        ws_planning[f'{col}{row}'].border = thin_border
        ws_planning.column_dimensions[col].width = 25
    
    ws_planning.row_dimensions[row].height = 30
    row += 1
    
    # Ejercicios
    for i in range(max_ejercicios):
        for j, dia in enumerate(dias_necesarios):
            col = get_column_letter(j + 1)
            ejercicios = routine['dias'].get(dia, {}).get('ejercicios', [])
            
            if i < len(ejercicios):
                ws_planning[f'{col}{row}'] = ejercicios[i]['ejercicio']
            else:
                ws_planning[f'{col}{row}'] = ""
            
            ws_planning[f'{col}{row}'].font = small_font
            ws_planning[f'{col}{row}'].alignment = top_align
            ws_planning[f'{col}{row}'].border = thin_border
        
        ws_planning.row_dimensions[row].height = 30
        row += 1
    
    row += 2
    
    # INSTRUCCIONES RESUMIDAS
    ws_planning[f'A{row}'] = "INSTRUCCIONES DE USO - RESUMEN"
    ws_planning[f'A{row}'].font = subtitle_font
    ws_planning[f'A{row}'].alignment = center_align
    ws_planning.merge_cells(f'A{row}:H{row}')
    ws_planning.row_dimensions[row].height = 20
    
    row += 1
    
    instrucciones = [
        ("1. REGISTRO DE DATOS", 
         "Anota en cada serie: Peso utilizado (kg), Repeticiones completadas, RIR (0=fallo, 1=podías 1 más, etc.), y Notas sobre sensaciones."),
        
        ("2. PROGRESIÓN CON RANGO (ej: 6-8 reps)", 
         "Empieza con el mínimo (6 reps) al RIR objetivo. Aumenta reps hasta el máximo (8) con el mismo peso. "
         "Cuando logres 8 reps al RIR objetivo, aumenta peso y vuelve a 6 reps."),
        
        ("3. PROGRESIÓN CON NÚMERO FIJO (ej: 10 reps)", 
         "Aumenta solo el peso cuando completes todas las series al RIR objetivo con buena técnica."),
        
        ("4. CALENTAMIENTO", 
         "IMPORTANTE: Siempre calienta con gomas o pesas de bajo peso e intensidad baja durante 5-10 minutos.")
    ]
    
    for titulo, texto in instrucciones:
        ws_planning[f'A{row}'] = titulo
        ws_planning[f'A{row}'].font = Font(name='Calibri', size=10, bold=True)
        ws_planning[f'A{row}'].alignment = left_align
        ws_planning.merge_cells(f'A{row}:H{row}')
        ws_planning.row_dimensions[row].height = 20
        
        row += 1
        
        ws_planning[f'A{row}'] = texto
        ws_planning[f'A{row}'].font = small_font
        ws_planning[f'A{row}'].alignment = left_align
        ws_planning.merge_cells(f'A{row}:H{row}')
        ws_planning.row_dimensions[row].height = 40
        
        row += 1
    
    # ========== HOJAS POR SEMANA ==========
    for semana in range(1, routine['semanas'] + 1):
        ws = wb.create_sheet(title=f"Semana {semana}")
        
        # Título de la semana
        ws['A1'] = f"RUTINA {cliente_nombre.upper()} - SEMANA {semana}"
        ws['A1'].font = title_font
        ws['A1'].alignment = center_align
        ws.merge_cells('A1:H1')
        ws.row_dimensions[1].height = 25
        
        current_row = 3
        
        # Por cada día de la semana
        for dia in dias_necesarios:
            if routine['dias'].get(dia, {}).get('ejercicios', []):
                dia_nombre = routine['dias'][dia].get('nombre', '')
                
                # Título del día
                titulo_dia = f"{dia.upper()}" + (f" - {dia_nombre.upper()}" if dia_nombre else "")
                ws[f'A{current_row}'] = titulo_dia
                ws[f'A{current_row}'].font = subtitle_font
                ws[f'A{current_row}'].alignment = center_align
                ws[f'A{current_row}'].fill = subheader_fill
                ws.merge_cells(f'A{current_row}:H{current_row}')
                ws.row_dimensions[current_row].height = 20
                
                current_row += 1
                
                # Headers de la tabla
                headers = ['Ejercicio', 'Series', 'Repeticiones', 'Peso (kg)', 'RIR', 'Descanso', 'Notas']
                for i, header in enumerate(headers):
                    col = get_column_letter(i + 1)
                    ws[f'{col}{current_row}'] = header
                    ws[f'{col}{current_row}'].font = header_font
                    ws[f'{col}{current_row}'].fill = header_fill
                    ws[f'{col}{current_row}'].alignment = center_align
                    ws[f'{col}{current_row}'].border = thin_border
                
                # Anchos de columna
                ws.column_dimensions['A'].width = 35
                ws.column_dimensions['B'].width = 12
                ws.column_dimensions['C'].width = 15
                ws.column_dimensions['D'].width = 12
                ws.column_dimensions['E'].width = 12
                ws.column_dimensions['F'].width = 12
                ws.column_dimensions['G'].width = 25
                
                ws.row_dimensions[current_row].height = 30
                current_row += 1
                
                # Ejercicios del día
                for ej in routine['dias'][dia]['ejercicios']:
                    descanso_min = segundos_a_minutos(ej['descanso'])
                    
                    # Fila con datos del ejercicio
                    # Ejercicio (A)
                    ws[f'A{current_row}'] = ej['ejercicio']
                    ws[f'A{current_row}'].font = normal_font
                    ws[f'A{current_row}'].alignment = left_align
                    ws[f'A{current_row}'].border = thin_border
                    
                    # Series (B) - Objetivo arriba
                    ws[f'B{current_row}'] = f"Obj: {ej['series']}\n\nReal:"
                    ws[f'B{current_row}'].font = small_font
                    ws[f'B{current_row}'].alignment = top_align
                    ws[f'B{current_row}'].border = thin_border
                    
                    # Repeticiones (C) - Objetivo arriba
                    ws[f'C{current_row}'] = f"Obj: {ej['reps']}\n\nReal:"
                    ws[f'C{current_row}'].font = small_font
                    ws[f'C{current_row}'].alignment = top_align
                    ws[f'C{current_row}'].border = thin_border
                    
                    # Peso (D) - En blanco completo
                    ws[f'D{current_row}'] = ""
                    ws[f'D{current_row}'].border = thin_border
                    
                    # RIR (E) - Objetivo arriba
                    ws[f'E{current_row}'] = f"Obj: RIR {ej['rir']}\n\nReal:"
                    ws[f'E{current_row}'].font = small_font
                    ws[f'E{current_row}'].alignment = top_align
                    ws[f'E{current_row}'].border = thin_border
                    
                    # Descanso (F)
                    ws[f'F{current_row}'] = descanso_min
                    ws[f'F{current_row}'].font = normal_font
                    ws[f'F{current_row}'].alignment = center_align
                    ws[f'F{current_row}'].border = thin_border
                    
                    # Notas (G)
                    ws[f'G{current_row}'] = ej['notas'] if ej['notas'] else ""
                    ws[f'G{current_row}'].font = small_font
                    ws[f'G{current_row}'].alignment = left_align
                    ws[f'G{current_row}'].border = thin_border
                    
                    # Altura de fila generosa para escribir
                    ws.row_dimensions[current_row].height = 50
                    
                    current_row += 1
                
                # Espacio entre días
                current_row += 1
    
    # Guardar en buffer
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================================================
# INICIALIZACIÓN SESSION STATE
# ============================================================================

if 'routine' not in st.session_state:
    st.session_state.routine = {
        'cliente': '',
        'semanas': 8,
        'dias': {},
        'metadata': {},
        'num_dias': 4,
        'vista': 'planificacion'
    }
# ============================================================================
# NAVEGACIÓN SUPERIOR
# ============================================================================

col_nav1, col_nav2, col_nav3 = st.columns([1, 1, 4])

with col_nav1:
    if st.button("📅 Planificación", use_container_width=True, 
                 type="primary" if st.session_state.routine['vista'] == 'planificacion' else "secondary"):
        st.session_state.routine['vista'] = 'planificacion'
        st.rerun()

with col_nav2:
    if st.button("👁️ Visualización", use_container_width=True,
                 type="primary" if st.session_state.routine['vista'] == 'visualizacion' else "secondary"):
        st.session_state.routine['vista'] = 'visualizacion'
        st.rerun()

st.markdown("---")

# ============================================================================
# SIDEBAR: CONFIGURACIÓN
# ============================================================================
with st.sidebar:
    st.header("📋 Configuración")
    
    st.session_state.routine['cliente'] = st.text_input(
        "Nombre completo", 
        st.session_state.routine.get('cliente', ''),
        placeholder="Ej: Juan Pérez"
    )
    
    col1, col2 = st.columns(2)
    with col1:
        semanas = st.number_input("Semanas", 4, 16, 8)
    with col2:
        num_dias_nuevo = st.number_input("Días/sem", 2, 7, st.session_state.routine['num_dias'])
        
        # SI CAMBIA EL NÚMERO DE DÍAS
        if num_dias_nuevo != st.session_state.routine['num_dias']:
            dias_actuales = list(st.session_state.routine['dias'].keys())
            dias_necesarios = [f'Día {i+1}' for i in range(num_dias_nuevo)]
            
            # ELIMINAR días que sobran
            for dia in dias_actuales:
                if dia not in dias_necesarios:
                    del st.session_state.routine['dias'][dia]
            
            # CREAR días que faltan
            for dia in dias_necesarios:
                if dia not in st.session_state.routine['dias']:
                    st.session_state.routine['dias'][dia] = {'nombre': '', 'ejercicios': []}
            
            # Actualizar contador
            st.session_state.routine['num_dias'] = num_dias_nuevo
            st.rerun()
    
    st.session_state.routine['semanas'] = semanas
    
    st.markdown("---")
    
    sexo = st.selectbox("Sexo", ["Hombre", "Mujer"])
    experiencia = st.selectbox(
        "Experiencia",
        ["Novato (<1 año)", "Intermedio (1-3 años)", "Avanzado (3+ años)"]
    )
    
    st.session_state.routine['metadata'] = {
        'sexo': sexo,
        'experiencia': experiencia,
        'fecha_creacion': datetime.now().strftime("%Y-%m-%d %H:%M")
    }
    
    st.markdown("---")
    
    total_ejercicios = sum([len(data.get('ejercicios', [])) for data in st.session_state.routine['dias'].values()])
    if total_ejercicios > 0:
        st.success(f"✅ {total_ejercicios} ejercicios")
    else:
        st.info("ℹ️ Sin ejercicios")
    
    st.caption(f"📅 {st.session_state.routine['metadata']['fecha_creacion']}")

# ============================================================================
# VISTA: PLANIFICACIÓN SEMANAL
# ============================================================================
if st.session_state.routine['vista'] == 'planificacion':
    
    st.title("📅 Planificación Semanal")
    st.caption(f"Cliente: **{st.session_state.routine['cliente'] if st.session_state.routine['cliente'] else 'Sin nombre'}**")
    
    st.markdown("---")
    
    routine = st.session_state.routine
    
    # Inicializar días con estructura mejorada
    dias_necesarios = [f'Día {i+1}' for i in range(routine['num_dias'])]
    
    # IMPORTANTE: Asegurar que todos los días existen
    for dia in dias_necesarios:
        if dia not in routine['dias']:
            routine['dias'][dia] = {'nombre': '', 'ejercicios': []}
        # Validar estructura
        elif 'ejercicios' not in routine['dias'][dia]:
            routine['dias'][dia]['ejercicios'] = []
        elif not isinstance(routine['dias'][dia]['ejercicios'], list):
            routine['dias'][dia]['ejercicios'] = []
    
    # Limpiar días extras
    dias_actuales = list(routine['dias'].keys())
    for dia in dias_actuales:
        if dia not in dias_necesarios:
            del routine['dias'][dia]
    
    # TABS POR DÍAS
    tabs = st.tabs(dias_necesarios)
    
    for idx, dia in enumerate(dias_necesarios):
        with tabs[idx]:
            # Nombre del día
            col_titulo, col_nombre = st.columns([1, 3])
            with col_titulo:
                st.markdown(f"### {dia}")
            with col_nombre:
                routine['dias'][dia]['nombre'] = st.text_input(
                    "Nombre del día (opcional)", 
                    routine['dias'][dia].get('nombre', ''),
                    key=f"nombre_{dia}",
                    placeholder="Ej: Pecho, Espalda, Tren Superior..."
                )
            
            st.markdown("---")
            
            # Método 1: Selección jerárquica
            st.markdown("#### ➕ Método 1: Selección por categorías")
            
            col1, col2, col3 = st.columns(3)
            
            with col1:
                grupo_sel = st.selectbox("Grupo muscular", GRUPOS_MUSCULARES, key=f"grupo_{dia}")
            
            with col2:
                regiones_disponibles = REGIONES_MAP.get(grupo_sel, [])
                region_sel = st.selectbox("Región específica", regiones_disponibles, key=f"region_{dia}")
            
            with col3:
                ejercicios_disponibles = EJERCICIOS_DB.get(grupo_sel, {}).get(region_sel, [])
                if ejercicios_disponibles:
                    ejercicio_sel = st.selectbox("Ejercicio", ejercicios_disponibles, key=f"ej_{dia}")
                else:
                    st.warning("No hay ejercicios")
                    ejercicio_sel = None
            
            # Método 2: Búsqueda por nombre
            st.markdown("#### 🔍 Método 2: Búsqueda rápida por nombre")
            
            busqueda = st.text_input(
                "Escribe el nombre del ejercicio", 
                key=f"busqueda_{dia}",
                placeholder="Ej: press banca, dominadas, sentadilla..."
            )
            
            ejercicio_buscado = None
            if busqueda:
                resultados = [ej for ej in TODOS_EJERCICIOS if busqueda.lower() in ej['nombre'].lower()]
                if resultados:
                    nombres_resultados = [ej['nombre'] for ej in resultados]
                    ejercicio_seleccionado = st.selectbox(
                        f"Resultados ({len(resultados)} encontrados)",
                        nombres_resultados,
                        key=f"resultado_{dia}"
                    )
                    ejercicio_buscado = next(ej for ej in resultados if ej['nombre'] == ejercicio_seleccionado)
                else:
                    st.info("No se encontraron ejercicios con ese nombre")
            
            # Formulario común para ambos métodos
            st.markdown("---")
            
            with st.form(f"form_{dia}", clear_on_submit=False):
                col_p1, col_p2, col_p3, col_p4 = st.columns(4)
                
                with col_p1:
                    series = st.number_input("Series", 1, 12, 3, key=f"series_{dia}")
                with col_p2:
                    reps = st.text_input("Reps", "8-12", key=f"reps_{dia}")
                with col_p3:
                    rir = st.selectbox("RIR", [0, 1, 2, 3, 4], index=2, key=f"rir_{dia}")
                with col_p4:
                    descanso_min = st.number_input("Descanso (min)", 0.5, 5.0, 2.0, step=0.5, key=f"desc_{dia}")
                
                notas = st.text_input("📝 Notas", key=f"notas_{dia}", placeholder="Opcional...")
                
                col_btn1, col_btn2 = st.columns(2)
                with col_btn1:
                    submitted1 = st.form_submit_button("➕ Añadir desde categorías", use_container_width=True, type="primary")
                with col_btn2:
                    submitted2 = st.form_submit_button("➕ Añadir desde búsqueda", use_container_width=True)
                
                # Convertir minutos a segundos para almacenar
                descanso_segundos = int(descanso_min * 60)
                
                if submitted1 and ejercicio_sel:
                    grupo_directo = get_grupo_directo(grupo_sel, region_sel)
                    nuevo_ej = {
                        'ejercicio': ejercicio_sel,
                        'grupo': grupo_sel,
                        'region': region_sel,
                        'grupo_directo': grupo_directo,
                        'series': series,
                        'reps': reps,
                        'rir': rir,
                        'descanso': descanso_segundos,
                        'notas': notas
                    }
                    routine['dias'][dia]['ejercicios'].append(nuevo_ej)
                    st.success(f"✅ {ejercicio_sel} añadido")
                    st.rerun()
                
                if submitted2 and ejercicio_buscado:
                    grupo_directo = get_grupo_directo(ejercicio_buscado['grupo'], ejercicio_buscado['region'])
                    nuevo_ej = {
                        'ejercicio': ejercicio_buscado['nombre'],
                        'grupo': ejercicio_buscado['grupo'],
                        'region': ejercicio_buscado['region'],
                        'grupo_directo': grupo_directo,
                        'series': series,
                        'reps': reps,
                        'rir': rir,
                        'descanso': descanso_segundos,
                        'notas': notas
                    }
                    routine['dias'][dia]['ejercicios'].append(nuevo_ej)
                    st.success(f"✅ {ejercicio_buscado['nombre']} añadido")
                    st.rerun()
            
            st.markdown("---")
            
            # Mostrar ejercicios
            if routine['dias'][dia]['ejercicios']:
                st.markdown("#### 📋 Ejercicios programados")
                
                ejercicios_data = []
                for i, ej in enumerate(routine['dias'][dia]['ejercicios']):
                    ejercicios_data.append({
                        'N°': i+1,
                        'Ejercicio': ej['ejercicio'],
                        'Series': ej['series'],
                        'Reps': ej['reps'],
                        'RIR': ej['rir'],
                        'Desc (min)': round(ej['descanso'] / 60, 1),
                        'Notas': ej['notas'][:20] + "..." if len(ej['notas']) > 20 else ej['notas']
                    })
                
                df_ejercicios = pd.DataFrame(ejercicios_data)
                
                edited_df = st.data_editor(
                    df_ejercicios,
                    use_container_width=True,
                    hide_index=True,
                    num_rows="fixed",
                    disabled=["N°", "Ejercicio"],
                    key=f"editor_{dia}",
                    column_config={
                        "N°": st.column_config.NumberColumn("N°", width="small"),
                        "Ejercicio": st.column_config.TextColumn("Ejercicio", width="large"),
                        "Series": st.column_config.NumberColumn("Series", width="small", min_value=1, max_value=12),
                        "Reps": st.column_config.TextColumn("Reps", width="small"),
                        "RIR": st.column_config.NumberColumn("RIR", width="small", min_value=0, max_value=4),
                        "Desc (min)": st.column_config.NumberColumn("Desc (min)", width="small", min_value=0.5, max_value=5.0, step=0.5),
                        "Notas": st.column_config.TextColumn("Notas", width="medium"),
                    }
                )
                
                if not edited_df.equals(df_ejercicios):
                    for i, row in edited_df.iterrows():
                        routine['dias'][dia]['ejercicios'][i]['series'] = int(row['Series'])
                        routine['dias'][dia]['ejercicios'][i]['reps'] = str(row['Reps'])
                        routine['dias'][dia]['ejercicios'][i]['rir'] = int(row['RIR'])
                        routine['dias'][dia]['ejercicios'][i]['descanso'] = int(row['Desc (min)'] * 60)
                        routine['dias'][dia]['ejercicios'][i]['notas'] = str(row['Notas']).replace("...", "")
                
                st.markdown("##### Acciones:")
                cols_action = st.columns([2, 2, 2, 2])
                
                with cols_action[0]:
                    ejercicio_mover = st.selectbox("Ejercicio N°", range(1, len(routine['dias'][dia]['ejercicios']) + 1), key=f"mover_{dia}")
                with cols_action[1]:
                    if st.button("⬆️", key=f"up_{dia}", use_container_width=True):
                        idx_move = ejercicio_mover - 1
                        if idx_move > 0:
                            routine['dias'][dia]['ejercicios'][idx_move], routine['dias'][dia]['ejercicios'][idx_move-1] = \
                                routine['dias'][dia]['ejercicios'][idx_move-1], routine['dias'][dia]['ejercicios'][idx_move]
                            st.rerun()
                with cols_action[2]:
                    if st.button("⬇️", key=f"down_{dia}", use_container_width=True):
                        idx_move = ejercicio_mover - 1
                        if idx_move < len(routine['dias'][dia]['ejercicios']) - 1:
                            routine['dias'][dia]['ejercicios'][idx_move], routine['dias'][dia]['ejercicios'][idx_move+1] = \
                                routine['dias'][dia]['ejercicios'][idx_move+1], routine['dias'][dia]['ejercicios'][idx_move]
                            st.rerun()
                with cols_action[3]:
                    if st.button("🗑️", key=f"del_{dia}", use_container_width=True):
                        routine['dias'][dia]['ejercicios'].pop(ejercicio_mover - 1)
                        st.rerun()
                
                st.markdown("---")
                
                # Volumen del día
                volumen_dia = calcular_volumen_por_dia(routine, dia)
                volumen_dia_activo = {k: v for k, v in volumen_dia.items() if v > 0}
                
                if volumen_dia_activo:
                    st.markdown("##### 📊 Volumen del día")
                    num_cols = min(len(volumen_dia_activo), 5)
                    cols = st.columns(num_cols)
                    for i, (grupo, series_vol) in enumerate(volumen_dia_activo.items()):
                        with cols[i % num_cols]:
                            st.metric(grupo, f"{series_vol}s")
            else:
                st.info("👆 No hay ejercicios programados")
    
    # Resumen rápido
    st.markdown("---")
    st.header("📊 Resumen Rápido")
    
    volumen_semanal = calcular_volumen_semanal(routine)
    deficit, exceso, optimo = validar_volumen(volumen_semanal)
    
    col1, col2, col3, col4 = st.columns(4)
    with col1:
        total_series = sum([v for v in volumen_semanal.values()])
        st.metric("Total series/semana", total_series)
    with col2:
        st.metric("Grupos óptimos", len(optimo))
    with col3:
        st.metric("Déficit", len(deficit), delta_color="inverse")
    with col4:
        st.metric("Exceso", len(exceso), delta_color="inverse")
    
    if deficit:
        st.warning(f"⚠️ **Déficit:** {', '.join(deficit)}")
    if exceso:
        st.error(f"🔴 **Exceso:** {', '.join(exceso)}")
    if not deficit and not exceso and total_series > 0:
        st.success("✅ Todos los grupos en rango óptimo")

# ============================================================================
# VISTA: VISUALIZACIÓN COMPLETA
# ============================================================================
elif st.session_state.routine['vista'] == 'visualizacion':
    
    st.title("👁️ Visualización Completa")
    
    routine = st.session_state.routine
    
    # Verificar que hay ejercicios
    total_ejercicios = sum([len(data.get('ejercicios', [])) for data in routine['dias'].values()])
    if total_ejercicios == 0:
        st.warning("⚠️ No hay ejercicios en la rutina. Ve a Planificación Semanal para añadir ejercicios.")
        st.stop()
    
    st.markdown("---")
    
    # Datos cliente
    col_info1, col_info2, col_info3, col_info4 = st.columns(4)
    with col_info1:
        st.metric("Cliente", routine['cliente'] if routine['cliente'] else "Sin nombre")
    with col_info2:
        st.metric("Duración", f"{routine['semanas']} semanas")
    with col_info3:
        st.metric("Frecuencia", f"{routine['num_dias']} días/semana")
    with col_info4:
        st.metric("Sexo", routine['metadata']['sexo'])
    
    st.markdown("---")
    
    # Rutina completa
    st.header("📅 Rutina Semanal Completa")
    
    dias_necesarios = [f"Día {i+1}" for i in range(routine['num_dias'])]
    
    for dia in dias_necesarios:
        if routine['dias'].get(dia, {}).get('ejercicios', []):
            dia_nombre = routine['dias'][dia].get('nombre', '')
            titulo_dia = f"{dia}" + (f" - {dia_nombre}" if dia_nombre else "")
            st.subheader(titulo_dia)
            
            ejercicios_data = []
            for i, ej in enumerate(routine['dias'][dia]['ejercicios']):
                ejercicios_data.append({
                    'N°': i+1,
                    'Ejercicio': ej['ejercicio'],
                    'Grupo': f"{ej['grupo']} → {ej['region']}",
                    'Series': ej['series'],
                    'Reps': ej['reps'],
                    'RIR': ej['rir'],
                    'Descanso': segundos_a_minutos(ej['descanso']),
                    'Notas': ej['notas'] if ej['notas'] else "-"
                })
            
            df_dia = pd.DataFrame(ejercicios_data)
            
            st.dataframe(
                df_dia,
                use_container_width=True,
                hide_index=True,
                column_config={
                    "N°": st.column_config.NumberColumn("N°", width="small"),
                    "Ejercicio": st.column_config.TextColumn("Ejercicio", width="large"),
                    "Grupo": st.column_config.TextColumn("Grupo", width="medium"),
                    "Series": st.column_config.NumberColumn("Series", width="small"),
                    "Reps": st.column_config.TextColumn("Reps", width="small"),
                    "RIR": st.column_config.NumberColumn("RIR", width="small"),
                    "Descanso": st.column_config.TextColumn("Descanso", width="small"),
                    "Notas": st.column_config.TextColumn("Notas", width="medium"),
                }
            )
            
            # Volumen del día
            volumen_dia = calcular_volumen_por_dia(routine, dia)
            volumen_dia_activo = {k: v for k, v in volumen_dia.items() if v > 0}
            
            if volumen_dia_activo:
                st.markdown(f"**📊 Volumen {dia}:**")
                cols = st.columns(len(volumen_dia_activo))
                for i, (grupo, series_vol) in enumerate(volumen_dia_activo.items()):
                    with cols[i]:
                        st.metric(grupo, f"{series_vol} series")
            
            st.markdown("---")
    
    # Análisis volumen semanal
    st.header("📊 Análisis de Volumen Semanal")
    
    volumen_semanal = calcular_volumen_semanal(routine)
    deficit, exceso, optimo = validar_volumen(volumen_semanal)
    
    col_m1, col_m2, col_m3, col_m4 = st.columns(4)
    with col_m1:
        total_series = sum([v for v in volumen_semanal.values()])
        st.metric("Total series/semana", total_series)
    with col_m2:
        st.metric("Grupos óptimos", len(optimo))
    with col_m3:
        st.metric("Déficit", len(deficit), delta_color="inverse")
    with col_m4:
        st.metric("Exceso", len(exceso), delta_color="inverse")
    
    st.markdown("---")
    
    col_tabla, col_grafico = st.columns([1, 2])
    
    with col_tabla:
        st.markdown("### Tabla de volumen")
        
        df_volumen = pd.DataFrame([
            {
                'Grupo': k,
                'Series': v,
                'Min': VOLUMEN_OPTIMO[k]['min'],
                'Max': VOLUMEN_OPTIMO[k]['max'],
                '✓': '✅' if VOLUMEN_OPTIMO[k]['min'] <= v <= VOLUMEN_OPTIMO[k]['max'] and v > 0
                     else ('⚠️' if v < VOLUMEN_OPTIMO[k]['min'] and v > 0 else ('🔴' if v > VOLUMEN_OPTIMO[k]['max'] else '-'))
            }
            for k, v in volumen_semanal.items()
        ])
        
        df_volumen = df_volumen.sort_values('Series', ascending=False)
        
        st.dataframe(df_volumen, use_container_width=True, hide_index=True, height=500)
    
    with col_grafico:
        st.markdown("### Gráfico de volumen")
        
        df_volumen_grafico = df_volumen[df_volumen['Series'] > 0]
        
        if not df_volumen_grafico.empty:
            fig = go.Figure()
            
            colors_list = []
            for _, row in df_volumen_grafico.iterrows():
                if row['Series'] >= row['Min'] and row['Series'] <= row['Max']:
                    colors_list.append('green')
                elif row['Series'] < row['Min']:
                    colors_list.append('orange')
                else:
                    colors_list.append('red')
            
            fig.add_trace(go.Bar(
                x=df_volumen_grafico['Grupo'],
                y=df_volumen_grafico['Series'],
                name='Series actuales',
                marker_color=colors_list,
                text=df_volumen_grafico['Series'],
                textposition='outside',
            ))
            
            fig.add_trace(go.Scatter(
                x=df_volumen_grafico['Grupo'],
                y=df_volumen_grafico['Min'],
                name='Mínimo',
                mode='lines+markers',
                line=dict(color='blue', dash='dash', width=2),
                marker=dict(size=6)
            ))
            
            fig.add_trace(go.Scatter(
                x=df_volumen_grafico['Grupo'],
                y=df_volumen_grafico['Max'],
                name='Máximo',
                mode='lines+markers',
                line=dict(color='red', dash='dash', width=2),
                marker=dict(size=6)
            ))
            
            fig.update_layout(
                xaxis_title='Grupo Muscular',
                yaxis_title='Series semanales',
                barmode='group',
                height=500,
                hovermode='x unified',
                legend=dict(orientation="h", yanchor="bottom", y=1.02, xanchor="right", x=1)
            )
            
            st.plotly_chart(fig, use_container_width=True)
    
    st.markdown("---")
    
    col_alert1, col_alert2 = st.columns(2)
    with col_alert1:
        if deficit:
            st.warning(f"**⚠️ Déficit:**\n" + "\n".join([f"- {g}" for g in deficit]))
        else:
            st.success("✅ Sin déficit")
    with col_alert2:
        if exceso:
            st.error(f"**🔴 Exceso:**\n" + "\n".join([f"- {g}" for g in exceso]))
        else:
            st.success("✅ Sin exceso")
    
    # Botones exportación
    st.markdown("---")
    st.header("📤 Exportar Rutina")
    
    col_btn1, col_btn2 = st.columns(2)
    
    with col_btn1:
        if st.button("📄 Generar PDF", use_container_width=True, type="primary"):
            with st.spinner("Generando PDF profesional..."):
                pdf_buffer = generar_pdf_rutina(routine)
                st.success("✅ PDF generado correctamente")
                
                st.download_button(
                    label="⬇️ Descargar PDF",
                    data=pdf_buffer,
                    file_name=f"Rutina_{routine['cliente'].replace(' ', '_') if routine['cliente'] else 'Cliente'}_{datetime.now().strftime('%Y%m%d')}.pdf",
                    mime="application/pdf",
                    use_container_width=True
                )
    
    with col_btn2:
        if st.button("📊 Generar Excel Seguimiento", use_container_width=True, type="primary"):
            with st.spinner("Generando Excel profesional..."):
                excel_buffer = generar_excel_rutina(routine)
                st.success("✅ Excel generado correctamente")
                
                st.download_button(
                    label="⬇️ Descargar Excel",
                    data=excel_buffer,
                    file_name=f"Seguimiento_{routine['cliente'].replace(' ', '_') if routine['cliente'] else 'Cliente'}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                    mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    use_container_width=True
                )
# ============================================================================
# FOOTER
# ============================================================================
st.markdown("---")
st.caption("🔬 **Constructor de Rutina** | David López - Entrenador Personal")

