import math
import streamlit as st
import pandas as pd
import plotly.graph_objects as go
from datetime import date, datetime
import sqlite3
import pickle
import os

from reportlab.lib.pagesizes import A4
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch, cm
from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, PageBreak, Image as RLImage
from reportlab.lib import colors
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_JUSTIFY
import io

import plotly.graph_objects as go
import matplotlib.pyplot as plt
import matplotlib

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

st.set_page_config(page_title="Evaluación Integral", page_icon="💪", layout="wide")

# ============================================
# FUNCIONES AUXILIARES
# ============================================

def epley_1rm(weight, reps):
    if weight is None or reps is None or reps <= 0:
        return None
    return weight * (1 + reps / 30.0)

def brzycki_1rm(weight, reps):
    if weight is None or reps is None or reps <= 0 or reps >= 37:
        return None
    return weight / (1.0278 - 0.0278 * reps)

def navy_bodyfat_male(waist_cm, neck_cm, height_cm):
    try:
        if waist_cm <= neck_cm or height_cm <= 0:
            return None
        return 86.010 * math.log10(waist_cm - neck_cm) - 70.041 * math.log10(height_cm) + 36.76
    except:
        return None

def navy_bodyfat_female(waist_cm, hip_cm, neck_cm, height_cm):
    try:
        if (waist_cm + hip_cm) <= neck_cm or height_cm <= 0:
            return None
        return 163.205 * math.log10(waist_cm + hip_cm - neck_cm) - 97.684 * math.log10(height_cm) - 78.387
    except:
        return None

def mifflin_bmr(sex, weight_kg, height_cm, age):
    if sex == "Hombre":
        return 10 * weight_kg + 6.25 * height_cm - 5 * age + 5
    else:
        return 10 * weight_kg + 6.25 * height_cm - 5 * age - 161

def tdee_from_factor(bmr, factor):
    return bmr * factor

def safe_select_index(options, value, default_index=0):
    if value is None:
        return default_index
    try:
        return options.index(value)
    except (ValueError, AttributeError):
        pass
    value_lower = str(value).lower()
    for idx, option in enumerate(options):
        if value_lower in option.lower() or option.lower() in value_lower:
            return idx
    return default_index

# ============================================
# ENVÍO AUTOMÁTICO POR EMAIL
# ============================================

import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

def send_evaluation_email(client_name, client_email, professional_email, pdf_buffer, excel_buffer, 
                          smtp_server="smtp.gmail.com", smtp_port=587, 
                          sender_email=None, sender_password=None):
    """
    Envía los informes PDF y Excel por email al profesional y al cliente
    
    Args:
        client_name: Nombre del cliente
        client_email: Email del cliente
        professional_email: Email del profesional (donde se envían los informes)
        pdf_buffer: Buffer del PDF generado
        excel_buffer: Buffer del Excel generado
        smtp_server: Servidor SMTP (por defecto Gmail)
        smtp_port: Puerto SMTP (por defecto 587 para TLS)
        sender_email: Email del remitente (tu email profesional)
        sender_password: Contraseña de aplicación de Gmail
    """
    
    try:
        # Crear mensaje
        msg = MIMEMultipart()
        msg['From'] = sender_email
        msg['To'] = professional_email
        msg['Cc'] = client_email  # Copia al cliente
        msg['Subject'] = f"Evaluación Física Completa - {client_name}"
        
        # Cuerpo del email
        body = f"""
Hola,

Se ha completado la evaluación física inicial de {client_name}.

📋 RESUMEN DE LA EVALUACIÓN
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
• Cliente: {client_name}
• Email: {client_email}
• Fecha y hora: {datetime.now().strftime('%d/%m/%Y a las %H:%M')}

📎 DOCUMENTACIÓN ADJUNTA
━━━━━━━━━━━━━━━━━━━━━━━━━━━━
✓ Informe Completo PDF 

✓ Base de Datos Excel 


───────────────────────────────────────────
Sistema de Evaluación Integral v5.0

Este correo ha sido generado automáticamente.
Para cualquier consulta, por favor responda a este email.
"""
        
        msg.attach(MIMEText(body, 'plain'))
        
        # Adjuntar PDF
        pdf_buffer.seek(0)
        pdf_attachment = MIMEBase('application', 'pdf')
        pdf_attachment.set_payload(pdf_buffer.read())
        encoders.encode_base64(pdf_attachment)
        pdf_attachment.add_header(
            'Content-Disposition',
            f'attachment; filename=informe_{client_name.replace(" ", "_")}.pdf'
        )
        msg.attach(pdf_attachment)
        
        # Adjuntar Excel
        excel_buffer.seek(0)
        excel_attachment = MIMEBase('application', 'vnd.openxmlformats-officedocument.spreadsheetml.sheet')
        excel_attachment.set_payload(excel_buffer.read())
        encoders.encode_base64(excel_attachment)
        excel_attachment.add_header(
            'Content-Disposition',
            f'attachment; filename=evaluacion_{client_name.replace(" ", "_")}.xlsx'
        )
        msg.attach(excel_attachment)
        
        # Conectar al servidor SMTP y enviar
        server = smtplib.SMTP(smtp_server, smtp_port)
        server.starttls()  # Seguridad TLS
        server.login(sender_email, sender_password)
        
        # Enviar a profesional y cliente
        recipients = [professional_email, client_email]
        server.sendmail(sender_email, recipients, msg.as_string())
        server.quit()
        
        return True, "Email enviado correctamente"
        
    except Exception as e:
        return False, f"Error al enviar email: {str(e)}"

# ============================================
# GENERACIÓN DE INFORME PDF COMPLETO Y PROFESIONAL
# ============================================


def generate_complete_professional_pdf(client_data, age, bmi, whr, bf, bmr, tdee_maintenance, 
                                       protein_avg, fat_avg, carbs_avg, alerts, strength_tests_data=None):
    """
    Genera un informe PDF COMPLETO con TODOS los datos del formulario + todos los cálculos + gráficos
    """
    buffer = io.BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4, rightMargin=2*cm, leftMargin=2*cm, topMargin=2*cm, bottomMargin=2*cm)
    
    styles = getSampleStyleSheet()
    
    # Estilos personalizados
    title_style = ParagraphStyle('CustomTitle', parent=styles['Heading1'], fontSize=24, 
                                 textColor=colors.HexColor('#1A5490'), spaceAfter=30, 
                                 alignment=TA_CENTER, fontName='Helvetica-Bold')
    
    heading_style = ParagraphStyle('CustomHeading', parent=styles['Heading2'], fontSize=16, 
                                   textColor=colors.HexColor('#2E86C1'), spaceAfter=12, 
                                   spaceBefore=12, fontName='Helvetica-Bold')
    
    subheading_style = ParagraphStyle('CustomSubHeading', parent=styles['Heading3'], fontSize=13, 
                                      textColor=colors.HexColor('#34495E'), spaceAfter=10, fontName='Helvetica-Bold')
    
    body_style = ParagraphStyle('CustomBody', parent=styles['BodyText'], fontSize=10, leading=14, 
                                spaceAfter=10, alignment=TA_JUSTIFY)
    
    story = []
    
    # ============================================
    # PORTADA
    # ============================================
    story.append(Spacer(1, 1.5*inch))
    story.append(Paragraph("INFORME COMPLETO DE EVALUACIÓN FÍSICA", title_style))
    story.append(Spacer(1, 0.3*inch))
    story.append(Paragraph(f"<b>Cliente:</b> {client_data.get('name', 'N/A')}", 
                          ParagraphStyle('client', parent=body_style, alignment=TA_CENTER, fontSize=14)))
    story.append(Paragraph(f"<b>Fecha del informe:</b> {datetime.now().strftime('%d de %B de %Y')}", 
                          ParagraphStyle('date', parent=body_style, alignment=TA_CENTER, fontSize=12)))
    story.append(Spacer(1, 0.5*inch))
    story.append(Paragraph("Sistema de Evaluación Integral | Basado en Evidencia Científica 2024-2025", 
                          ParagraphStyle('footer', alignment=TA_CENTER, fontSize=8, textColor=colors.grey)))
    story.append(PageBreak())
    
    # ============================================
    # PARTE 1: DATOS COMPLETOS DEL CLIENTE
    # ============================================
    story.append(Paragraph("PARTE 1: INFORMACIÓN COMPLETA DEL CLIENTE", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    # 1.1 INFORMACIÓN PERSONAL
    story.append(Paragraph("1.1 Información Personal", subheading_style))
    personal_data = [
        ['Campo', 'Valor'],
        ['Nombre completo', client_data.get('name', 'N/A')],
        ['Email', client_data.get('email', 'N/A')],
        ['Teléfono', client_data.get('phone', 'N/A')],
        ['Ciudad', client_data.get('country_city', 'N/A')],
        ['Fecha de nacimiento', str(client_data.get('dob', 'N/A'))],
        ['Edad', f"{age} años" if age else 'N/A'],
        ['Sexo biológico', client_data.get('sex', 'N/A')],
    ]
    t = Table(personal_data, colWidths=[7*cm, 9*cm])
    t.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2E86C1')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EBF5FB')),
    ]))
    story.append(t)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.2 ESTADO DE SALUD
    story.append(Paragraph("1.2 Estado de Salud", subheading_style))
    health_data = [
        ['Aspecto', 'Detalles'],
        ['Condiciones médicas', client_data.get('conditions', 'Ninguna reportada') or 'Ninguna'],
        ['Lesiones recientes', client_data.get('injuries', 'Ninguna reportada') or 'Ninguna'],
        ['Medicación/Suplementos', client_data.get('meds', 'Ninguna reportada') or 'Ninguna'],
        ['Aprobación médica', client_data.get('clearance', 'N/A')],
    ]
    t2 = Table(health_data, colWidths=[5*cm, 11*cm])
    t2.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FADBD8')),
    ]))
    story.append(t2)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.3 MEDIDAS CORPORALES
    story.append(Paragraph("1.3 Medidas Corporales", subheading_style))
    measures_data = [
        ['Medida', 'Valor'],
        ['Peso', f"{client_data.get('weight', 0):.1f} kg"],
        ['Estatura', f"{client_data.get('height', 0):.0f} cm"],
        ['Cintura', f"{client_data.get('waist', 0):.1f} cm" if client_data.get('waist') else 'No medida'],
        ['Cadera', f"{client_data.get('hip', 0):.1f} cm" if client_data.get('hip') else 'No medida'],
        ['Cuello', f"{client_data.get('neck', 0):.1f} cm" if client_data.get('neck') else 'No medida'],
        ['Brazo', f"{client_data.get('arm', 0):.1f} cm" if client_data.get('arm') else 'No medida'],
        ['%Grasa conocido', f"{client_data.get('bf_known_value', 0):.1f}%" if client_data.get('bf_known_value') else 'No disponible'],
    ]
    t3 = Table(measures_data, colWidths=[7*cm, 9*cm])
    t3.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8F8F5')),
    ]))
    story.append(t3)
    story.append(Spacer(1, 0.2*inch))
    
    story.append(PageBreak())
    
    # 1.4 INFORMACIÓN HORMONAL
    story.append(Paragraph("1.4 Información Hormonal", subheading_style))
    if client_data.get('sex') == "Mujer":
        hormonal_data = [
            ['Aspecto', 'Información'],
            ['Ciclo menstrual', client_data.get('cycle', 'N/A')],
            ['Síntomas del ciclo (1-5)', str(client_data.get('cycle_symptoms', 'N/A'))],
            ['Anticonceptivos', client_data.get('contraceptives', 'N/A')],
            ['Embarazos/Posparto', client_data.get('pregnancy_history', 'N/A') or 'N/A'],
        ]
    else:
        hormonal_data = [
            ['Aspecto', 'Información'],
            ['Analítica hormonal', client_data.get('hormonal_panel', 'No disponible') or 'No disponible'],
            ['Síntomas baja energía (1-5)', str(client_data.get('reds_symptoms', 'N/A'))],
        ]
    
    t4 = Table(hormonal_data, colWidths=[6*cm, 10*cm])
    t4.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EBDEF0')),
    ]))
    story.append(t4)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.5 ESTILO DE VIDA
    story.append(Paragraph("1.5 Estilo de Vida y Recuperación", subheading_style))
    lifestyle_data = [
        ['Factor', 'Valor'],
        ['Horas de sueño promedio', f"{client_data.get('sleep_hours', 0):.1f} horas"],
        ['Calidad del sueño (1-5)', f"{client_data.get('sleep_quality', 'N/A')}/5"],
        ['Nivel de estrés (1-5)', f"{client_data.get('stress', 'N/A')}/5"],
        ['Pasos diarios promedio', client_data.get('steps', 'N/A')],
        ['Tipo de trabajo', client_data.get('work_type', 'N/A')],
    ]
    t5 = Table(lifestyle_data, colWidths=[8*cm, 8*cm])
    t5.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#F39C12')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FEF5E7')),
    ]))
    story.append(t5)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.6 EXPERIENCIA EN ENTRENAMIENTO
    story.append(Paragraph("1.6 Experiencia en Entrenamiento", subheading_style))
    training_data = [
        ['Factor', 'Información'],
        ['Años de experiencia', client_data.get('exp_years', 'N/A')],
        ['Continuidad reciente', client_data.get('continuity', 'N/A')],
        ['Días de entrenamiento/semana', f"{client_data.get('freq', 0)} días"],
        ['Equipamiento disponible', client_data.get('equipment', 'N/A')],
    ]
    t6 = Table(training_data, colWidths=[7*cm, 9*cm])
    t6.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#16A085')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D1F2EB')),
    ]))
    story.append(t6)
    story.append(Spacer(1, 0.2*inch))
    
    story.append(PageBreak())
    
    # 1.7 OBJETIVOS
    story.append(Paragraph("1.7 Objetivos y Expectativas", subheading_style))
    goals_data = [
        ['Aspecto', 'Detalle'],
        ['Objetivo principal', client_data.get('main_goal', 'N/A')],
        ['Plazo para alcanzarlo', client_data.get('horizon', 'N/A')],
        ['Objetivos secundarios', ', '.join(client_data.get('secondary', [])) or 'Ninguno'],
        ['Enfoque Estética/Rendimiento', f"{client_data.get('aesthetics_vs_perf', 5)}/10"],
    ]
    t7 = Table(goals_data, colWidths=[6*cm, 10*cm])
    t7.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E67E22')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FDEBD0')),
    ]))
    story.append(t7)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.8 NUTRICIÓN
    story.append(Paragraph("1.8 Hábitos Nutricionales Actuales", subheading_style))
    nutrition_data = [
        ['Aspecto', 'Información'],
        ['Comidas principales/día', client_data.get('meals', 'N/A')],
        ['Tipo de alimentación', client_data.get('diet_type', 'N/A')],
        ['Alergias/Intolerancias', client_data.get('allergies', 'Ninguna') or 'Ninguna'],
        ['Proteína diaria estimada', client_data.get('protein_est', 'N/A')],
        ['Fuentes de proteína', ', '.join(client_data.get('protein_sources', [])) or 'N/A'],
        ['Relación con la comida (1-5)', f"{client_data.get('food_relation', 'N/A')}/5"],
        ['Nivel de actividad total', client_data.get('factor_label', 'N/A')],
    ]
    t8 = Table(nutrition_data, colWidths=[6*cm, 10*cm])
    t8.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#27AE60')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#D5F4E6')),
    ]))
    story.append(t8)
    story.append(Spacer(1, 0.2*inch))
    
    # 1.9 SUPLEMENTACIÓN
    story.append(Paragraph("1.9 Suplementación Actual", subheading_style))
    supps_text = ', '.join(client_data.get('supps', [])) if client_data.get('supps') else 'Ninguna'
    story.append(Paragraph(f"<b>Suplementos actuales:</b> {supps_text}", body_style))
    if client_data.get('supp_details'):
        story.append(Paragraph(f"<b>Detalles:</b> {client_data.get('supp_details')}", body_style))
    story.append(Spacer(1, 0.2*inch))
    
    # 1.10 ADHERENCIA
    story.append(Paragraph("1.10 Adherencia y Preferencias", subheading_style))
    adherence_data = [
        ['Factor', 'Valor'],
        ['Nivel de compromiso (1-5)', f"{client_data.get('commitment', 'N/A')}/5"],
        ['Barreras principales', ', '.join(client_data.get('barriers', [])) or 'Ninguna'],
        ['Estilo de coaching preferido', client_data.get('coaching_style', 'N/A')],
        ['Frecuencia de seguimiento', client_data.get('feedback_freq', 'N/A')],
    ]
    t9 = Table(adherence_data, colWidths=[7*cm, 9*cm])
    t9.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#8E44AD')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#E8DAEF')),
    ]))
    story.append(t9)
    
    story.append(PageBreak())
    
    # ============================================
    # PARTE 2: ANÁLISIS Y RESULTADOS PROFESIONALES
    # ============================================
    story.append(Paragraph("PARTE 2: ANÁLISIS Y RESULTADOS PROFESIONALES", heading_style))
    story.append(Spacer(1, 0.2*inch))
    
    # 2.1 ANÁLISIS ANTROPOMÉTRICO
    story.append(Paragraph("2.1 Análisis Antropométrico", subheading_style))
    anthro_results = [
        ['Parámetro', 'Valor', 'Clasificación/Riesgo'],
        ['Peso', f"{client_data.get('weight', 0):.1f} kg", ''],
        ['Estatura', f"{client_data.get('height', 0):.0f} cm", ''],
        ['IMC', f"{bmi:.1f}" if bmi else 'N/A', 
         'Normal (18.5-24.9)' if bmi and 18.5 <= bmi < 25 else 
         ('Sobrepeso (25-29.9)' if bmi and 25 <= bmi < 30 else 
          ('Obesidad (≥30)' if bmi and bmi >= 30 else 'Bajo peso (<18.5)'))],
        ['WHR (Cintura/Cadera)', f"{whr:.2f}" if whr else 'N/A',
         f"{'Normal' if whr and ((client_data.get('sex')=='Hombre' and whr<0.90) or (client_data.get('sex')=='Mujer' and whr<0.85)) else 'Riesgo cardiovascular elevado'}" if whr else 'N/A'],
        ['% Grasa Corporal (Navy)', f"{bf:.1f}%" if bf else 'N/A', 
         'Método: Fórmula US Navy' if bf else ''],
    ]
    
    t_anthro = Table(anthro_results, colWidths=[5*cm, 4*cm, 7*cm])
    t_anthro.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('GRID', (0, 0), (-1, -1), 1, colors.grey),
        ('FONTSIZE', (0, 1), (-1, -1), 9),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 1), (-1, -1), 6),
        ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
    ]))
    story.append(t_anthro)
    story.append(Spacer(1, 0.3*inch))
    
    # 2.2 GASTO ENERGÉTICO
    story.append(Paragraph("2.2 Gasto Energético y Requerimientos Calóricos", subheading_style))
    
    if bmr and tdee_maintenance:
        energy_results = [
            ['Parámetro', 'Valor', 'Descripción'],
            ['BMR (Metabolismo Basal)', f"{bmr:.0f} kcal/día", 'Energía en reposo absoluto'],
            ['TDEE (Gasto Total Diario)', f"{tdee_maintenance:.0f} kcal/día", f"Incluye actividad: {client_data.get('factor_label', 'N/A')}"],
            ['', '', ''],
            ['ESCENARIOS CALÓRICOS', '', ''],
            ['Superávit Conservador (+7.5%)', f"{tdee_maintenance * 1.075:.0f} kcal/día", 'Ganancia muscular controlada'],
            ['Superávit Agresivo (+17.5%)', f"{tdee_maintenance * 1.175:.0f} kcal/día", 'Máxima ganancia muscular'],
            ['Déficit Conservador (-7.5%)', f"{tdee_maintenance * 0.925:.0f} kcal/día", 'Pérdida de grasa gradual'],
            ['Déficit Agresivo (-17.5%)', f"{tdee_maintenance * 0.825:.0f} kcal/día", 'Pérdida de grasa acelerada'],
        ]
        
        t_energy = Table(energy_results, colWidths=[6*cm, 4*cm, 6*cm])
        t_energy.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#FADBD8')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
        ]))
        story.append(t_energy)
        story.append(Spacer(1, 0.3*inch))
    
    story.append(PageBreak())
    
    # 2.3 DISTRIBUCIÓN DE MACRONUTRIENTES
    story.append(Paragraph("2.3 Distribución de Macronutrientes (Mantenimiento)", subheading_style))
    
    if protein_avg and fat_avg and carbs_avg and tdee_maintenance:
        weight = client_data.get('weight', 70)
        
        # Tabla de macros
        macro_results = [
            ['Macronutriente', 'Cantidad (g/día)', 'Ratio (g/kg)', 'Calorías', '% del Total'],
            ['Proteína', f"{protein_avg:.0f}g", f"{protein_avg/weight:.2f}", f"{protein_avg*4:.0f} kcal", 
             f"{(protein_avg*4/tdee_maintenance)*100:.1f}%"],
            ['Grasas', f"{fat_avg:.0f}g", f"{fat_avg/weight:.2f}", f"{fat_avg*9:.0f} kcal",
             f"{(fat_avg*9/tdee_maintenance)*100:.1f}%"],
            ['Carbohidratos', f"{carbs_avg:.0f}g", f"{carbs_avg/weight:.2f}", f"{carbs_avg*4:.0f} kcal",
             f"{(carbs_avg*4/tdee_maintenance)*100:.1f}%"],
            ['TOTAL', '', '', f"{tdee_maintenance:.0f} kcal", '100%'],
        ]
        
        t_macros = Table(macro_results, colWidths=[4*cm, 3*cm, 2.5*cm, 3*cm, 3.5*cm])
        t_macros.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2ECC71')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTNAME', (0, 4), (-1, 4), 'Helvetica-Bold'),
            ('BACKGROUND', (0, 4), (-1, 4), colors.HexColor('#D5F4E6')),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (0, 3), colors.HexColor('#FADBD8')),
            ('BACKGROUND', (1, 1), (-1, 3), colors.HexColor('#E8F8F5')),
        ]))
        story.append(t_macros)
        story.append(Spacer(1, 0.2*inch))
        
        # GRÁFICO DE MACROS con matplotlib (más compatible)
        try:
            matplotlib.use('Agg')  # Backend sin GUI
            
            # Crear gráfico de pastel
            fig, ax = plt.subplots(figsize=(6, 4))
            sizes = [protein_avg*4, fat_avg*9, carbs_avg*4]
            labels = ['Proteína', 'Grasas', 'Carbohidratos']
            pie_colors = ['#FF6B6B', '#4ECDC4', '#95E1D3']
            explode = (0.05, 0.05, 0.05)
            
            ax.pie(sizes, labels=labels, colors=pie_colors, autopct='%1.1f%%',
                   startangle=90, explode=explode, textprops={'fontsize': 10})
            ax.set_title('Distribución Calórica de Macronutrientes', fontsize=12, fontweight='bold')
            
            # Guardar en buffer
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            
            # Añadir al PDF
            img = RLImage(img_buffer, width=5*inch, height=3*inch)
            story.append(img)
            story.append(Spacer(1, 0.2*inch))
        except Exception as e:
            story.append(Paragraph(f"<i>Nota: Gráfico de macros no disponible</i>", body_style))

    
    # 2.4 TESTS DE FUERZA
    if strength_tests_data and len(strength_tests_data) > 0:
        story.append(Paragraph("2.4 Tests de Fuerza - Estimación de 1RM", subheading_style))
        
        strength_results = [['Ejercicio', 'Carga (kg)', 'Reps', 'RPE', '1RM Promedio (kg)']]
        for lift, vals in strength_tests_data.items():
            strength_results.append([
                lift,
                f"{vals['carga']:.1f}",
                str(vals['reps']),
                f"{vals['rpe']:.1f}",
                f"{vals['Promedio']:.1f}"
            ])
        
        t_strength = Table(strength_results, colWidths=[5*cm, 3*cm, 2*cm, 2*cm, 4*cm])
        t_strength.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#FADBD8')),
        ]))
        story.append(t_strength)
        story.append(Spacer(1, 0.2*inch))
        
        # GRÁFICO DE BARRAS 1RM
        try:
            matplotlib.use('Agg')
            
            # Datos
            lifts = list(strength_tests_data.keys())
            rms = [vals['Promedio'] for vals in strength_tests_data.values()]
            
            # Crear gráfico
            fig, ax = plt.subplots(figsize=(7, 4))
            bars = ax.bar(lifts, rms, color='#667EEA', edgecolor='black', linewidth=1.2)
            
            # Añadir valores encima de las barras
            for bar in bars:
                height = bar.get_height()
                ax.text(bar.get_x() + bar.get_width()/2., height,
                       f'{height:.1f} kg',
                       ha='center', va='bottom', fontsize=9, fontweight='bold')
            
            ax.set_xlabel('Ejercicio', fontsize=11, fontweight='bold')
            ax.set_ylabel('1RM (kg)', fontsize=11, fontweight='bold')
            ax.set_title('1RM Estimado por Ejercicio', fontsize=12, fontweight='bold')
            ax.grid(axis='y', alpha=0.3)
            plt.xticks(rotation=45, ha='right', fontsize=9)
            plt.tight_layout()
            
            # Guardar en buffer
            img_buffer = io.BytesIO()
            plt.savefig(img_buffer, format='png', dpi=150, bbox_inches='tight')
            img_buffer.seek(0)
            plt.close()
            
            # Añadir al PDF
            img = RLImage(img_buffer, width=5.5*inch, height=3.5*inch)
            story.append(img)
            story.append(Spacer(1, 0.2*inch))
        except Exception as e:
            story.append(Paragraph(f"<i>Nota: Gráfico de fuerza no disponible</i>", body_style))

        
        # Zonas de entrenamiento
        story.append(Paragraph("2.4.1 Zonas de Entrenamiento Recomendadas", subheading_style))
        zones_results = [['Ejercicio', 'Fuerza (85-95%)', 'Hipertrofia (65-85%)', 'Resistencia (50-65%)']]
        
        for lift, vals in strength_tests_data.items():
            rm = vals['Promedio']
            zones_results.append([
                lift,
                f"{rm*0.85:.0f}-{rm*0.95:.0f} kg",
                f"{rm*0.65:.0f}-{rm*0.85:.0f} kg",
                f"{rm*0.50:.0f}-{rm*0.65:.0f} kg"
            ])
        
        t_zones = Table(zones_results, colWidths=[5*cm, 3.5*cm, 3.5*cm, 4*cm])
        t_zones.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#9B59B6')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('GRID', (0, 0), (-1, -1), 1, colors.grey),
            ('FONTSIZE', (0, 0), (-1, -1), 8),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            ('BACKGROUND', (0, 1), (-1, -1), colors.HexColor('#EBDEF0')),
        ]))
        story.append(t_zones)
    
    story.append(PageBreak())
    
    # 2.5 ALERTAS Y RECOMENDACIONES
    story.append(Paragraph("2.5 Alertas y Recomendaciones Clínicas", subheading_style))
    
    if alerts and len(alerts) > 0:
        for alert in alerts:
            if "🔴" in alert:
                p = Paragraph(f"<b>⚠ ALERTA CRÍTICA:</b> {alert.replace('🔴', '')}", 
                             ParagraphStyle('alert', parent=body_style, textColor=colors.HexColor('#C0392B')))
                story.append(p)
            elif "🟡" in alert:
                p = Paragraph(f"<b>⚠ ATENCIÓN:</b> {alert.replace('🟡', '')}", 
                             ParagraphStyle('warning', parent=body_style, textColor=colors.HexColor('#D68910')))
                story.append(p)
            else:
                p = Paragraph(f"• {alert.replace('💡', '')}", body_style)
                story.append(p)
    else:
        story.append(Paragraph("✓ No se detectaron alertas críticas. El cliente se encuentra en condiciones óptimas para iniciar el programa de entrenamiento.", 
                              ParagraphStyle('ok', parent=body_style, textColor=colors.HexColor('#229954'))))
    
    story.append(Spacer(1, 0.5*inch))
    
    # PIE DE PÁGINA FINAL
    story.append(Spacer(1, 1*inch))
    story.append(Paragraph("_" * 80, ParagraphStyle('line', alignment=TA_CENTER, fontSize=8)))
    story.append(Paragraph(f"Informe completo generado el {datetime.now().strftime('%d/%m/%Y a las %H:%M')}", 
                          ParagraphStyle('footer', alignment=TA_CENTER, fontSize=8, textColor=colors.grey)))
    story.append(Paragraph("Sistema de Evaluación Integral v5.0 | Basado en evidencia científica 2024-2025", 
                          ParagraphStyle('footer2', alignment=TA_CENTER, fontSize=7, textColor=colors.grey)))
    story.append(Paragraph(f"Cliente: {client_data.get('name', 'N/A')} | Profesional responsable: [Nombre del profesional]", 
                          ParagraphStyle('footer3', alignment=TA_CENTER, fontSize=7, textColor=colors.grey)))
    
    # Construir PDF
    doc.build(story)
    buffer.seek(0)
    return buffer



# ================================================================================
# GENERACIÓN DE EXCEL COMPLETO Y PROFESIONAL
# =================================================================================


def generate_complete_excel(client_data, age, bmi, whr, bf, bmr, tdee_maintenance, 
                            protein_avg, fat_avg, carbs_avg, alerts, strength_tests_data=None):
    """
    Genera un archivo Excel completo con múltiples hojas y formato profesional
    """
    wb = Workbook()
    
    # Eliminar hoja por defecto
    if 'Sheet' in wb.sheetnames:
        wb.remove(wb['Sheet'])
    
    # Estilos reutilizables
    header_fill = PatternFill(start_color="2E86C1", end_color="2E86C1", fill_type="solid")
    header_font = Font(name='Calibri', bold=True, color="FFFFFF", size=12)
    subheader_fill = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
    subheader_font = Font(name='Calibri', bold=True, size=11)
    data_font = Font(name='Calibri', size=10)
    center_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    left_alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ============================================
    # HOJA 1: RESUMEN EJECUTIVO
    # ============================================
    ws_resumen = wb.create_sheet("Resumen Ejecutivo", 0)
    ws_resumen.column_dimensions['A'].width = 30
    ws_resumen.column_dimensions['B'].width = 40
    
    # Título
    ws_resumen['A1'] = "INFORME DE EVALUACIÓN FÍSICA"
    ws_resumen['A1'].font = Font(name='Calibri', bold=True, size=16, color="1A5490")
    ws_resumen.merge_cells('A1:B1')
    ws_resumen['A1'].alignment = center_alignment
    
    # Fecha y cliente
    ws_resumen['A2'] = "Cliente:"
    ws_resumen['B2'] = client_data.get('name', 'N/A')
    ws_resumen['A3'] = "Fecha del informe:"
    ws_resumen['B3'] = datetime.now().strftime('%d/%m/%Y')
    
    for row in range(2, 4):
        ws_resumen[f'A{row}'].font = Font(bold=True)
    
    ws_resumen['A5'] = "DATOS CLAVE"
    ws_resumen['A5'].font = subheader_font
    ws_resumen['A5'].fill = subheader_fill
    ws_resumen.merge_cells('A5:B5')
    
    # Datos clave
    key_data = [
        ("Edad", f"{age} años" if age else "N/A"),
        ("Sexo", client_data.get('sex', 'N/A')),
        ("Peso", f"{client_data.get('weight', 0):.1f} kg"),
        ("Estatura", f"{client_data.get('height', 0):.0f} cm"),
        ("IMC", f"{bmi:.1f}" if bmi else "N/A"),
        ("% Grasa", f"{bf:.1f}%" if bf else "N/A"),
        ("BMR", f"{bmr:.0f} kcal/día" if bmr else "N/A"),
        ("TDEE", f"{tdee_maintenance:.0f} kcal/día" if tdee_maintenance else "N/A"),
        ("Objetivo", client_data.get('main_goal', 'N/A')),
        ("Experiencia", client_data.get('exp_years', 'N/A')),
    ]
    
    row_num = 6
    for label, value in key_data:
        ws_resumen[f'A{row_num}'] = label
        ws_resumen[f'B{row_num}'] = value
        ws_resumen[f'A{row_num}'].font = Font(bold=True)
        ws_resumen[f'A{row_num}'].border = thin_border
        ws_resumen[f'B{row_num}'].border = thin_border
        row_num += 1
    
    # ============================================
    # HOJA 2: DATOS PERSONALES COMPLETOS
    # ============================================
    ws_personal = wb.create_sheet("Datos Personales")
    ws_personal.column_dimensions['A'].width = 30
    ws_personal.column_dimensions['B'].width = 50
    
    ws_personal['A1'] = "INFORMACIÓN PERSONAL COMPLETA"
    ws_personal['A1'].font = subheader_font
    ws_personal['A1'].fill = subheader_fill
    ws_personal.merge_cells('A1:B1')
    ws_personal['A1'].alignment = center_alignment
    
    personal_info = [
        ("", ""),
        ("DATOS BÁSICOS", ""),
        ("Nombre completo", client_data.get('name', 'N/A')),
        ("Email", client_data.get('email', 'N/A')),
        ("Teléfono", client_data.get('phone', 'N/A')),
        ("Ciudad", client_data.get('country_city', 'N/A')),
        ("Fecha de nacimiento", str(client_data.get('dob', 'N/A'))),
        ("Edad", f"{age} años" if age else 'N/A'),
        ("Sexo biológico", client_data.get('sex', 'N/A')),
        ("", ""),
        ("SALUD", ""),
        ("Condiciones médicas", client_data.get('conditions', 'Ninguna') or 'Ninguna'),
        ("Lesiones recientes", client_data.get('injuries', 'Ninguna') or 'Ninguna'),
        ("Medicación", client_data.get('meds', 'Ninguna') or 'Ninguna'),
        ("Aprobación médica", client_data.get('clearance', 'N/A')),
        ("", ""),
        ("ESTILO DE VIDA", ""),
        ("Horas de sueño", f"{client_data.get('sleep_hours', 0):.1f} horas"),
        ("Calidad del sueño (1-5)", f"{client_data.get('sleep_quality', 'N/A')}/5"),
        ("Nivel de estrés (1-5)", f"{client_data.get('stress', 'N/A')}/5"),
        ("Pasos diarios", client_data.get('steps', 'N/A')),
        ("Tipo de trabajo", client_data.get('work_type', 'N/A')),
    ]
    
    row_num = 2
    for label, value in personal_info:
        if label and not value:  # Es un subtítulo
            ws_personal[f'A{row_num}'] = label
            ws_personal[f'A{row_num}'].font = Font(bold=True, size=11)
            ws_personal[f'A{row_num}'].fill = PatternFill(start_color="D6EAF8", end_color="D6EAF8", fill_type="solid")
            ws_personal.merge_cells(f'A{row_num}:B{row_num}')
        else:
            ws_personal[f'A{row_num}'] = label
            ws_personal[f'B{row_num}'] = value
            ws_personal[f'A{row_num}'].font = Font(bold=True)
            ws_personal[f'A{row_num}'].border = thin_border
            ws_personal[f'B{row_num}'].border = thin_border
            ws_personal[f'B{row_num}'].alignment = left_alignment
        row_num += 1
    
    # ============================================
    # HOJA 3: MEDIDAS CORPORALES
    # ============================================
    ws_medidas = wb.create_sheet("Medidas Corporales")
    ws_medidas.column_dimensions['A'].width = 30
    ws_medidas.column_dimensions['B'].width = 20
    ws_medidas.column_dimensions['C'].width = 30
    
    ws_medidas['A1'] = "Parámetro"
    ws_medidas['B1'] = "Valor"
    ws_medidas['C1'] = "Clasificación"
    
    for col in ['A1', 'B1', 'C1']:
        ws_medidas[col].font = header_font
        ws_medidas[col].fill = header_fill
        ws_medidas[col].alignment = center_alignment
        ws_medidas[col].border = thin_border
    
    medidas_data = [
        ("Peso", f"{client_data.get('weight', 0):.1f} kg", ""),
        ("Estatura", f"{client_data.get('height', 0):.0f} cm", ""),
        ("IMC", f"{bmi:.1f}" if bmi else "N/A", 
         "Normal" if bmi and 18.5 <= bmi < 25 else ("Sobrepeso" if bmi and bmi >= 25 else "Bajo peso")),
        ("Cintura", f"{client_data.get('waist', 0):.1f} cm" if client_data.get('waist') else "No medida", ""),
        ("Cadera", f"{client_data.get('hip', 0):.1f} cm" if client_data.get('hip') else "No medida", ""),
        ("WHR", f"{whr:.2f}" if whr else "N/A", 
         "Normal" if whr and ((client_data.get('sex')=='Hombre' and whr<0.90) or (client_data.get('sex')=='Mujer' and whr<0.85)) else "Riesgo elevado" if whr else "N/A"),
        ("Cuello", f"{client_data.get('neck', 0):.1f} cm" if client_data.get('neck') else "No medida", ""),
        ("Brazo", f"{client_data.get('arm', 0):.1f} cm" if client_data.get('arm') else "No medida", ""),
        ("% Grasa (Navy)", f"{bf:.1f}%" if bf else "N/A", "Método: Fórmula US Navy" if bf else ""),
    ]
    
    row_num = 2
    for param, valor, clasif in medidas_data:
        ws_medidas[f'A{row_num}'] = param
        ws_medidas[f'B{row_num}'] = valor
        ws_medidas[f'C{row_num}'] = clasif
        
        for col in ['A', 'B', 'C']:
            ws_medidas[f'{col}{row_num}'].border = thin_border
            ws_medidas[f'{col}{row_num}'].alignment = center_alignment if col == 'B' else left_alignment
        
        row_num += 1
    
    # ============================================
    # HOJA 4: GASTO ENERGÉTICO Y MACROS
    # ============================================
    ws_energia = wb.create_sheet("Energía y Macros")
    ws_energia.column_dimensions['A'].width = 35
    ws_energia.column_dimensions['B'].width = 20
    ws_energia.column_dimensions['C'].width = 30
    
    ws_energia['A1'] = "GASTO ENERGÉTICO"
    ws_energia['A1'].font = subheader_font
    ws_energia['A1'].fill = subheader_fill
    ws_energia.merge_cells('A1:C1')
    ws_energia['A1'].alignment = center_alignment
    
    if bmr and tdee_maintenance:
        energia_data = [
            ("Parámetro", "Valor", "Descripción"),
            ("BMR (Metabolismo Basal)", f"{bmr:.0f} kcal/día", "Energía en reposo"),
            ("TDEE (Gasto Total)", f"{tdee_maintenance:.0f} kcal/día", f"Actividad: {client_data.get('factor_label', 'N/A')}"),
            ("", "", ""),
            ("ESCENARIOS CALÓRICOS", "", ""),
            ("Superávit Conservador (+7.5%)", f"{tdee_maintenance * 1.075:.0f} kcal/día", "Ganancia muscular controlada"),
            ("Superávit Agresivo (+17.5%)", f"{tdee_maintenance * 1.175:.0f} kcal/día", "Máxima ganancia"),
            ("Déficit Conservador (-7.5%)", f"{tdee_maintenance * 0.925:.0f} kcal/día", "Pérdida gradual"),
            ("Déficit Agresivo (-17.5%)", f"{tdee_maintenance * 0.825:.0f} kcal/día", "Pérdida acelerada"),
        ]
        
        row_num = 2
        for i, (param, valor, desc) in enumerate(energia_data):
            ws_energia[f'A{row_num}'] = param
            ws_energia[f'B{row_num}'] = valor
            ws_energia[f'C{row_num}'] = desc
            
            if i == 0:  # Header
                for col in ['A', 'B', 'C']:
                    ws_energia[f'{col}{row_num}'].font = header_font
                    ws_energia[f'{col}{row_num}'].fill = header_fill
                    ws_energia[f'{col}{row_num}'].alignment = center_alignment
            elif i == 4:  # Subtítulo
                ws_energia.merge_cells(f'A{row_num}:C{row_num}')
                ws_energia[f'A{row_num}'].font = Font(bold=True)
                ws_energia[f'A{row_num}'].fill = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")
                ws_energia[f'A{row_num}'].alignment = center_alignment
            
            for col in ['A', 'B', 'C']:
                ws_energia[f'{col}{row_num}'].border = thin_border
            
            row_num += 1
    
    # MACRONUTRIENTES
    ws_energia[f'A{row_num+1}'] = "DISTRIBUCIÓN DE MACRONUTRIENTES (MANTENIMIENTO)"
    ws_energia[f'A{row_num+1}'].font = subheader_font
    ws_energia[f'A{row_num+1}'].fill = subheader_fill
    ws_energia.merge_cells(f'A{row_num+1}:C{row_num+1}')
    ws_energia[f'A{row_num+1}'].alignment = center_alignment
    
    row_num += 2
    
    if protein_avg and fat_avg and carbs_avg and tdee_maintenance:
        weight = client_data.get('weight', 70)
        
        macros_data = [
            ("Macronutriente", "Cantidad", "Ratio", "Calorías", "% Total"),
            ("Proteína", f"{protein_avg:.0f}g", f"{protein_avg/weight:.2f} g/kg", f"{protein_avg*4:.0f} kcal", f"{(protein_avg*4/tdee_maintenance)*100:.1f}%"),
            ("Grasas", f"{fat_avg:.0f}g", f"{fat_avg/weight:.2f} g/kg", f"{fat_avg*9:.0f} kcal", f"{(fat_avg*9/tdee_maintenance)*100:.1f}%"),
            ("Carbohidratos", f"{carbs_avg:.0f}g", f"{carbs_avg/weight:.2f} g/kg", f"{carbs_avg*4:.0f} kcal", f"{(carbs_avg*4/tdee_maintenance)*100:.1f}%"),
            ("TOTAL", "", "", f"{tdee_maintenance:.0f} kcal", "100%"),
        ]
        
        ws_energia.column_dimensions['D'].width = 15
        ws_energia.column_dimensions['E'].width = 15
        
        for i, row_data in enumerate(macros_data):
            for j, value in enumerate(row_data):
                col_letter = chr(65 + j)  # A, B, C, D, E
                ws_energia[f'{col_letter}{row_num}'] = value
                
                if i == 0:  # Header
                    ws_energia[f'{col_letter}{row_num}'].font = header_font
                    ws_energia[f'{col_letter}{row_num}'].fill = PatternFill(start_color="2ECC71", end_color="2ECC71", fill_type="solid")
                elif i == 4:  # Total
                    ws_energia[f'{col_letter}{row_num}'].font = Font(bold=True)
                    ws_energia[f'{col_letter}{row_num}'].fill = PatternFill(start_color="D5F4E6", end_color="D5F4E6", fill_type="solid")
                
                ws_energia[f'{col_letter}{row_num}'].border = thin_border
                ws_energia[f'{col_letter}{row_num}'].alignment = center_alignment
            
            row_num += 1
    
    # ============================================
    # HOJA 5: TESTS DE FUERZA
    # ============================================
    if strength_tests_data and len(strength_tests_data) > 0:
        ws_fuerza = wb.create_sheet("Tests de Fuerza")
        ws_fuerza.column_dimensions['A'].width = 25
        ws_fuerza.column_dimensions['B'].width = 15
        ws_fuerza.column_dimensions['C'].width = 10
        ws_fuerza.column_dimensions['D'].width = 10
        ws_fuerza.column_dimensions['E'].width = 18
        
        ws_fuerza['A1'] = "TESTS DE FUERZA - ESTIMACIÓN 1RM"
        ws_fuerza['A1'].font = subheader_font
        ws_fuerza['A1'].fill = subheader_fill
        ws_fuerza.merge_cells('A1:E1')
        ws_fuerza['A1'].alignment = center_alignment
        
        headers = ["Ejercicio", "Carga (kg)", "Reps", "RPE", "1RM Promedio (kg)"]
        for i, header in enumerate(headers):
            col_letter = chr(65 + i)
            ws_fuerza[f'{col_letter}2'] = header
            ws_fuerza[f'{col_letter}2'].font = header_font
            ws_fuerza[f'{col_letter}2'].fill = header_fill
            ws_fuerza[f'{col_letter}2'].alignment = center_alignment
            ws_fuerza[f'{col_letter}2'].border = thin_border
        
        row_num = 3
        for lift, vals in strength_tests_data.items():
            ws_fuerza[f'A{row_num}'] = lift
            ws_fuerza[f'B{row_num}'] = f"{vals['carga']:.1f}"
            ws_fuerza[f'C{row_num}'] = vals['reps']
            ws_fuerza[f'D{row_num}'] = f"{vals['rpe']:.1f}"
            ws_fuerza[f'E{row_num}'] = f"{vals['Promedio']:.1f}"
            
            for col in ['A', 'B', 'C', 'D', 'E']:
                ws_fuerza[f'{col}{row_num}'].border = thin_border
                ws_fuerza[f'{col}{row_num}'].alignment = center_alignment
            
            row_num += 1
        
        # Zonas de entrenamiento
        row_num += 2
        ws_fuerza[f'A{row_num}'] = "ZONAS DE ENTRENAMIENTO"
        ws_fuerza[f'A{row_num}'].font = subheader_font
        ws_fuerza[f'A{row_num}'].fill = subheader_fill
        ws_fuerza.merge_cells(f'A{row_num}:D{row_num}')
        ws_fuerza[f'A{row_num}'].alignment = center_alignment
        
        row_num += 1
        zone_headers = ["Ejercicio", "Fuerza (85-95%)", "Hipertrofia (65-85%)", "Resistencia (50-65%)"]
        for i, header in enumerate(zone_headers):
            col_letter = chr(65 + i)
            ws_fuerza[f'{col_letter}{row_num}'] = header
            ws_fuerza[f'{col_letter}{row_num}'].font = header_font
            ws_fuerza[f'{col_letter}{row_num}'].fill = PatternFill(start_color="9B59B6", end_color="9B59B6", fill_type="solid")
            ws_fuerza[f'{col_letter}{row_num}'].alignment = center_alignment
            ws_fuerza[f'{col_letter}{row_num}'].border = thin_border
        
        row_num += 1
        for lift, vals in strength_tests_data.items():
            rm = vals['Promedio']
            ws_fuerza[f'A{row_num}'] = lift
            ws_fuerza[f'B{row_num}'] = f"{rm*0.85:.0f}-{rm*0.95:.0f} kg"
            ws_fuerza[f'C{row_num}'] = f"{rm*0.65:.0f}-{rm*0.85:.0f} kg"
            ws_fuerza[f'D{row_num}'] = f"{rm*0.50:.0f}-{rm*0.65:.0f} kg"
            
            for col in ['A', 'B', 'C', 'D']:
                ws_fuerza[f'{col}{row_num}'].border = thin_border
                ws_fuerza[f'{col}{row_num}'].alignment = center_alignment
            
            row_num += 1
    
    # ============================================
    # HOJA 6: OBJETIVOS Y NUTRICIÓN
    # ============================================
    ws_objetivos = wb.create_sheet("Objetivos y Nutrición")
    ws_objetivos.column_dimensions['A'].width = 30
    ws_objetivos.column_dimensions['B'].width = 50
    
    ws_objetivos['A1'] = "OBJETIVOS"
    ws_objetivos['A1'].font = subheader_font
    ws_objetivos['A1'].fill = subheader_fill
    ws_objetivos.merge_cells('A1:B1')
    ws_objetivos['A1'].alignment = center_alignment
    
    objetivos_data = [
        ("Objetivo principal", client_data.get('main_goal', 'N/A')),
        ("Plazo", client_data.get('horizon', 'N/A')),
        ("Objetivos secundarios", ', '.join(client_data.get('secondary', [])) or 'Ninguno'),
        ("Enfoque (Estética/Rendimiento)", f"{client_data.get('aesthetics_vs_perf', 5)}/10"),
        ("", ""),
        ("NUTRICIÓN", ""),
        ("Comidas/día", client_data.get('meals', 'N/A')),
        ("Tipo de alimentación", client_data.get('diet_type', 'N/A')),
        ("Alergias", client_data.get('allergies', 'Ninguna') or 'Ninguna'),
        ("Proteína estimada actual", client_data.get('protein_est', 'N/A')),
        ("Fuentes de proteína", ', '.join(client_data.get('protein_sources', [])) or 'N/A'),
        ("Suplementos actuales", ', '.join(client_data.get('supps', [])) or 'Ninguna'),
    ]
    
    row_num = 2
    for label, value in objetivos_data:
        if label and not value:  # Subtítulo
            ws_objetivos[f'A{row_num}'] = label
            ws_objetivos[f'A{row_num}'].font = Font(bold=True, size=11)
            ws_objetivos[f'A{row_num}'].fill = PatternFill(start_color="FDEBD0", end_color="FDEBD0", fill_type="solid")
            ws_objetivos.merge_cells(f'A{row_num}:B{row_num}')
        elif label and value:
            ws_objetivos[f'A{row_num}'] = label
            ws_objetivos[f'B{row_num}'] = value
            ws_objetivos[f'A{row_num}'].font = Font(bold=True)
            ws_objetivos[f'A{row_num}'].border = thin_border
            ws_objetivos[f'B{row_num}'].border = thin_border
            ws_objetivos[f'B{row_num}'].alignment = left_alignment
        row_num += 1
    
    # ============================================
    # HOJA 7: ALERTAS
    # ============================================
    if alerts and len(alerts) > 0:
        ws_alertas = wb.create_sheet("Alertas Clínicas")
        ws_alertas.column_dimensions['A'].width = 15
        ws_alertas.column_dimensions['B'].width = 70
        
        ws_alertas['A1'] = "Nivel"
        ws_alertas['B1'] = "Alerta / Recomendación"
        
        for col in ['A1', 'B1']:
            ws_alertas[col].font = header_font
            ws_alertas[col].fill = header_fill
            ws_alertas[col].alignment = center_alignment
            ws_alertas[col].border = thin_border
        
        row_num = 2
        for alert in alerts:
            if "🔴" in alert:
                ws_alertas[f'A{row_num}'] = "CRÍTICA"
                ws_alertas[f'A{row_num}'].fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
                ws_alertas[f'A{row_num}'].font = Font(bold=True, color="FFFFFF")
                ws_alertas[f'B{row_num}'] = alert.replace('🔴', '')
            elif "🟡" in alert:
                ws_alertas[f'A{row_num}'] = "ATENCIÓN"
                ws_alertas[f'A{row_num}'].fill = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
                ws_alertas[f'A{row_num}'].font = Font(bold=True)
                ws_alertas[f'B{row_num}'] = alert.replace('🟡', '')
            else:
                ws_alertas[f'A{row_num}'] = "INFO"
                ws_alertas[f'A{row_num}'].fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
                ws_alertas[f'A{row_num}'].font = Font(bold=True, color="FFFFFF")
                ws_alertas[f'B{row_num}'] = alert.replace('💡', '')
            
            for col in ['A', 'B']:
                ws_alertas[f'{col}{row_num}'].border = thin_border
                ws_alertas[f'{col}{row_num}'].alignment = left_alignment
            
            row_num += 1
    
    # Guardar en buffer
    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return buffer


# ============================================
# PERSISTENCIA
# ============================================

DRAFT_FILE = "form_draft.pkl"

def save_draft():
    data_to_save = {k: v for k, v in st.session_state.items() 
                   if not k.startswith('_') and k not in ['FormSubmitter']}
    try:
        with open(DRAFT_FILE, 'wb') as f:
            pickle.dump(data_to_save, f)
        return True
    except Exception as e:
        st.error(f"Error guardando: {e}")
        return False

def load_draft():
    try:
        if os.path.exists(DRAFT_FILE):
            with open(DRAFT_FILE, 'rb') as f:
                return pickle.load(f)
    except Exception as e:
        st.error(f"Error cargando: {e}")
    return None

def delete_draft():
    try:
        if os.path.exists(DRAFT_FILE):
            os.remove(DRAFT_FILE)
            return True
    except Exception as e:
        st.error(f"Error eliminando: {e}")
    return False

def save_strength_tests(nombre, email, lifts_dict):
    """Guarda tests de fuerza en base de datos"""
    conn = sqlite3.connect('clientes_evaluaciones.db')
    c = conn.cursor()
    fecha = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
    for ejercicio, vals in lifts_dict.items():
        c.execute('''INSERT INTO tests_fuerza VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
                  (fecha, nombre, email, ejercicio, vals['carga'], vals['reps'], vals['rpe'],
                   vals['Epley'], vals['Brzycki'], vals['Promedio']))
    conn.commit()
    conn.close()

def save_to_database(data_dict):
    """Guarda evaluación completa"""
    conn = sqlite3.connect('clientes_evaluaciones.db')
    c = conn.cursor()
    c.execute('''INSERT INTO evaluaciones VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)''',
              (data_dict['fecha'], data_dict['nombre'], data_dict['email'], data_dict['sexo'],
               data_dict['edad'], data_dict['peso'], data_dict['estatura'], data_dict['imc'],
               data_dict['whr'], data_dict['bf_navy'], data_dict['bmr'], data_dict['tdee'],
               data_dict['calorias_objetivo'], data_dict['proteina_g'], data_dict['grasa_g'],
               data_dict['carbohidratos_g'], data_dict['objetivo_principal'], data_dict['experiencia']))
    conn.commit()
    conn.close()

# ============================================
# BASE DE DATOS
# ============================================

def init_database():
    conn = sqlite3.connect('clientes_evaluaciones.db')
    c = conn.cursor()
    c.execute('''CREATE TABLE IF NOT EXISTS evaluaciones
                 (fecha TEXT, nombre TEXT, email TEXT, sexo TEXT, edad INTEGER,
                  peso REAL, estatura REAL, imc REAL, whr REAL, bf_navy REAL,
                  bmr REAL, tdee REAL, calorias_objetivo REAL,
                  proteina_g REAL, grasa_g REAL, carbohidratos_g REAL,
                  objetivo_principal TEXT, experiencia TEXT)''')
    c.execute('''CREATE TABLE IF NOT EXISTS tests_fuerza
                 (fecha TEXT, nombre TEXT, email TEXT, ejercicio TEXT,
                  carga REAL, repeticiones INTEGER, rpe REAL,
                  rm_epley REAL, rm_brzycki REAL, rm_promedio REAL)''')
    conn.commit()
    conn.close()

init_database()


# ============================================
# INICIALIZACIÓN (CON DETECCIÓN DE RESETEO)
# ============================================

# Detectar si acabamos de resetear
just_reset = st.session_state.get('_resetting', False)

if just_reset:
    # Si acabamos de resetear, NO cargar el draft
    st.session_state['draft_loaded'] = False
    st.session_state['form_submitted'] = False
    st.session_state['client_data'] = {}
    st.session_state['strength_tests'] = {}
    st.session_state['initialized'] = True
    # Limpiar el flag
    del st.session_state['_resetting']
else:
    # Flujo normal: cargar draft si existe
    if 'initialized' not in st.session_state:
        draft_exists = os.path.exists(DRAFT_FILE)
        
        if draft_exists:
            draft = load_draft()
            if draft:
                st.session_state.update(draft)
                st.session_state['draft_loaded'] = True
            else:
                st.session_state['draft_loaded'] = False
        else:
            st.session_state['draft_loaded'] = False
        
        # Inicializar variables críticas
        if 'form_submitted' not in st.session_state:
            st.session_state['form_submitted'] = False
        if 'client_data' not in st.session_state:
            st.session_state['client_data'] = {}
        if 'strength_tests' not in st.session_state:
            st.session_state['strength_tests'] = {}
        
        st.session_state['initialized'] = True



# ============================================
# NAVEGACIÓN
# ============================================

tab1, tab2 = st.tabs(["📋 Formulario del Cliente", "📊 Panel Profesional"])

# ============================================
# TAB 1: TU CÓDIGO SIN MODIFICAR
# ============================================

with tab1:
    st.title("Formulario de Evaluación Inicial 💪")
    
    if st.session_state.get('draft_loaded'):
        st.success("✅ Formulario recuperado automáticamente.")
    
    st.markdown("**Datos guardados automáticamente al pulsar 'Guardar Borrador'**")
    
    # Botones
    col1, col2 = st.columns([2, 2])
    with col1:
        if st.button("💾 Guardar Borrador", use_container_width=True):
            if save_draft():
                st.success("✅ Guardado")
    with col2:
        if st.button("🔄 Resetear", use_container_width=True):
            delete_draft()
            st.session_state['_resetting'] = True
            keys_to_delete = [key for key in list(st.session_state.keys()) if key != '_resetting']
            for key in keys_to_delete:
                del st.session_state[key]
            st.rerun()
    
    st.divider()
    
    # NUEVO: Si acabamos de resetear, no mostrar el formulario en este frame
    if st.session_state.get('_resetting'):
        del st.session_state['_resetting']
        st.rerun()
    
    # Resto del formulario (exactamente igual que antes)...

    
    # -------- 1. INFORMACIÓN PERSONAL --------
    st.header("1️⃣ Información Personal")
    col1, col2 = st.columns(2)
    with col1:
        st.text_input("Nombre completo *", key='name')
        st.date_input("Fecha nacimiento *", key='dob', value=st.session_state.get('dob', date(2000,1,1)), 
                     min_value=date(1940,1,1), max_value=date.today())
        sex_opts = ["Hombre", "Mujer"]
        st.selectbox("Sexo *", sex_opts, key='sex', index=safe_select_index(sex_opts, st.session_state.get('sex')))
    with col2:
        st.text_input("Email *", key='email')
        st.text_input("Teléfono", key='phone')
        st.text_input("Ciudad", key='country_city')
    
    st.divider()
    
    # -------- 2. CONSENTIMIENTO --------
    st.header("2️⃣ Consentimiento")
    st.checkbox("✓ Acepto términos *", key='consent')
    
    st.divider()
    
    # -------- 3. SALUD --------
    st.header("3️⃣ Salud")
    col1, col2 = st.columns(2)
    with col1:
        st.text_area("Condiciones médicas", key='conditions')
        st.text_input("Lesiones recientes", key='injuries')
    with col2:
        st.text_area("Medicación", key='meds')
        clearance_opts = ["No necesaria", "Sí aprobado", "Pendiente"]
        st.selectbox("Aprobación médica", clearance_opts, key='clearance',
                    index=safe_select_index(clearance_opts, st.session_state.get('clearance')))
    
    st.divider()
    
    # -------- 4. MEDIDAS --------
    st.header("4️⃣ Medidas Corporales")
    col1, col2 = st.columns(2)
    with col1:
        st.number_input("Peso (kg) *", key='weight', value=float(st.session_state.get('weight', 70.0)), step=0.1, min_value=30.0)
    with col2:
        st.number_input("Estatura (cm) *", key='height', value=float(st.session_state.get('height', 170.0)), step=0.1, min_value=130.0)
    
    with st.expander("📐 Medidas adicionales", expanded=st.session_state.get('measures_expanded', False)):
        col3, col4 = st.columns(2)
        with col3:
            st.number_input("Cintura (cm)", key='waist', value=float(st.session_state.get('waist', 80.0)), step=0.1)
            st.number_input("Cadera (cm)", key='hip', value=float(st.session_state.get('hip', 95.0)), step=0.1)
        with col4:
            st.number_input("Cuello (cm)", key='neck', value=float(st.session_state.get('neck', 35.0)), step=0.1)
            st.number_input("Brazo (cm)", key='arm', value=float(st.session_state.get('arm', 30.0)), step=0.1)
        st.session_state['measures_expanded'] = True
    
    with st.expander("📊 Conozco %grasa", expanded=st.session_state.get('bf_expanded', False)):
        st.number_input("%Grasa", key='bf_known_value', value=float(st.session_state.get('bf_known_value', 15.0)), step=0.1)
        st.session_state['bf_expanded'] = True
    
    st.divider()
    
    # -------- 5. HORMONAL --------
    st.header("5️⃣ Información Hormonal")
    
    if st.session_state.get('sex') == "Mujer":
        col1, col2 = st.columns(2)
        with col1:
            cycle_opts = ["Regular", "Irregular", "Amenorrea", "Menopausia", "No aplica"]
            st.selectbox("Ciclo menstrual", cycle_opts, key='cycle', 
                        index=safe_select_index(cycle_opts, st.session_state.get('cycle')))
            st.slider("Síntomas ciclo", 1, 5, st.session_state.get('cycle_symptoms', 3), key='cycle_symptoms')
        with col2:
            contra_opts = ["No", "Píldora combinada", "Píldora progestágeno", "DIU hormonal", "DIU no hormonal", "Implante", "Otros"]
            st.selectbox("Anticonceptivos", contra_opts, key='contraceptives', 
                        index=safe_select_index(contra_opts, st.session_state.get('contraceptives')))
            st.text_input("Embarazos", key='pregnancy_history')
    else:
        st.text_area("Analítica hormonal", key='hormonal_panel')
        st.slider("Síntomas baja energía", 1, 5, st.session_state.get('reds_symptoms', 1), key='reds_symptoms')
    
    st.divider()
    
    # -------- 6. ESTILO DE VIDA --------
    st.header("6️⃣ Estilo de Vida")
    col1, col2, col3 = st.columns(3)
    with col1:
        st.slider("Horas sueño", 4.0, 10.0, float(st.session_state.get('sleep_hours', 7.0)), 0.5, key='sleep_hours')
        st.slider("Calidad sueño", 1, 5, st.session_state.get('sleep_quality', 3), key='sleep_quality')
    with col2:
        st.slider("Estrés", 1, 5, st.session_state.get('stress', 3), key='stress')
        steps_opts = ["<5000", "5000-8000", "8000-12000", ">12000"]
        st.selectbox("Pasos/día", steps_opts, key='steps',
                    index=safe_select_index(steps_opts, st.session_state.get('steps')))
    with col3:
        work_opts = ["Sedentario", "Moderadamente activo", "Muy activo"]
        st.selectbox("Tipo trabajo", work_opts, key='work_type',
                    index=safe_select_index(work_opts, st.session_state.get('work_type')))
    
    st.divider()
    
    # -------- 7. EXPERIENCIA --------
    st.header("7️⃣ Experiencia")
    col1, col2, col3 = st.columns(3)
    with col1:
        exp_opts = ["0-1", "1-3", "3-5", "5+"]
        st.selectbox("Años", exp_opts, key='exp_years',
                    index=safe_select_index(exp_opts, st.session_state.get('exp_years')))
    with col2:
        cont_opts = ["0-3 meses", "3-6 meses", "6-12 meses", "12+ meses"]
        st.selectbox("Continuidad", cont_opts, key='continuity',
                    index=safe_select_index(cont_opts, st.session_state.get('continuity')))
    with col3:
        st.slider("Días/sem", 0, 7, st.session_state.get('freq', 3), key='freq')
    
    equip_opts = ["Casa básico", "Casa completo", "Gimnasio comercial", "Gimnasio especializado"]
    st.selectbox("Equipamiento", equip_opts, key='equipment',
                index=safe_select_index(equip_opts, st.session_state.get('equipment')))
    
    st.divider()
    
    # -------- 8. OBJETIVOS --------
    st.header("8️⃣ Objetivos")
    col1, col2 = st.columns(2)
    with col1:
        goal_opts = ["Ganar músculo", "Perder grasa", "Ganar fuerza", "Rendimiento", "Salud"]
        st.selectbox("Objetivo principal *", goal_opts, key='main_goal',
                    index=safe_select_index(goal_opts, st.session_state.get('main_goal')))
        horiz_opts = ["4 semanas", "3 meses", "6 meses", "1 año"]
        st.selectbox("Plazo", horiz_opts, key='horizon',
                    index=safe_select_index(horiz_opts, st.session_state.get('horizon')))
    with col2:
        st.multiselect("Secundarios", ["Músculo", "Grasa", "Fuerza", "Movilidad", "Postura", "Estética"], 
                      key='secondary', default=st.session_state.get('secondary', []))
        st.slider("Estética/Rendimiento", 1, 10, st.session_state.get('aesthetics_vs_perf', 5), key='aesthetics_vs_perf')
    
    st.divider()
    
    # -------- 9. NUTRICIÓN --------
    st.header("9️⃣ Nutrición")
    col1, col2 = st.columns(2)
    with col1:
        meals_opts = ["2", "3", "4-5", "6+"]
        st.selectbox("Comidas/día", meals_opts, key='meals',
                    index=safe_select_index(meals_opts, st.session_state.get('meals')))
        diet_opts = ["Omnívora", "Vegetariana", "Vegana", "Flexitariana", "Keto", "Otra"]
        st.selectbox("Tipo", diet_opts, key='diet_type',
                    index=safe_select_index(diet_opts, st.session_state.get('diet_type')))
        st.text_input("Alergias", key='allergies')
    with col2:
        prot_opts = ["No sé", "<50g", "50-100g", "100-150g", "150-200g", "200+g"]
        st.selectbox("Proteína actual", prot_opts, key='protein_est',
                    index=safe_select_index(prot_opts, st.session_state.get('protein_est')))
        st.multiselect("Fuentes", ["Carne", "Pollo", "Pescado", "Huevos", "Lácteos", "Legumbres", "Tofu", "Suplementos"], 
                      key='protein_sources', default=st.session_state.get('protein_sources', []))
        st.slider("Relación comida", 1, 5, st.session_state.get('food_relation', 4), key='food_relation')
    
    factor_opts = {
        "Sedentario": 1.2,
        "Ligero (1-3d)": 1.375,
        "Moderado (3-5d)": 1.55,
        "Activo (6-7d)": 1.725,
        "Muy activo": 1.9
    }
    st.selectbox("Actividad total", list(factor_opts.keys()), key='factor_label',
                index=safe_select_index(list(factor_opts.keys()), st.session_state.get('factor_label')))
    
    st.divider()
    
    # -------- 10. SUPLEMENTOS --------
    st.header("🔟 Suplementación")
    st.multiselect("Suplementos", ["Creatina", "Cafeína", "Proteína", "Beta-alanina", "Omega-3", "Vitamina D", "Otros"], 
                  key='supps', default=st.session_state.get('supps', []))
    st.text_area("Detalles", key='supp_details')
    
    st.divider()
    
    # -------- 11. ADHERENCIA --------
    st.header("1️⃣1️⃣ Adherencia")
    col1, col2 = st.columns(2)
    with col1:
        st.slider("Compromiso", 1, 5, st.session_state.get('commitment', 4), key='commitment')
        st.multiselect("Barreras", ["Tiempo", "Dolor", "Equipamiento", "Conocimiento", "Motivación", "Entorno"], 
                      key='barriers', default=st.session_state.get('barriers', []))
    with col2:
        coach_opts = ["Directivo", "Educativo", "Colaborativo"]
        st.selectbox("Coaching", coach_opts, key='coaching_style',
                    index=safe_select_index(coach_opts, st.session_state.get('coaching_style')))
        feed_opts = ["Semanal", "Quincenal", "Mensual"]
        st.selectbox("Seguimiento", feed_opts, key='feedback_freq',
                    index=safe_select_index(feed_opts, st.session_state.get('feedback_freq')))
    
    st.divider()
    
    # -------- ENVIAR --------
    st.markdown("### ¡Listo! 🎉")
    
        
    if st.button("✅ ENVIAR FORMULARIO", type="primary", use_container_width=True):
        errors = []
        
        if not st.session_state.get('name', '').strip():
            errors.append("Nombre")
        if not st.session_state.get('email') or '@' not in st.session_state.get('email', ''):
            errors.append("Email")
        if not st.session_state.get('consent'):
            errors.append("Consentimiento")
        if st.session_state.get('weight', 0) <= 0 or st.session_state.get('height', 0) <= 0:
            errors.append("Peso/Estatura")
        
        if errors:
            st.error(f"❌ Faltan: {', '.join(errors)}")
        else:
            # Guardar datos
            st.session_state['client_data'] = {k: v for k, v in st.session_state.items() 
                                               if not k.startswith('_') and k not in ['FormSubmitter', 'initialized', 
                                                                                       'draft_loaded', 'measures_expanded', 
                                                                                       'bf_expanded', 'client_data', 
                                                                                       'strength_tests', 'form_submitted']}
            st.session_state['form_submitted'] = True
            save_draft()
            
            # NUEVO: Generar y enviar informes por email
            st.info("📧 Generando y enviando informes por email...")
            
            # Obtener datos del cliente
            client_data = st.session_state['client_data']
            
            # Calcular métricas necesarias
            age = None
            if client_data.get('dob'):
                try:
                    age = int((date.today() - client_data['dob']).days / 365.25)
                except:
                    pass
            
            height_m = client_data.get('height', 170) / 100
            weight = client_data.get('weight', 70)
            bmi = weight / (height_m ** 2) if height_m > 0 else None
            
            waist = client_data.get('waist')
            hip = client_data.get('hip')
            whr = (waist / hip) if (waist and hip and hip > 0) else None
            
            # %Grasa Navy
            bf = None
            neck = client_data.get('neck')
            if client_data.get('sex') == "Hombre":
                if waist and neck and waist > neck:
                    bf = navy_bodyfat_male(waist, neck, client_data.get('height', 170))
            else:
                if waist and hip and neck and (waist + hip) > neck:
                    bf = navy_bodyfat_female(waist, hip, neck, client_data.get('height', 170))
            
            if client_data.get('bf_known_value'):
                bf = client_data.get('bf_known_value')
            
            # BMR y TDEE
            factor_map = {"Sedentario": 1.2, "Ligero (1-3d)": 1.375, "Moderado (3-5d)": 1.55, 
                         "Activo (6-7d)": 1.725, "Muy activo": 1.9}
            factor = factor_map.get(client_data.get('factor_label', 'Moderado (3-5d)'), 1.55)
            
            bmr = None
            tdee_maintenance = None
            if age:
                bmr = mifflin_bmr(client_data.get('sex', 'Hombre'), weight, client_data.get('height', 170), age)
                tdee_maintenance = tdee_from_factor(bmr, factor)
            
            # Macros
            protein_avg = None
            fat_avg = None
            carbs_avg = None
            if tdee_maintenance:
                if client_data.get('sex') == "Mujer":
                    protein_avg = weight * 1.75
                    fat_avg = weight * 0.9
                else:
                    protein_avg = weight * 2.25
                    fat_avg = weight * 0.9
                
                protein_kcal = protein_avg * 4
                fat_kcal = fat_avg * 9
                carbs_avg = max(0, (tdee_maintenance - protein_kcal - fat_kcal) / 4)
            
            try:
                # Generar PDF
                with st.spinner("Generando PDF..."):
                    pdf_buffer = generate_complete_professional_pdf(
                        client_data=client_data,
                        age=age,
                        bmi=bmi,
                        whr=whr,
                        bf=bf,
                        bmr=bmr,
                        tdee_maintenance=tdee_maintenance,
                        protein_avg=protein_avg,
                        fat_avg=fat_avg,
                        carbs_avg=carbs_avg,
                        alerts=[],
                        strength_tests_data=st.session_state.get('strength_tests', {})
                    )
                
                # Generar Excel
                with st.spinner("Generando Excel..."):
                    excel_buffer = generate_complete_excel(
                        client_data=client_data,
                        age=age,
                        bmi=bmi,
                        whr=whr,
                        bf=bf,
                        bmr=bmr,
                        tdee_maintenance=tdee_maintenance,
                        protein_avg=protein_avg,
                        fat_avg=fat_avg,
                        carbs_avg=carbs_avg,
                        alerts=[],
                        strength_tests_data=st.session_state.get('strength_tests', {})
                    )
                
                # CONFIGURACIÓN DE EMAIL (CAMBIA ESTOS VALORES)
                PROFESSIONAL_EMAIL = "davidlopeztrain@gmail.com"  
                SENDER_EMAIL = "davidlopeztrain@gmail.com" 
                SENDER_PASSWORD = "lekc ilrg rqty xclh "
                
                # Enviar email
                with st.spinner("Enviando por email..."):
                    success, message = send_evaluation_email(
                        client_name=client_data.get('name', 'Cliente'),
                        client_email=client_data.get('email', ''),
                        professional_email=PROFESSIONAL_EMAIL,
                        pdf_buffer=pdf_buffer,
                        excel_buffer=excel_buffer,
                        sender_email=SENDER_EMAIL,
                        sender_password=SENDER_PASSWORD
                    )
                
                if success:
                    st.success("✅ Formulario enviado y informes enviados por email correctamente")
                    st.info(f"📧 Se han enviado los informes a:\n- **Profesional:** {PROFESSIONAL_EMAIL}\n- **Cliente:** {client_data.get('email', '')}")
                
                else:
                    st.warning(f"⚠️ Formulario guardado pero error al enviar email: {message}")
                    st.info("Los informes se pueden descargar desde el Panel Profesional")
            
            except Exception as e:
                st.error(f"❌ Error al generar/enviar informes: {e}")
                st.info("El formulario se guardó correctamente. Puedes generar los informes desde el Panel Profesional.")


# ============================================
# TAB 2: PANEL PROFESIONAL (COMPLETO Y CORREGIDO)
# ============================================

with tab2:
    st.title("Panel Profesional 📊")
    
    if not st.session_state.get('form_submitted'):
        st.info("ℹ️ Completa el formulario primero")
    else:
        st.success("✅ Datos disponibles")
        
        client_data = st.session_state.get('client_data', {})
        
        if not client_data:
            st.warning("⚠️ No hay datos disponibles")
        else:   
            # Calcular edad
            age = None
            if client_data.get('dob'):
                try:
                    age = int((date.today() - client_data['dob']).days / 365.25)
                except:
                    pass
            
            # Obtener factor numérico
            factor_map = {
                "Sedentario": 1.2,
                "Ligero (1-3d)": 1.375,
                "Moderado (3-5d)": 1.55,
                "Activo (6-7d)": 1.725,
                "Muy activo": 1.9
            }
            factor = factor_map.get(client_data.get('factor_label', 'Moderado (3-5d)'), 1.55)
            
            # ============================================
            # SECCIÓN 1: RESUMEN DEL CLIENTE
            # ============================================
            
            st.header("👤 Resumen del Cliente")
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Nombre", client_data.get('name', 'N/A'))
                st.metric("Edad", f"{age} años" if age else "—")
            with col2:
                st.metric("Sexo", client_data.get('sex', 'N/A'))
                st.metric("Email", client_data.get('email', 'N/A'))
            with col3:
                st.metric("Experiencia", client_data.get('exp_years', 'N/A'))
                st.metric("Continuidad", client_data.get('continuity', 'N/A'))
            with col4:
                main_goal = client_data.get('main_goal', 'N/A')
                st.metric("Objetivo Principal", main_goal.split('(')[0].strip() if main_goal != 'N/A' else 'N/A')
                st.metric("Plazo", client_data.get('horizon', 'N/A'))
            
            st.divider()
            
            # ============================================
            # SECCIÓN 2: ANÁLISIS ANTROPOMÉTRICO
            # ============================================
            
            st.header("📏 Análisis Antropométrico")
            
            # Cálculos base
            height = client_data.get('height', 170)
            weight = client_data.get('weight', 70)
            height_m = height / 100
            bmi = weight / (height_m ** 2) if height_m > 0 else None
            
            waist = client_data.get('waist')
            hip = client_data.get('hip')
            whr = None
            if waist and hip and hip > 0:
                whr = waist / hip
            
            # %Grasa Navy
            bf = None
            neck = client_data.get('neck')
            if client_data.get('sex') == "Hombre":
                if waist and neck and height and waist > neck:
                    bf = navy_bodyfat_male(waist, neck, height)
            else:
                if waist and hip and neck and height and (waist + hip) > neck:
                    bf = navy_bodyfat_female(waist, hip, neck, height)
            
            if client_data.get('bf_known_value'):
                bf = client_data.get('bf_known_value')
            
            # BMR y TDEE
            bmr = None
            tdee_maintenance = None
            if age:
                bmr = mifflin_bmr(client_data.get('sex', 'Hombre'), weight, height, age)
                tdee_maintenance = tdee_from_factor(bmr, factor)
            
            # Mostrar métricas
            col1, col2, col3, col4, col5 = st.columns(5)
            with col1:
                st.metric("Peso", f"{weight:.1f} kg")
            with col2:
                st.metric("Estatura", f"{height:.0f} cm")
            with col3:
                st.metric("IMC", f"{bmi:.1f}" if bmi else "—")
            with col4:
                st.metric("WHR", f"{whr:.2f}" if whr else "No disponible")
            with col5:
                st.metric("% Grasa", f"{bf:.1f}%" if bf else "No disponible")
            
            st.divider()
            
            # ============================================
            # SECCIÓN 3: GASTO ENERGÉTICO Y MACROS
            # ============================================
            
            st.header("🔥 Gasto Energético y Distribución de Macronutrientes")
            
            # Mostrar BMR y TDEE de mantenimiento
            col1, col2 = st.columns(2)
            with col1:
                st.metric("BMR (Mifflin-St Jeor)", f"{bmr:.0f} kcal/día" if bmr else "—")
            with col2:
                st.metric("🎯 TDEE Mantenimiento", f"{tdee_maintenance:.0f} kcal/día" if tdee_maintenance else "—")
            
            st.markdown("---")
            
            # CALCULAR MACROS SEGÚN EVIDENCIA (POR SEXO)
            if tdee_maintenance:
                st.subheader("📊 Distribución de Macronutrientes (Mantenimiento)")
                
                # Rangos según sexo
                if client_data.get('sex') == "Mujer":
                    protein_min = weight * 1.6
                    protein_max = weight * 1.9
                    protein_avg = (protein_min + protein_max) / 2
                    fat_min = weight * 0.8
                    fat_max = weight * 1.0
                    fat_avg = (fat_min + fat_max) / 2
                else:  # Hombre
                    protein_min = weight * 2.0
                    protein_max = weight * 2.5
                    protein_avg = (protein_min + protein_max) / 2
                    fat_min = weight * 0.8
                    fat_max = weight * 1.0
                    fat_avg = (fat_min + fat_max) / 2
                
                # Carbohidratos: resto hasta completar calorías
                protein_kcal = protein_avg * 4
                fat_kcal = fat_avg * 9
                carbs_kcal = tdee_maintenance - (protein_kcal + fat_kcal)
                carbs_avg = carbs_kcal / 4
                
                col1, col2, col3 = st.columns(3)
                with col1:
                    st.metric(
                        "Proteína",
                        f"{protein_avg:.0f} g/día",
                        delta=f"Rango: {protein_min:.0f}-{protein_max:.0f} g"
                    )
                    st.caption(f"{protein_avg/weight:.1f} g/kg")
                with col2:
                    st.metric(
                        "Grasas",
                        f"{fat_avg:.0f} g/día",
                        delta=f"Rango: {fat_min:.0f}-{fat_max:.0f} g"
                    )
                    st.caption(f"{fat_avg/weight:.1f} g/kg")
                with col3:
                    st.metric(
                        "Carbohidratos",
                        f"{carbs_avg:.0f} g/día",
                        delta="Resto hasta completar kcal"
                    )
                
                # Gráfico de distribución
                fig_maintenance = go.Figure(data=[go.Pie(
                    labels=['Proteína', 'Grasas', 'Carbohidratos'],
                    values=[protein_kcal, fat_kcal, carbs_kcal],
                    hole=.4,
                    marker_colors=['#FF6B6B', '#4ECDC4', '#95E1D3']
                )])
                fig_maintenance.update_layout(
                    title="Distribución Calórica - Mantenimiento",
                    showlegend=True,
                    height=350,
                    margin=dict(t=40, b=0, l=0, r=0)
                )
                st.plotly_chart(fig_maintenance, use_container_width=True)
                
                st.markdown("---")
                
                # ============================================
                # SUPERÁVIT Y DÉFICIT
                # ============================================
                
                st.subheader("📈 Escenarios: Superávit y Déficit Calórico")
                
                # SUPERÁVIT
                st.markdown("#### 🔼 Superávit Calórico (Ganancia Muscular)")
                
                surplus_conservative = tdee_maintenance * 1.075
                surplus_aggressive = tdee_maintenance * 1.175
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric(
                        "Superávit Conservador",
                        f"{surplus_conservative:.0f} kcal/día",
                        delta=f"+{surplus_conservative - tdee_maintenance:.0f} kcal (+5-10%)"
                    )
                    carbs_surplus_cons = ((surplus_conservative - (protein_kcal + fat_kcal)) / 4)
                    st.caption(f"Proteína: {protein_avg:.0f}g | Grasas: {fat_avg:.0f}g | Carbs: {carbs_surplus_cons:.0f}g")
                
                with col2:
                    st.metric(
                        "Superávit Agresivo",
                        f"{surplus_aggressive:.0f} kcal/día",
                        delta=f"+{surplus_aggressive - tdee_maintenance:.0f} kcal (+15-20%)"
                    )
                    carbs_surplus_aggr = ((surplus_aggressive - (protein_kcal + fat_kcal)) / 4)
                    st.caption(f"Proteína: {protein_avg:.0f}g | Grasas: {fat_avg:.0f}g | Carbs: {carbs_surplus_aggr:.0f}g")
                
                st.markdown("---")
                
                # DÉFICIT
                st.markdown("#### 🔽 Déficit Calórico (Pérdida de Grasa)")
                
                deficit_conservative = tdee_maintenance * 0.925
                deficit_aggressive = tdee_maintenance * 0.825
                
                # Para déficit, ajustar proteína al alza
                if client_data.get('sex') == "Mujer":
                    protein_deficit = weight * 1.9
                else:
                    protein_deficit = weight * 2.5
                
                protein_deficit_kcal = protein_deficit * 4
                
                col1, col2 = st.columns(2)
                with col1:
                    st.metric(
                        "Déficit Conservador",
                        f"{deficit_conservative:.0f} kcal/día",
                        delta=f"{deficit_conservative - tdee_maintenance:.0f} kcal (-5-10%)"
                    )
                    carbs_deficit_cons = ((deficit_conservative - (protein_deficit_kcal + fat_kcal)) / 4)
                    st.caption(f"Proteína: {protein_deficit:.0f}g | Grasas: {fat_avg:.0f}g | Carbs: {max(0, carbs_deficit_cons):.0f}g")
                
                with col2:
                    st.metric(
                        "Déficit Agresivo",
                        f"{deficit_aggressive:.0f} kcal/día",
                        delta=f"{deficit_aggressive - tdee_maintenance:.0f} kcal (-15-20%)"
                    )
                    carbs_deficit_aggr = ((deficit_aggressive - (protein_deficit_kcal + fat_kcal)) / 4)
                    st.caption(f"Proteína: {protein_deficit:.0f}g | Grasas: {fat_avg:.0f}g | Carbs: {max(0, carbs_deficit_aggr):.0f}g")
                
                st.info("💡 **Nota:** En déficit calórico, la proteína se aumenta al extremo superior del rango para preservar masa muscular.")
            
            st.divider()
            
            # ============================================
            # SECCIÓN 4: TESTS DE FUERZA
            # ============================================
            
            st.header("💪 Tests de Fuerza Submáximos")
            st.markdown("Registra los tests de fuerza del cliente para estimar 1RM y establecer zonas de entrenamiento.")
            
            st.subheader("Registro de Tests")
            
            lift_names = ["Sentadilla", "Peso muerto", "Press banca", "Press militar", "Remo barra", "Dominadas"]
            
            for lift_name in lift_names:
                with st.expander(f"📊 {lift_name}", expanded=False):
                    col1, col2, col3, col4 = st.columns([3, 2, 2, 1])
                    
                    with col1:
                        register = st.checkbox(f"Registrar {lift_name}", key=f"register_{lift_name}")
                    
                    if register:
                        with col2:
                            carga = st.number_input(
                                "Carga (kg)",
                                min_value=0.0,
                                max_value=500.0,
                                value=50.0,
                                step=2.5,
                                key=f"carga_{lift_name}"
                            )
                        with col3:
                            reps = st.number_input(
                                "Repeticiones",
                                min_value=1,
                                max_value=20,
                                value=5,
                                step=1,
                                key=f"reps_{lift_name}"
                            )
                        with col4:
                            rpe = st.number_input(
                                "RPE",
                                min_value=6.0,
                                max_value=10.0,
                                value=8.0,
                                step=0.5,
                                key=f"rpe_{lift_name}"
                            )
                        
                        epl = epley_1rm(carga, reps)
                        brz = brzycki_1rm(carga, reps)
                        avg = (epl + brz) / 2.0 if (epl and brz) else (epl or brz)
                        
                        if avg:
                            st.success(f"✅ 1RM estimado: **{avg:.1f} kg** (Epley: {epl:.1f} kg | Brzycki: {brz:.1f} kg)")
                            
                            st.session_state.strength_tests[lift_name] = {
                                'carga': carga,
                                'reps': reps,
                                'rpe': rpe,
                                'Epley': epl,
                                'Brzycki': brz,
                                'Promedio': avg
                            }
            
            if st.button("💾 Guardar Todos los Tests en Base de Datos", type="primary", use_container_width=True):
                if st.session_state.strength_tests:
                    try:
                        save_strength_tests(client_data.get('name', 'cliente'), client_data.get('email', ''), st.session_state.strength_tests)
                        st.success(f"✅ {len(st.session_state.strength_tests)} test(s) guardado(s)")
                    except Exception as e:
                        st.error(f"❌ Error: {e}")
                else:
                    st.warning("⚠️ No hay tests registrados")
            
            if st.session_state.strength_tests:
                st.markdown("---")
                st.subheader("📋 Resumen de Tests Registrados")
                
                rm_table = []
                for lift, vals in st.session_state.strength_tests.items():
                    rm_table.append({
                        "Ejercicio": lift,
                        "Carga (kg)": vals['carga'],
                        "Reps": vals['reps'],
                        "RPE": vals['rpe'],
                        "1RM Epley (kg)": f"{vals['Epley']:.1f}",
                        "1RM Brzycki (kg)": f"{vals['Brzycki']:.1f}",
                        "1RM Promedio (kg)": f"{vals['Promedio']:.1f}"
                    })
                
                df_rm = pd.DataFrame(rm_table)
                st.dataframe(df_rm, use_container_width=True, hide_index=True)
                
                fig_rm = go.Figure(data=[
                    go.Bar(
                        x=[d['Ejercicio'] for d in rm_table],
                        y=[float(d['1RM Promedio (kg)']) for d in rm_table],
                        marker_color='#667EEA',
                        text=[f"{d['1RM Promedio (kg)']} kg" for d in rm_table],
                        textposition='outside'
                    )
                ])
                fig_rm.update_layout(
                    title="1RM Estimado por Ejercicio",
                    xaxis_title="Ejercicio",
                    yaxis_title="1RM (kg)",
                    height=400,
                    showlegend=False
                )
                st.plotly_chart(fig_rm, use_container_width=True)
                
                st.subheader("🎯 Zonas de Entrenamiento (% del 1RM)")
                zones_table = []
                for lift, vals in st.session_state.strength_tests.items():
                    rm = vals['Promedio']
                    zones_table.append({
                        "Ejercicio": lift,
                        "Fuerza (85-95%)": f"{rm*0.85:.0f}-{rm*0.95:.0f} kg",
                        "Hipertrofia (65-85%)": f"{rm*0.65:.0f}-{rm*0.85:.0f} kg",
                        "Resistencia (50-65%)": f"{rm*0.50:.0f}-{rm*0.65:.0f} kg"
                    })
                
                df_zones = pd.DataFrame(zones_table)
                st.dataframe(df_zones, use_container_width=True, hide_index=True)
            
            st.divider()
            
            # ============================================
            # SECCIÓN 5: ALERTAS
            # ============================================
            
            st.header("⚠️ Alertas y Recomendaciones Clínicas")
            
            alerts = []
            
            if whr:
                if client_data.get('sex') == "Hombre" and whr >= 0.93:
                    alerts.append("🔴 WHR ≥0.93: Riesgo cardiovascular ELEVADO. Priorizar pérdida de grasa abdominal.")
                elif client_data.get('sex') == "Mujer" and whr >= 0.85:
                    alerts.append("🔴 WHR ≥0.85: Riesgo cardiovascular ELEVADO. Priorizar pérdida de grasa abdominal.")
            
            if bmi:
                if bmi < 18.5:
                    alerts.append("🟡 IMC <18.5: Posible bajo peso. Considerar superávit calórico y evaluación médica.")
                elif bmi >= 30:
                    alerts.append("🟡 IMC ≥30: Obesidad. Enfoque gradual con déficit moderado y supervisión médica recomendada.")
            
            if client_data.get('sex') == "Mujer" and "Amenorrea" in str(client_data.get('cycle', '')):
                alerts.append("🔴 AMENORREA detectada: OBLIGATORIO evaluación médica antes de déficit calórico. Riesgo RED-S.")
            
            if client_data.get('sex') == "Hombre" and client_data.get('reds_symptoms', 0) >= 4:
                alerts.append("🔴 Síntomas severos de baja energía/RED-S. Evaluación hormonal y evitar déficit calórico.")
            
            if client_data.get('steps') == "<5000":
                alerts.append("💡 NEAT bajo (<5000 pasos). Aumentar progresivamente a 7000-10000 pasos/día.")
            
            if client_data.get('sleep_hours', 7) < 6.5 or client_data.get('sleep_quality', 5) <= 2:
                alerts.append("💡 Sueño insuficiente. Priorizar 7-9h de sueño para optimizar recuperación.")
            
            if client_data.get('stress', 0) >= 4:
                alerts.append("💡 Estrés elevado. Considerar técnicas de gestión (meditación, reducir volumen).")
            
            if alerts:
                for alert in alerts:
                    if "🔴" in alert:
                        st.error(alert)
                    elif "🟡" in alert:
                        st.warning(alert)
                    else:
                        st.info(alert)
            else:
                st.success("✅ No se detectaron alertas críticas. Cliente en condiciones óptimas para iniciar programa.")
            
            st.divider()
            
            
            # ============================================
            # SECCIÓN 6: EXPORTACIÓN
            # ============================================
            
            st.header("💾 Exportar Evaluación Completa")
            
            col1, col2 = st.columns(2)
            
            with col1:
                # EXCEL COMPLETO (NUEVO)
                if st.button("📊 Generar Excel Completo", use_container_width=True, type="primary"):
                    with st.spinner("Generando Excel profesional..."):
                        try:
                            excel_buffer = generate_complete_excel(
                                client_data=client_data,
                                age=age,
                                bmi=bmi,
                                whr=whr,
                                bf=bf,
                                bmr=bmr,
                                tdee_maintenance=tdee_maintenance,
                                protein_avg=protein_avg if 'protein_avg' in locals() else None,
                                fat_avg=fat_avg if 'fat_avg' in locals() else None,
                                carbs_avg=carbs_avg if 'carbs_avg' in locals() else None,
                                alerts=alerts if 'alerts' in locals() else [],
                                strength_tests_data=st.session_state.get('strength_tests', {})
                            )
                            
                            st.success("✅ Excel generado correctamente")
                            
                            st.download_button(
                                label="⬇️ Descargar Excel Completo",
                                data=excel_buffer,
                                file_name=f"evaluacion_completa_{client_data.get('name', 'cliente').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.xlsx",
                                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                use_container_width=True
                            )
                            
                        except Exception as e:
                            st.error(f"❌ Error: {e}")
                            import traceback
                            st.code(traceback.format_exc())
            
            with col2:
                # PDF PROFESIONAL
                if st.button("📄 Generar Informe PDF", use_container_width=True, type="primary"):
                    with st.spinner("Generando PDF..."):
                        try:
                            pdf_buffer = generate_complete_professional_pdf(
                                client_data=client_data,
                                age=age,
                                bmi=bmi,
                                whr=whr,
                                bf=bf,
                                bmr=bmr,
                                tdee_maintenance=tdee_maintenance,
                                protein_avg=protein_avg if 'protein_avg' in locals() else None,
                                fat_avg=fat_avg if 'fat_avg' in locals() else None,
                                carbs_avg=carbs_avg if 'carbs_avg' in locals() else None,
                                alerts=alerts if 'alerts' in locals() else [],
                                strength_tests_data=st.session_state.get('strength_tests', {})
                            )
                            
                            st.success("✅ PDF generado")
                            
                            st.download_button(
                                label="⬇️ Descargar PDF",
                                data=pdf_buffer,
                                file_name=f"informe_{client_data.get('name', 'cliente').replace(' ', '_')}_{datetime.now().strftime('%Y%m%d')}.pdf",
                                mime="application/pdf",
                                use_container_width=True
                            )
                            
                        except Exception as e:
                            st.error(f"❌ Error: {e}")



st.markdown("---")
st.caption("💪 Sistema de Evaluación Integral v5.1 | Basado en evidencia científica 2024")

